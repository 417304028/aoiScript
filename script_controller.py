import tkinter as tk
from tkinter import ttk
from threading import Thread, current_thread
from loguru import logger
import importlib
from scripts import yjk, lxbj, jbgn, kjj, spc, tccs
from utils import running_event, test_case_status
import ctypes
import inspect
from utils import setup_logger
import sys
import win32event
import win32api
import winerror
import openpyxl
import os
from datetime import datetime

# 创建一个全局的命名互斥体，确保同一时间只能有一个脚本控制器窗口打开
mutex = win32event.CreateMutex(None, False, "Global\\ScriptControllerMutex")
if win32api.GetLastError() == winerror.ERROR_ALREADY_EXISTS:
    logger.error("脚本控制器窗口已经在运行，不能同时打开多个实例。")
    sys.exit(0)

# 定义 Excel 文件路径
EXCEL_FILE_PATH = "test_results.xlsx"

# 定义列名
EXCEL_COLUMNS = [
    "设备类型", "执行系统", "用例模块", "用例编号", "执行操作",
    "执行结果", "预期结果", "执行时间", "备注"
]

def create_or_read_excel():
    # 如果 Excel 文件不存在，则创建一个新的文件
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = openpyxl.Workbook()
        wb.save(EXCEL_FILE_PATH)

def update_excel(module_name, method_name, operation, result, error_step, execution_time, remarks):
    # 读取现有的 Excel 文件
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    if module_name not in wb.sheetnames:
        wb.create_sheet(module_name)
        sheet = wb[module_name]
        sheet.append(EXCEL_COLUMNS)
    else:
        sheet = wb[module_name]
    
    sheet.append([
        "AOI",
        "",  # 先不管
        module_name,
        method_name,
        operation,
        result,
        "",  # 先不管
        execution_time,
        remarks
    ])
    wb.save(EXCEL_FILE_PATH)

def screenshot_error_to_excel(method_name, error_content, test_case_status):
    # 捕捉错误并更新 Excel
    execution_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    update_excel(
        module_name=self.case_combobox.get(),
        method_name=method_name,
        operation=error_content,
        result=test_case_status,
        error_step="",  # 根据需要填写
        execution_time=execution_time,
        remarks=""
    )

# 在程序启动时调用
create_or_read_excel()

class HomeScreen(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.selected_icon = tk.StringVar(value="AOI")

        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 12), padding=10)
        style.configure("Selected.TButton", background="darkblue", foreground="white", borderwidth=2, relief="solid")
        style.configure("Unselected.TButton", background="lightgray", foreground="black", borderwidth=2, relief="ridge")

        self.label = tk.Label(self, text="自动化测试", font=("Segoe UI", 18))
        self.label.pack(pady=20)

        self.icon_frame = tk.Frame(self)
        self.icon_frame.pack(pady=20)

        self.aoi_icon = ttk.Button(
            self.icon_frame, text="AOI", command=self.toggle_aoi, style="Selected.TButton"
        )
        self.aoi_icon.pack(padx=10, pady=10)

        self.button_frame = tk.Frame(self)
        self.button_frame.pack(pady=20)

        self.confirm_button = ttk.Button(self.button_frame, text="确认", command=self.open_aoi_detail)
        self.confirm_button.pack(side="left", padx=10)

        self.exit_button = ttk.Button(self.button_frame, text="退出", command=self.master.quit)
        self.exit_button.pack(side="right", padx=10)

    def toggle_aoi(self):
        if self.selected_icon.get() == "AOI":
            self.selected_icon.set("")
            self.aoi_icon.configure(style="Unselected.TButton")
        else:
            self.selected_icon.set("AOI")
            self.aoi_icon.configure(style="Selected.TButton")

    def open_aoi_detail(self):
        if self.selected_icon.get() == "AOI":
            self.master.switch_frame(AOIDetailScreen)

class AOIDetailScreen(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.excel_path = EXCEL_FILE_PATH
        self.results = self.load_excel()
        self.pack()
        self.create_widgets()

    def load_excel(self):
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            wb.save(self.excel_path)
            return {}
        
        wb = openpyxl.load_workbook(self.excel_path)
        # 删除第一个空的工作表
        if 'Sheet' in wb.sheetnames and not wb['Sheet'].max_row > 1:
            wb.remove(wb['Sheet'])
            wb.save(self.excel_path)

        results = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            results[sheet_name] = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                results[sheet_name].append(row)
        return results

    def save_to_excel(self, data, module_name):
        wb = openpyxl.load_workbook(self.excel_path)
        if module_name not in wb.sheetnames:
            wb.create_sheet(module_name)
            sheet = wb[module_name]
            sheet.append(EXCEL_COLUMNS)
        else:
            sheet = wb[module_name]
        
        sheet.append(data)
        wb.save(self.excel_path)

    def create_widgets(self):
        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 10), padding=5, background="#f0f0f0", foreground="#333", borderwidth=1, relief="solid")
        style.map("TButton",
                  background=[("active", "#e0e0e0"), ("disabled", "#f0f0f0")],
                  foreground=[("active", "#000"), ("disabled", "#999")])

        self.system_label = ttk.Label(self, text="系统:")
        self.system_label.grid(row=0, column=0, padx=5, pady=5, sticky="e")

        self.system_combobox = ttk.Combobox(self, values=["UI", "中间层", "MES", "SPC", "RV"], width=10)
        self.system_combobox.current(0)
        self.system_combobox.grid(row=0, column=1, padx=5, pady=5)

        self.case_label = ttk.Label(self, text="用例:")
        self.case_label.grid(row=0, column=2, padx=5, pady=5, sticky="e")

        self.modules = {
            "离线编辑": lxbj,
            "元件库": yjk,
            "基本功能": jbgn,
            "快捷键": kjj,
            "SPC": spc
        }

        self.case_combobox = ttk.Combobox(self, values=list(self.modules.keys()), width=10)
        self.case_combobox.bind("<<ComboboxSelected>>", self.update_table)
        self.case_combobox.bind("<<ComboboxSelected>>", self.insert_methods_to_excel)
        self.case_combobox.grid(row=0, column=3, padx=5, pady=5)

        self.search_frame = tk.Frame(self)
        self.search_frame.grid(row=1, column=0, columnspan=4, pady=10, padx=20, sticky="w")

        self.search_entry = ttk.Entry(self.search_frame, width=20, foreground="gray")
        self.search_entry.insert(0, "输入用例编码搜索")
        self.search_entry.bind("<FocusIn>", self.clear_placeholder)
        self.search_entry.bind("<FocusOut>", self.add_placeholder)
        self.search_entry.pack(side="left", fill="x", expand=True)

        self.search_button = ttk.Button(self.search_frame, text="搜索", style="TButton", width=5, command=self.search_methods)
        self.search_button.pack(side="left", fill="y", padx=0)

        self.start_button = ttk.Button(self.search_frame, text="开始执行", style="TButton", width=10)
        self.start_button.pack(side="left", padx=5)

        self.stop_button = ttk.Button(self.search_frame, text="停止执行", style="TButton", width=10)
        self.stop_button.pack(side="left", padx=5)

        self.table = ttk.Treeview(self, columns=("编码", "状态", "操作"), show="headings", height=10)
        self.table.heading("编码", text="编码")
        self.table.heading("状态", text="状态")
        self.table.heading("操作", text="操作")
        self.table.grid(row=2, column=0, columnspan=6, padx=5, pady=5, sticky="nsew")

        self.table.column("编码", anchor="center", width=100)
        self.table.column("状态", anchor="center", width=100)
        self.table.column("操作", anchor="center", width=200)

        self.more_label = tk.Label(self, text="点击查询更多...", fg="blue", cursor="hand2")
        self.more_label.grid(row=3, column=0, columnspan=6, pady=10)
        self.more_label.bind("<Button-1>", self.more_info)

        # 加载所有模块的方法到Excel
        self.load_all_modules_to_excel()

    def load_all_modules_to_excel(self):
        for module_name, module in self.modules.items():
            methods = [func for func in dir(module) if callable(getattr(module, func)) and not func.startswith("__")]
            for method in methods:
                update_excel(module_name, method, "", "未执行", "", "", "")

    def update_table(self, event=None):
        # 更新表格时读取 Excel
        module_name = self.case_combobox.get()
        if module_name in self.results:
            self.table.delete(*self.table.get_children())
            for result in self.results[module_name]:
                self.table.insert("", "end", values=(result[3], result[5], "操作"))

    def insert_methods_to_excel(self, event=None):
        module_name = self.case_combobox.get()
        module = self.modules.get(module_name)
        if module:
            methods = [func for func in dir(module) if callable(getattr(module, func)) and not func.startswith("__")]
            for method in methods:
                update_excel(module_name, method, "", "未执行", "", "", "")

    def add_table_row(self, method_name):
        item = self.table.insert("", "end", values=(method_name, "未执行"))
        self.add_buttons_to_row(item, method_name)

    def add_buttons_to_row(self, item, method_name):
        # 获取Treeview的bbox来确定按钮位置
        bbox = self.table.bbox(item, column="操作")
        if not bbox:
            return

        # 创建一个Canvas来放置按钮
        canvas = tk.Canvas(self.table, width=bbox[2], height=bbox[3])
        canvas.place(x=bbox[0], y=bbox[1])

        # 创建按钮，设置样式和大小
        execute_button = tk.Button(
            canvas, text="执行这条", command=lambda: self.execute_method(method_name),
            height=1, width=8, font=("Segoe UI", 9)
        )
        execute_down_button = tk.Button(
            canvas, text="这条往下", command=lambda: self.execute_down(method_name),
            height=1, width=8, font=("Segoe UI", 9)
        )

        # 布局按钮
        execute_button.pack(side="left", padx=5)
        execute_down_button.pack(side="left", padx=5)

    def execute_method(self, method_name):
        # 执行用例逻辑
        result = "成功"  # 假设执行成功
        error_step = ""  # 假设没有错误步骤
        # 如果有错误，捕获并记录
        try:
            # 执行用例
            pass
        except Exception as e:
            result = "失败"
            error_step = str(e)

        # 保存结果到 Excel
        data = ["AOI", "", "", method_name, error_step, result, "", "", ""]
        self.save_to_excel(data, self.case_combobox.get())
        print(f"执行方法: {method_name}")

    def execute_down(self, method_name):
        print(f"执行从方法: {method_name} 开始的所有方法")

    def search_methods(self):
        search_text = self.search_entry.get().strip()
        for item in self.table.get_children():
            method_name = self.table.item(item, "values")[0]
            if search_text.lower() in method_name.lower() or not search_text:
                self.table.reattach(item, '', 'end')
            else:
                self.table.detach(item)

    def clear_placeholder(self, event):
        if self.search_entry.get() == "输入用例编码搜索":
            self.search_entry.delete(0, tk.END)
            self.search_entry.configure(foreground="black")

    def add_placeholder(self, event):
        if not self.search_entry.get():
            self.search_entry.insert(0, "输入用例编码搜索")
            self.search_entry.configure(foreground="gray")

    def more_info(self, event):
        # TODO: Add functionality for more info
        pass

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("自动化测试")
        self.geometry("600x600")  # 调整窗口大小
        self.center_window()
        self.current_frame = None
        self.switch_frame(HomeScreen)  # 先进入HomeScreen

    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def switch_frame(self, frame_class):
        new_frame = frame_class(self)
        if self.current_frame is not None:
            self.current_frame.destroy()
        self.current_frame = new_frame
        self.current_frame.pack()
        if isinstance(new_frame, AOIDetailScreen):
            self.geometry("600x600")  # Detail界面大小
        else:
            self.geometry("400x400")  # Home界面大小

if __name__ == "__main__":
    setup_logger()
    app = Application()
    app.mainloop()