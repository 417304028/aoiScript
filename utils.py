import ctypes
import random
import re
import shutil
import subprocess
import sys
# sys.coinit_flags = 2
import time
import psutil
import pyperclip
import cv2
import numpy as np
import win32gui
import pyautogui
import config
import functools
import datetime
import threading
import win32con
import win32com.client
import os
import tkinter as tk
from threading import Event
from pywinauto import Application
from PIL import ImageGrab
from PIL import Image as PILImage
from PIL import ImageChops
from paddleocr import PaddleOCR
from screeninfo import get_monitors
from loguru import logger
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from sklearn.cluster import KMeans
import win32api
import win32process
pyautogui.FAILSAFE = False

# ctypes.windll.shcore.SetProcessDpiAwareness(0)  # 解决使用pyautowin时缩放问题
running_event = Event()


# ============================日志=======================
# 设置日志记录器
def setup_logger():
    log_dir = os.path.join(os.getcwd(), "log")
    print("Log Directory:", log_dir)
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, "script_log_{time:YYYY-MM-DD}.log")
    logger.add(log_file, rotation="100 MB", retention="10 days",
               format="{time:YYYY-MM-DD HH:mm:ss.SSS} {level} {message}", level="INFO")
    logger.info("已开启日志记录")

try:
    logger.info("paddleocr 和 PaddleOCR 已成功导入。")
    # 确定路径
    dev_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'models')
    if os.path.exists(dev_path):
        # 开发环境
        logger.info("使用开发环境")
        internal_path = dev_path
    elif getattr(sys, 'frozen', True):
        # 打包后的环境
        logger.info("使用打包后环境")
        internal_path = os.path.join(os.path.abspath(os.path.dirname(sys.executable)), 'models')

    # 检查路径中 '_internal' 的出现次数
    internal_count = internal_path.count('_internal')
    logger.debug(f"最终 internal_path: {internal_path}")

    if os.path.exists(internal_path):
        # 使用 _internal 目录下的模型
        det_model_dir = os.path.join(internal_path, 'ch_PP-OCRv4_det_infer')
        rec_model_dir = os.path.join(internal_path, 'ch_PP-OCRv4_rec_infer')
        cls_model_dir = os.path.join(internal_path, 'ch_ppocr_mobile_v2.0_cls_infer')
    else:
        logger.error("模型目录不存在，使用绝对路径（仅用于开发环境）")
        det_model_dir = 'D:\\work\\models\\ch_PP-OCRv4_det_infer'
        rec_model_dir = 'D:\\work\\models\\ch_PP-OCRv4_rec_infer'
        cls_model_dir = 'D:\\work\\models\\ch_ppocr_mobile_v2.0_cls_infer'

    logger.debug(f"det_model_dir: {det_model_dir}")
    logger.debug(f"rec_model_dir: {rec_model_dir}")
    logger.debug(f"cls_model_dir: {cls_model_dir}")

    # 初始化 PaddleOCR
    paddleOCR = PaddleOCR(
        use_angle_cls=True,
        lang='ch',
        use_gpu=False,
        show_log=True,
        det_model_dir=det_model_dir,
        rec_model_dir=rec_model_dir,
        cls_model_dir=cls_model_dir
    )
except ImportError as e:
    logger.error(f"导入错误: {e}")

# ============================窗格处理===================
# 连接窗口（好像只能用win32连接窗口,uia不行）
def connect_aoi_window():
    try:
        # 使用 uia 后端进行连接
        app = Application(backend="uia").connect(title_re="Sinic-Tek 3D AOI", auto_id="MainForm")
        main_window = app.window(auto_id="MainForm")
        if main_window.exists(timeout=20):  # 增加超时时间
            logger.info("成功连接到窗口")
            return main_window
        else:
            error_message = "未找到窗口"
            logger.error(error_message)
            raise Exception(error_message)
    except Exception as e:
        error_message = f"连接窗口时发生错误: {e}"
        logger.error(error_message)
        raise Exception(error_message)
def window_interactive(title_re=None, auto_id=None, class_name=None, control_type="Button", name=None, action="click", text=None):
    try:
        logger.info(f"传入参数 - title_re: {title_re}, auto_id: {auto_id}, class_name: {class_name}, control_type: {control_type}, name: {name}, action: {action}, text: {text}")
        main_window = connect_aoi_window()
        logger.info("成功连接到主窗口")

        # 检查并组合存在的参数
        search_criteria = {}
        if title_re:
            search_criteria['title_re'] = f".*{title_re}.*"
        if auto_id:
            search_criteria['auto_id'] = auto_id
        if class_name:
            search_criteria['class_name'] = class_name
        if control_type:
            search_criteria['control_type'] = control_type
        if name:
            search_criteria['title'] = name
        
        if not search_criteria:
            raise ValueError("必须提供至少一个参数用于筛选控件")
        
        logger.info(f"使用筛选条件: {search_criteria} 查找目标控件")
        try:
            target_control = main_window.child_window(**search_criteria).wait('exists', timeout=10)
            logger.info(f"找到目标控件: {target_control}")
        except Exception as e:
            logger.error(f"查找控件时发生错误: {e}")
            raise Exception(f"查找控件时发生错误: {e}")
        
        # 检查控件是否存在
        if not target_control.exists():
            raise Exception("未找到符合条件的控件")
        
        target_control = target_control.wrapper_object()
        
        logger.info(f"目标控件信息: {target_control}")
        logger.info("找到目标控件，检查其可见性和启用状态")
        if not (target_control.is_visible() and target_control.is_enabled()):
            raise Exception("目标控件不可见或不可用")
        
        logger.info("目标控件可见且可用，准备执行操作")
        if action == "click":
            target_control.click()
            logger.info("成功点击目标控件")
        elif action == "set_text" and text is not None:
            target_control.set_text(text)
            logger.info(f"成功设置文本: {text}")
        else:
            raise ValueError("无效的操作或缺少文本内容")
    except Exception as e:
        error_message = f"操作目标控件时发生错误: {e}"
        logger.error(error_message)
        raise Exception(error_message)
# 处理登录时的一系列预处理（有可能为打开程式的界面）
def login_process():
    logger.warning("开始登录预处理")
    minimize_service_process_manager()
    if not search_symbol(config.LOGINING, 2, tolerance= 0.5) and not search_symbol(config.AOI_TOPIC,3,tolerance=0.5):
        minimize_service_process_manager()
        time.sleep(3)
    bring_window_to_foreground("AOI.exe")
    delete_documents(config.SHARE_LIB_PATH)
    delete_documents(config.ELEMENTS_LIB_PATH)
    if search_symbol(config.LOGIN_ENGLISH,2):
        logger.info("检测到aoi英文登录界面")
        pyautogui.write("sinictek")
        pyautogui.press("enter")
        time.sleep(1)
    if search_symbol(config.RV_PASSWORD,2,tolerance=0.7):
        time.sleep(1)
        logger.info("输入密码")
        pyautogui.write("000")
        pyautogui.press("enter")
        time.sleep(1)
        pyautogui.press("enter")
        time.sleep(2)
    click_by_ocr("取消",timeout=3)
    if search_symbol(config.NO,1):
        click_by_png(config.NO)
    while search_symbol(config.PROGRAM_LOADING,3):
        time.sleep(3)
    try:
        # 获取屏幕分辨率
        user32 = ctypes.windll.user32
        current_width = user32.GetSystemMetrics(0)
        current_height = user32.GetSystemMetrics(1)
        
        # 如果当前分辨率不是1920x1080，则进行修改
        if current_width != 1920 or current_height != 1080:
            logger.debug(f"当前分辨率为{current_width}x{current_height}，正在修改为1920x1080...")
            user32.ChangeDisplaySettingsW(None, 0)
            devmode = ctypes.create_string_buffer(148)
            ctypes.memset(devmode, 0, 148)
            ctypes.memmove(devmode, b'\x00' * 148, 148)
            ctypes.memmove(devmode, b'\x00' * 40, 40)
            ctypes.memmove(devmode[40:], b'\x80\x07\x38\x04', 4)
            ctypes.memmove(devmode[44:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[108:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[112:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[116:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[120:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[124:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[128:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[132:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[136:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[140:], b'\x00\x00\x00\x00', 4)
            ctypes.memmove(devmode[144:], b'\x00\x00\x00\x00', 4)
            user32.ChangeDisplaySettingsW(devmode, 0)
            logger.info("分辨率修改成功")
        else:
            logger.info("当前分辨率已为1920x1080，无需修改")
    except Exception as e:
        error_message = f"修改分辨率时发生错误: {e}"
        logger.error(error_message)
        raise Exception(error_message)
    # 这是为了前置窗口的 先保留
    # click_by_png(config.AOI_TOPIC, tolerance=0.65, type="bottom")
    # 确保登录后无弹框
    login_symbols = [
        (config.USER_LOGIN_LIGHT, 0.95),
        (config.USER_LOGIN_DARK, 0.95),
        (config.USER_LOGIN_CHINESE_LIGHT, 0.8),
        (config.USER_LOGIN_CHINESE_DARK, 0.8)
    ]
    minimize_service_process_manager()
    if any(search_symbol(symbol, 1.5, tolerance=tol) for symbol, tol in login_symbols):
        write_text((865,500),"000")
        time.sleep(0.5)
        pyautogui.press('enter')
        if search_symbol(config.INVALID_PASSWORD, 1.5, tolerance=0.95):
            # 关闭密码错误窗口
            pyautogui.press('enter')
            time.sleep(0.5)
            # 输入下个密码
            write_text((865,500),"sinictek")
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(1)
            # 关闭登录成功窗口
            pyautogui.press('enter')
        else:
            # 关闭登录成功窗口
            pyautogui.press('enter')
    # 防止登录aoi后直接就是打开程式界面
    if search_symbol(config.WINDOW_CANCEL, 3, tolerance=0.5):
        click_by_png(config.WINDOW_CANCEL, tolerance=0.5)
    # 关闭打开程式窗口
    click_by_ocr("取消")
    time.sleep(8)
    # 确保登录时没有开在线调参功能
    if search_symbol(config.RUN_DARK, 1):
        click_by_png(config.RUN_DARK, timeout=1.5)
        time.sleep(1)
    if search_symbol(config.CHANGE_PARAM_ONLINE_OPENED, 1):
        click_by_png(config.CHANGE_PARAM_ONLINE_OPENED, timeout=1.5)
        time.sleep(0.5)
    if search_symbol(config.QUESTION_MARK, 3, tolerance=0.75):
        if search_symbol(config.NO,2):
            click_by_png(config.NO,timeout=2)
        else:
            pyautogui.press("enter")
    logger.warning("登录预处理完成")

def bring_window_to_foreground(process_name):
    # 检查进程是否存在
    if not any(proc.info['name'] == process_name for proc in psutil.process_iter(['name'])):
        logger.error(f"未找到{process_name}进程")
        return False

    class Op:
        def __init__(self, pid, name):
            self.pid = pid
            self.name = name

        def setForegroundWindowByWin32GUI(self):
            try:
                # 获取窗口句柄
                hwnds = self.get_obj_hwnd()
                if not hwnds:
                    logger.error(f"未找到窗口句柄: {self.name}")
                    return False
                hwnd = hwnds[0]

                # 获取当前前台窗口信息
                hForeWnd = win32gui.GetForegroundWindow()
                current_thread_id = win32api.GetCurrentThreadId()
                fore_thread_id, _ = win32process.GetWindowThreadProcessId(hForeWnd)

                # 附加线程输入
                win32process.AttachThreadInput(fore_thread_id, current_thread_id, True)
                
                # 激活并最大化窗口
                win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
                result = win32gui.SetForegroundWindow(hwnd)
                logger.info(f"设置窗口前置: {self.name}, API返回结果: {result}")

                # 恢复线程输入状态
                win32process.AttachThreadInput(fore_thread_id, current_thread_id, False)

                return result != 0
            except Exception as e:
                logger.error(f"设置窗口前置时发生错误: {e}")
                return False
            finally:
                # 确保始终恢复线程输入状态
                if 'fore_thread_id' in locals() and 'current_thread_id' in locals():
                    win32process.AttachThreadInput(fore_thread_id, current_thread_id, False)

        def get_obj_hwnd(self):
            def _EnumWindowsCallback(hwnd, hwnds):
                _, pid = win32process.GetWindowThreadProcessId(hwnd)
                if pid == self.pid:
                    # 检查窗口是否可见且没有父窗口
                    if win32gui.IsWindowVisible(hwnd) and not win32gui.GetParent(hwnd):
                        # 进一步检查窗口标题是否匹配（可选）
                        # title = win32gui.GetWindowText(hwnd)
                        # if title:  # 如果有标题要求可以添加过滤条件
                        hwnds.append(hwnd)
            hwnds = []
            win32gui.EnumWindows(_EnumWindowsCallback, hwnds)
            return hwnds

    # 获取指定进程的ID
    pid = None
    try:
        processes = [proc for proc in psutil.process_iter(['pid', 'name']) if proc.info['name'] == process_name]
        if processes:
            # 随机选择一个进程实例（可根据需要调整选择逻辑）
            selected_proc = random.choice(processes)
            pid = selected_proc.info['pid']
    except Exception as e:
        logger.error(f"获取进程信息时出错: {e}")

    if pid:
        op = Op(pid, process_name)
        try:
            success = op.setForegroundWindowByWin32GUI()
        except Exception as e:
            logger.error(f"设置窗口前置时发生错误: {e}")
        # if success:
        #     logger.info(f"成功前置窗口: {process_name}")
        #     return True
        # else:
        #     logger.error("窗口前置失败")
        #     return False
    else:
        logger.error(f"未找到{process_name}进程")
        return False
# 确保rv打开并前置
def check_and_launch_rv():
    close_rv()
    logger.info("将所有窗口最小化")
    pyautogui.hotkey('win', 'd')
    time.sleep(1)
    logger.info("最小化完成...")
    if not bring_window_to_foreground('DVPro.UI.exe'):
        # 保存当前工作目录
        current_dir = os.getcwd()
        # 切换到目标目录并运行程序
        os.chdir(os.path.dirname(config.RV_EXE_PATH))
        subprocess.Popen(f'cmd /c "{os.path.basename(config.RV_EXE_PATH)}"', shell=True)
        # 切换回原来的工作目录
        os.chdir(current_dir)

        process_found = False
        while not process_found:
            for proc in psutil.process_iter(['pid', 'name']):
                if proc.info['name'] == os.path.basename(config.RV_EXE_PATH):
                    logger.info(f"{os.path.basename(config.RV_EXE_PATH)} 已启动")
                    process_found = True
                    break
            if not process_found:
                logger.info(f"等待 {os.path.basename(config.RV_EXE_PATH)} 启动...")
                time.sleep(1)
        time.sleep(5)
        # if search_symbol(config.RV_TASKBAR):
        #     click_by_png(config.RV_TASKBAR, tolerance=0.8)
        # else:
        #     logger.info("检测到DVPro.UI进程，正在将其前置...")
    else:
        logger.info("置顶rv成功")
    # pyautogui.press("enter")
    # 登录
    if search_symbol(config.RV_PASSWORD, 5, tolerance=0.7):
        logger.info("检测到登录界面")
    else:
        logger.info("未检测到登录界面，即将前置窗口")
        bring_window_to_foreground('DVPro.UI.exe')
    write_text_textbox(config.RV_PASSWORD, write_content=config.RV_PASSWORD_TEXT)
    pyautogui.press("enter")
    if not search_symbol(config.RV_ICON,5):
        pyautogui.press("enter")
        write_text_textbox(config.RV_PASSWORD, write_content="sinictek")
        pyautogui.press("enter")
    if not search_symbol(config.RV_ICON,25):
        raise Exception("未发现rv界面")

# 确保spc打开并前置
def check_and_launch_spc():
    close_spc()
    logger.info("将所有窗口最小化")
    pyautogui.hotkey('win', 'd')
    time.sleep(1)
    logger.info("最小化完成...")
    if not bring_window_to_foreground('SPCViewMain.exe'):
        spc_running = any("SPC" in p.name() for p in psutil.process_iter())
        logger.info(f"SPC程序运行状态: {'运行中' if spc_running else '未运行'}")
        if not spc_running:
            logger.info(f"SPC程序未运行,路径: {config.SPC_EXE_PATH}，正在启动...")
            if not os.path.exists(config.SPC_EXE_PATH):
                logger.error(f"SPC程序路径不存在: {config.SPC_EXE_PATH}")
                return
            plugins_dir = os.path.join(os.getcwd(), "PlugIns")
            if not os.path.exists(plugins_dir):
                os.makedirs(plugins_dir)
                logger.info(f"创建文件夹: {plugins_dir}")
            else:
                logger.info(f"文件夹已存在: {plugins_dir}")
            env = os.environ.copy()
            env['LANG'] = 'zh_CN.UTF-8'  # 设置为中文

            # # 方式1: 使用subprocess.run启动
            # try:
            #     subprocess.run([config.SPC_EXE_PATH], check=True, env=env)
            #     logger.info("SPC程序已启动（方式1）")
            # except subprocess.CalledProcessError as e:
            #     logger.error(f"启动SPC程序时发生错误（方式1）: {e}")

            # # 方式2: 使用subprocess.Popen启动
            # try:
            #     process = subprocess.Popen([config.SPC_EXE_PATH], env=env)
            #     process.wait()
            #     logger.info("SPC程序已启动（方式2）")
            # except subprocess.CalledProcessError as e:
            #     logger.error(f"启动SPC程序时发生错误（方式2）: {e}")

            # # 方式3: 使用os.system启动
            # try:
            #     result = os.system(f'"{config.SPC_EXE_PATH}"')
            #     if result == 0:
            #         logger.info("SPC程序已启动（方式3）")
            #     else:
            #         logger.error(f"启动SPC程序时发生错误（方式3）: 返回码 {result}")
            # except Exception as e:
            #     logger.error(f"启动SPC程序时发生错误（方式3）: {e}")
            # logger.info("等待SPC程序启动...")
            # time.sleep(5)

            # # 方式4: 使用ctypes库启动
            # try:
            #     ctypes.windll.shell32.ShellExecuteW(None, 'open', config.SPC_EXE_PATH, None, None, 1)
            #     logger.info("SPC程序已启动（方式4）")
            # except Exception as e:
            #     logger.error(f"启动SPC程序时发生错误（方式4）: {e}")

            # 方式5: 使用命令行运行SPCViewMain.exe
            current_dir = os.getcwd()
            os.chdir(os.path.dirname(config.SPC_EXE_PATH))
            subprocess.Popen(f'cmd /c "{os.path.basename(config.SPC_EXE_PATH)}"', shell=True)
            # 切换回原来的工作目录
            os.chdir(current_dir)
            process_found = False
            while not process_found:
                for proc in psutil.process_iter(['pid', 'name']):
                    if proc.info['name'] == os.path.basename(config.SPC_EXE_PATH):
                        logger.info(f"{os.path.basename(config.SPC_EXE_PATH)} 已启动")
                        process_found = True
                        time.sleep(3)
                        break
                if not process_found:
                    logger.info(f"等待 {os.path.basename(config.SPC_EXE_PATH)} 启动...")
                    time.sleep(3)
            # 中文界面
            def attempt_login():
                logger.info("检测到SPC中文登录界面")
                click_by_png(config.SPC_LOGIN_USERNAME_CHINESE, tolerance=0.7, type="right")
                click_by_png(config.SPC_USER_ADMIN)
                write_text_textbox(config.SPC_LOGIN_PASSWORD_CHINESE, write_content=config.SPC_USER_NAME)
                logger.info(f"输入密码: {config.SPC_USER_NAME}")
                if search_symbol(config.SPC_LOGIN_CHINESE, 5, tolerance=0.6):
                    click_by_png(config.SPC_LOGIN_CHINESE, tolerance=0.6)
                    logger.info("点击登录按钮")
                else:
                    pyautogui.press("enter")
                if not search_symbol(config.SPC_SYSTEM_SETTING, 2, tolerance=0.6):
                    logger.error("登录spc时密码错误 改用000")
                    pyautogui.press("enter")
                    write_text_textbox(config.SPC_LOGIN_PASSWORD_CHINESE, write_content="000")
                    click_by_png(config.SPC_LOGIN_CHINESE, tolerance=0.6)
                time.sleep(5)

            if search_symbol(config.SPC_LOGIN_CHINESE, 5, tolerance=0.6):
                attempt_login()
            else:
                logger.info("未检测到SPC登录界面，尝试将窗口前置")
                bring_window_to_foreground('SPCViewMain.exe')
                # 这里是把登录界面置顶了 但是没执行登录
                if search_symbol(config.SPC_LOGIN_CHINESE, 2, tolerance=0.6):
                    logger.debug("前置后找到登录界面了 开始登录")
                    attempt_login()
            time.sleep(3)
            if not search_symbol(config.SPC_SYSTEM_SETTING, 2):
                bring_window_to_foreground('SPCViewMain.exe')
                if not search_symbol(config.SPC_SYSTEM_SETTING, 2):
                    error_message = "疑似登录失败,未检测到SPC界面"
                    logger.error(error_message)
                    raise Exception(error_message)
            logger.info("SPC登录成功")
            if click_by_ocr("请输入Password", timeout=3):
                pyautogui.press("enter")
    
    # 将SPC应用最大化
    logger.info("尝试最大化SPC应用")
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] == "SPCViewMain.exe":
            pid = proc.pid
            def enum_handler(hwnd, lParam):
                if win32gui.IsWindowVisible(hwnd) and win32gui.IsWindowEnabled(hwnd):
                    _, window_pid = win32process.GetWindowThreadProcessId(hwnd)
                    if window_pid == pid:
                        logger.info(f"最大化窗口句柄: {hwnd}")
                        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
            win32gui.EnumWindows(enum_handler, None)
            break
        
def minimize_service_process_manager():
    """
    最小化 ServiceProcessManager.exe 的窗口
    """
    logger.info("尝试最小化 ServiceProcessManager.exe")
    # 通过 psutil 找到 ServiceProcessManager.exe 进程
    target_proc = None
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] == "ServiceProcessManager.exe":
            target_proc = proc
            break
    if not target_proc:
        logger.error("未找到 ServiceProcessManager.exe 进程")
        return

    pid = target_proc.pid


    def enum_handler(hwnd, lParam):
        # 检查该窗口是否属于目标进程
        if win32gui.IsWindowVisible(hwnd) and win32gui.IsWindowEnabled(hwnd):
            _, window_pid = win32process.GetWindowThreadProcessId(hwnd)
            if window_pid == pid:
                logger.info(f"最小化窗口句柄: {hwnd}")
                win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)

    try:
        win32gui.EnumWindows(enum_handler, None)
        logger.info("ServiceProcessManager.exe 的窗口已最小化")
    except Exception as e:
        logger.error(f"最小化 ServiceProcessManager.exe 时发生错误: {e}")

# 确保aoi打开并前置
def check_and_launch_aoi():
    logger.info("开始检查并启动AOI程序...")
    logger.info("将所有窗口最小化")
    pyautogui.hotkey('win', 'd')
    logger.info("最小化完成...")
    start_time = time.time()  # 开始计时
    for proc in psutil.process_iter(['pid', 'name']):
        if "AOI.exe" in proc.info['name']:
            try:
                minimize_service_process_manager()
                # logger.info(f"检测到AOI进程: {proc.info['name']} (PID: {proc.info['pid']})，准备杀掉...")
                parent_proc = psutil.Process(proc.info['pid'])
                # children = parent_proc.children(recursive=True)
                # for child in children:
                #     logger.info(f"杀掉子进程: {child.name()} (PID: {child.pid})")
                #     child.kill()
                #     child.wait()  # 确认子进程被杀掉
                # parent_proc.kill()
                # parent_proc.wait()  # 确认父进程被杀掉
                # logger.info(f"成功杀掉进程: {proc.info['name']} (PID: {proc.info['pid']})")
                
                # # 再次检查进程是否已被杀掉
                # if not psutil.pid_exists(proc.info['pid']):
                #     logger.info(f"确认进程 {proc.info['name']} (PID: {proc.info['pid']}) 已被成功杀掉")
                # else:
                #     logger.error(f"进行杀掉aoi进程操作之后，进程 {proc.info['name']} (PID: {proc.info['pid']}) 仍在运行")
                #     raise Exception(f"进行杀掉aoi进程操作之后，进程 {proc.info['name']} (PID: {proc.info['pid']}) 仍在运行")
            except Exception as kill_error:
                logger.error(f"无法杀掉进程 {proc.info['name']} (PID: {proc.info['pid']}): {kill_error}")
    elapsed_time = time.time() - start_time  # 计算耗时
    logger.info(f"杀掉AOI进程的操作耗时：{elapsed_time}秒")

    aoi_running = any("AOI.exe" == p.name() for p in psutil.process_iter())
    
    # aoi不存在的话，启动aoi
    if not aoi_running:
        logger.info("AOI程序未运行，正在启动...")
        # app = Application().start(config.AOI_EXE_PATH)
        # 使用命令行启动AOI程序
        current_dir = os.getcwd()
        os.chdir(os.path.dirname(config.AOI_EXE_PATH))
        subprocess.Popen(f'cmd /c "{os.path.basename(config.AOI_EXE_PATH)}"', shell=True)
        # 切换回原来的工作目录
        os.chdir(current_dir)
        logger.info("等待AOI程序启动...")
        minimize_service_process_manager()
        if search_symbol(config.WARNING, 5, tolerance=0.75):
            pyautogui.press("enter")
        timeout = time.time() + 90
        retry_count = 0
        max_retries = 2
        while time.time() < timeout:
            if search_symbol(config.LOGINING, 2, tolerance=0.5) or any("AOI.exe" == p.name() for p in psutil.process_iter()):
                logger.info("检测到登录进程，检查并杀掉'aoi_memory'进程...")
                minimize_service_process_manager()
                for proc in psutil.process_iter():
                    try:
                        if "aoi_memory" in proc.name():
                            logger.info(f"正在杀死进程 {proc.name()}...")
                            proc.kill()
                            proc.wait()
                            logger.info(f"进程 {proc.name()} 已被杀死")
                            break
                    except Exception as e:
                        logger.error(f"无法杀死进程 {proc.name()}: {e}")
                retry_count += 1  
            else:
                minimize_service_process_manager()
            if retry_count == max_retries:
                break
        logger.info("重新搜索加载界面")
        minimize_service_process_manager()
        for _ in range(2):
            if not click_by_ocr("思泰克智能",timeout=3):
                # bring_window_to_foreground("AOI.exe")
                bring_aoi_to_top()
            else:
                if search_symbol(config.WARNING, 2, tolerance=0.7):
                    pyautogui.press("enter")
                    break

    # aoi存在的话 关闭再重启aoi
    else:
        logger.info("再次检测AOI.exe是否存在...")
        time.sleep(3)
        minimize_service_process_manager()
        for proc in psutil.process_iter():
            if "AOI.exe" == proc.name():
                try:
                    logger.info(f"检测到进程 {proc.name()} (PID: {proc.pid})，正在杀死...")
                    # proc.kill()
                    # proc.wait(timeout=5)  # 增加超时时间
                    if not proc.is_running():
                        logger.info(f"进程 {proc.name()} (PID: {proc.pid}) 已被杀死")
                    else:
                        logger.error(f"进程 {proc.name()} (PID: {proc.pid}) 未能被杀死")
                except Exception as e:
                    logger.error(f"无法杀死进程 {proc.name()} (PID: {proc.pid}): {e}")
                break
    while click_by_ocr("思泰克智能",timeout=5):
        time.sleep(3)
    login_process()
    logger.info("AOI程序检查和启动完成。")
def close_aoi():
    try:
        for proc in psutil.process_iter():
            if "AOI.exe" == proc.name():
                logger.info("AOI程序正在运行,正在关闭...")
                proc.kill()
                proc.wait() 
                logger.info("AOI程序已关闭")
                break
    except Exception as e:
        logger.info(f"关闭aoi时出现错误:{e}")
def initialize_aoi():
    bring_window_to_foreground("AOI.exe")
    # 防止有各类弹框
    click_by_ocr("取消",timeout=3)
    # 防止在测试界面
    if search_symbol(config.TESTING_INTERFACE_INFORMATION,2,tolerance=0.6):
        click_by_ocr("停止",timeout=3)
        time.sleep(5)
        click_by_ocr("返回",timeout=3)
    open_program(if_specific=True)
    time.sleep(5)
    # 关闭其他软件
    # close_rv()
    close_spc()

def close_rv():
    try:
        for proc in psutil.process_iter():
            if "DVPro.UI.exe" == proc.name():
                logger.info("AOI程序正在运行,正在关闭...")
                proc.kill()
                proc.wait() 
                logger.info("AOI程序已关闭")
                break
    except Exception as e:
        logger.info(f"关闭rv时出现错误:{e}")
def close_spc():
    try:
        for proc in psutil.process_iter():
            if "SPCViewMain.exe" == proc.name():
                logger.info("SPC程序正在运行,正在关闭...")
                proc.kill()
                proc.wait() 
                logger.info("SPC程序已关闭")
                break
    except Exception as e:
        logger.info(f"关闭spc时出现错误:{e}")
# 确保在特定界面（通过特定标识物的存在）
def ensure_in_specific_window(name=None, auto_id=None, control_type=None):
    try:
        main_window = connect_aoi_window()
        criteria = {}
        if name:
            criteria['title'] = name
        if auto_id:
            criteria['auto_id'] = auto_id
        if control_type:
            criteria['control_type'] = control_type

        specific_symbol = main_window.child_window(**criteria)
        if specific_symbol.exists(timeout=3):
            logger.info(specific_symbol.get_properties())
            return True
        else:
            logger.info("未找到" + name + "窗格，可能目前不在指定的界面")
            return False
    except Exception as e:
        logger.error(f"发生错误: {e}")
        return False


# 点不到就报错（前提是能搜索到该按钮）
def click_by_controls(name=None, auto_id=None, control_type=None):
    try:
        main_window = connect_aoi_window()
        criteria = {}
        if name:
            criteria['title_re'] = f".*{name}.*"
        if auto_id:
            criteria['auto_id'] = auto_id
        if control_type:
            criteria['control_type'] = control_type

        button = main_window.child_window(**criteria)
        if button.exists(timeout=3):
            button.click_input()
            logger.info("已点击按钮：" + (name if name else "未指定名称"))
        else:
            logger.info("点击时未找到指定的按钮")
    except Exception as e:
        logger.error(f"发生错误: {e}")


# 确认程序无卡顿/闪退
def caton_or_flashback(window_name):
    try:
        # 首先检索AOI进程是否存在
        aoi_process = None
        for proc in psutil.process_iter():
            if proc.name() == window_name:
                aoi_process = proc
                break

        if not aoi_process:
            logger.warning(f"{window_name} 进程可能已经崩溃")
            raise RuntimeError(f"{window_name} 进程不存在")

        # 检查程序状态
        if aoi_process.status() == psutil.STATUS_STOPPED:
            logger.warning(f"{window_name} 程序已停止，可能已经崩溃")
            raise RuntimeError(f"{window_name} 程序已停止")

        # 检查CPU和内存占用，每秒检查一次，持续3秒
        zero_cpu_count = 0
        max_memory_usage = 0
        # for _ in range(3):
        cpu_usage = aoi_process.cpu_percent(interval=1)
        #     current_memory_usage = aoi_process.memory_info().rss / (1024 * 1024)
        #     max_memory_usage = max(max_memory_usage, current_memory_usage)
        #     if cpu_usage > 0:
        #         zero_cpu_count = 0
        #         logger.info(f"{window_name} 程序运行中，CPU占用: {cpu_usage}%")
        #     else:
        #         zero_cpu_count += 1

        if max_memory_usage > 16000:  # 假设内存占用大于16000MB为异常 
            logger.warning(f"{window_name} 程序内存占用异常，最高内存占用: {max_memory_usage:.2f} MB")
            raise MemoryError(f"{window_name} 程序内存占用过高")

        if zero_cpu_count == 3:
            logger.warning(f"{window_name} 程序无响应，CPU占用率连续3秒为0%")
            raise TimeoutError(f"{window_name} 程序无响应")

        logger.info(f"{window_name} 程序运行正常，最终CPU占用: {cpu_usage}%，最高内存占用: {max_memory_usage:.2f} MB")

    except (RuntimeError, TimeoutError, MemoryError) as e:
        logger.error(f"检测到程序异常: {e}")
        raise  # 可以重新抛出异常，或者进行其他处理


# ==============================识别处理===============================
# 寻找准星
def get_crosshair_center():
    logger.info("开始寻找准星")
    # 定义准星的颜色
    crosshair_color = (255, 69, 0)

    # 定义截图区域
    left = 540
    top = 150
    width = 1500 - 540
    height = 740 - 150

    # 截取指定区域的屏幕图像
    screenshot = pyautogui.screenshot(region=(left, top, width, height))
    screenshot_np = np.array(screenshot)

    # 转换颜色到HSV
    screenshot_hsv = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2HSV)
    crosshair_color_hsv = cv2.cvtColor(np.uint8([[crosshair_color]]), cv2.COLOR_RGB2HSV)[0][0]

    # 定义颜色的HSV范围，更严格的范围
    hue_variation = 5
    saturation_variation = 50
    value_variation = 50
    lower_bound = np.array([crosshair_color_hsv[0] - hue_variation, crosshair_color_hsv[1] - saturation_variation,
                            crosshair_color_hsv[2] - value_variation])
    upper_bound = np.array([crosshair_color_hsv[0] + hue_variation, crosshair_color_hsv[1] + saturation_variation,
                            crosshair_color_hsv[2] + value_variation])

    # 创建颜色掩码
    mask = cv2.inRange(screenshot_hsv, lower_bound, upper_bound)

    # 计算掩码的质心
    M = cv2.moments(mask)
    if M["m00"] != 0:
        cx = int(M["m10"] / M["m00"]) + left
        cy = int(M["m01"] / M["m00"]) + top
        logger.info(f"全屏准星位置: ({cx}, {cy})")
        return (cx, cy)
    else:
        logger.error("未找到准星")
        return None


# 检测屏幕是否有大量某个颜色
def check_color_expand():
    # 捕获屏幕截图
    screenshot = ImageGrab.grab()
    screenshot_np = np.array(screenshot)

    # 定义目标颜色和容差
    target_color = np.array([188, 157, 160])  # 注意：OpenCV使用BGR格式
    tolerance = 20  # 容差值

    # 计算颜色范围
    lower_bound = target_color - tolerance
    upper_bound = target_color + tolerance

    # 创建掩码
    mask = cv2.inRange(screenshot_np, lower_bound, upper_bound)

    # 计算掩码中的白色像素比例
    coverage = np.sum(mask) / (screenshot_np.shape[0] * screenshot_np.shape[1] * 255)

    # 判断是否有大量的目标颜色
    if coverage > 0.25:
        logger.info("元件四周遮罩出现大量粉色")
        return True
    else:
        logger.info("元件四周遮罩未出现大量粉色")
        return False

# 检测滚动条占比，以确认元件/芯片数量(换方法了 废弃)
def check_chip_coverage(scrollbar_color, scrollbar_region):
    # 截取指定区域的屏幕图像
    screenshot = ImageGrab.grab(bbox=scrollbar_region)
    screenshot_np = np.array(screenshot)

    # 定义目标颜色和容差
    target_color = np.array(scrollbar_color)  # 直接使用RGB颜色
    tolerance = 5 

    # 计算颜色范围
    lower_bound = target_color - tolerance
    upper_bound = target_color + tolerance

    # 创建掩码
    mask = cv2.inRange(screenshot_np, lower_bound, upper_bound)

    # 计算掩码中的白色像素比例
    coverage = np.sum(mask) / (screenshot_np.shape[0] * screenshot_np.shape[1] * 255)

    # 返回百分比
    return coverage
# 检测斜杠太麻烦了 暂时没时间，用颜色代替了
def if_checked(top_left, bottom_right):
    # 调整坐标
    adjusted_top_left, adjusted_bottom_right = adjust_coordinates(top_left, bottom_right)

    # 截取屏幕上的复选框区域
    screenshot = ImageGrab.grab(
        bbox=(adjusted_top_left[0], adjusted_top_left[1], adjusted_bottom_right[0], adjusted_bottom_right[1]))
    screenshot_np = np.array(screenshot)
    screenshot_np = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2BGR)

    # 先检测蓝色
    lower_blue = np.array([100, 0, 0])
    upper_blue = np.array([255, 100, 100])
    blue_mask = cv2.inRange(screenshot_np, lower_blue, upper_blue)

    if np.any(blue_mask):
        return True

    # 如果没有识别到蓝色，继续检测其他目标颜色
    target_colors = [(70, 70, 88), (32, 32, 54), (107, 107, 107)]
    tolerance = 15  # 容差值

    for color in target_colors:
        lower_bound = np.array(color) - tolerance
        upper_bound = np.array(color) + tolerance

        # 创建掩码
        mask = cv2.inRange(screenshot_np, lower_bound, upper_bound)

        # 检查是否有目标颜色的像素点
        if np.any(mask):
            return True
    return False
# 确保打勾框打勾情况
def is_checked(top_left, bottom_right, expect_checked, times=1):
    # 使用if_checked函数来检查是否勾选
    time.sleep(1)
    check = if_checked(top_left, bottom_right)
    if check == expect_checked:
        pass
    else:
        center_x = (top_left[0] + bottom_right[0]) // 2
        center_y = (top_left[1] + bottom_right[1]) // 2
        pyautogui.click(center_x, center_y, clicks=times)
# 过滤颜色 目的是过滤掉某个颜色的影响
def filter_color(image, target_color=None):
    # 将图像转换为numpy数组
    image_np = np.array(image)
    
    if target_color is not None:
        # 定义颜色范围
        lower_bound = np.array(target_color) - 1  # 缩小颜色范围
        upper_bound = np.array(target_color) + 1  # 缩小颜色范围
        # 创建颜色掩码
        mask = cv2.inRange(image_np, lower_bound, upper_bound)
        # 应用掩码
        filtered_image = cv2.bitwise_and(image_np, image_np, mask=mask)
    else:
        # 转换为灰度图像
        filtered_image = cv2.cvtColor(image_np, cv2.COLOR_BGR2GRAY)
    
    return PILImage.fromarray(filtered_image)
# 在屏幕中获取text的中心坐标并点击
def lcs_length(s1: str, s2: str) -> int:
    """
    计算两个字符串的最长公共子序列长度
    """
    m, n = len(s1), len(s2)
    # 初始化二维数组
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m):
        for j in range(n):
            if s1[i] == s2[j]:
                dp[i + 1][j + 1] = dp[i][j] + 1
            else:
                dp[i + 1][j + 1] = max(dp[i][j + 1], dp[i + 1][j])
    return dp[m][n]

def lcs_ratio(target: str, source: str) -> float:
    """
    计算目标字符串作为子序列出现在源字符串中的比例
    如果目标文本完全以正序包含则比例为1，否则小于1
    """
    if not target:
        return 0.0
    return lcs_length(target, source) / len(target)


# 在屏幕中获取文本的中心坐标并点击
def click_by_ocr(text, times=1, timeout=10, tolerance=0.6):
    start_time = time.time()
    logger.info(f"在{timeout}秒内使用 PaddleOCR 识别屏幕中的文字，要求包含率大于 {tolerance}，目标文本为：{text}")
    
    # 判断输入的文本是否为空
    if not text:
        logger.error("目标文本为空")
        return False

    # 根据文本中中文的比例决定语言
    chinese_count = sum(1 for char in text if '\u4e00' <= char <= '\u9fff')
    lang = 'ch' if (chinese_count / len(text)) >= 0.6 else 'en'

    while time.time() - start_time < timeout:
        try:
            logger.debug("开始截取屏幕截图")
            screenshot = pyautogui.screenshot()
            screenshot_cv = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
            logger.debug("使用 PaddleOCR 进行文字识别")
            # 直接调用全局变量 paddleOCR 的ocr方法
            result = paddleOCR.ocr(img=screenshot_cv, cls=True)
            
            # 扁平化OCR返回的结果，确保结构统一
            detections = []
            if result and isinstance(result[0][0][0], (int, float)):
                detections = result
            else:
                for sublist in result:
                    detections.extend(sublist)
            
            best_match = None
            highest_similarity = 0.0

            # 定义文本标准化函数（忽略标点符号影响），转换为字符串后处理
            def normalize_text(input_text):
                return re.sub(r'[^\w\s]', '', str(input_text))
            
            normalized_target_text = normalize_text(text)
            
            # 遍历所有检测项，计算目标文本与检测文本的包含率（按顺序匹配）
            for detection in detections:
                if len(detection) == 3:
                    bbox, detected_text, conf = detection
                elif len(detection) == 2:
                    bbox, detected_text = detection
                    conf = 1.0
                else:
                    logger.debug("忽略格式不符的检测项")
                    continue

                normalized_detected = normalize_text(detected_text)
                # 计算lcs_ratio，保证字符顺序一致
                similarity = lcs_ratio(normalized_target_text, normalized_detected)
                # logger.debug(f"检测到文本: {detected_text}，包含率: {similarity:.2f}")

                if similarity > highest_similarity:
                    highest_similarity = similarity
                    best_match = (bbox, detected_text, similarity)
            
            if best_match and highest_similarity >= tolerance:
                bbox, detected_text, similarity = best_match
                # 计算bbox的中心坐标
                x_coords = [point[0] for point in bbox]
                y_coords = [point[1] for point in bbox]
                center_x = int(sum(x_coords) / len(x_coords))
                center_y = int(sum(y_coords) / len(y_coords))
                pyautogui.click(center_x, center_y, clicks=times)
                logger.info(f"点击成功，识别到文本: {detected_text}，中心坐标: ({center_x}, {center_y})，包含率: {similarity:.2f}")
                return True
            else:
                logger.debug(f"未达到容差要求，最高包含率: {highest_similarity:.2f}。")
        except Exception as e:
            logger.error(f"在尝试点击过程中发生异常：{e}")
        time.sleep(1)
    
    # 如果超时未识别到目标文本，生成截图
    screenshot_path = f"{text}.png"
    screenshot.save(screenshot_path)
    logger.info(f"超时未识别并点击目标文本：{text}。已生成截图：{screenshot_path}。请检查OCR识别结果，可能原因：目标文本存在但包含率不足")
    return False

# 定义一个内部函数，用于获取指定窗口句柄对应的进程名称
def get_proc_name(handle):
    try:
        _, pid = win32process.GetWindowThreadProcessId(handle)
        return psutil.Process(pid).name()
    except Exception as exc:
        return f"无法获取进程名称: {exc}"
# def get_dpi_scaling_factor():
#     # 获取屏幕的 DPI，96 是标准 DPI
#     user32 = ctypes.windll.user32
#     hDC = user32.GetDC(0)
#     dpi = ctypes.windll.gdi32.GetDeviceCaps(hDC, 88)  # LOGPIXELSX = 88
#     scaling_factor = dpi / 96
#     return scaling_factor
# 识别图片并点击
def click_by_png(image_path, times=1, timeout=20, if_click_right=0, tolerance=0.8, region=None, instance=1, use_random=0, type="default", click_index=None, preference=None, if_filter_color=False, object_of_reference=None, direction=None):
    start_time = time.time()
    clicked = False  # 标记是否成功点击
    image_path = image_fit_screen(image_path)
    logger.info(f"正在识别并点击图片……图片路径为: {image_path}")

    # 读取目标图像并计算其平均颜色
    target_image = cv2.imread(image_path)
    target_avg_color = cv2.mean(target_image)[:3]

    # 如果提供了参考对象和方向，计算新的搜索区域
    if object_of_reference and direction:
        # 首先定位参考对象
        ref_location = pyautogui.locateOnScreen(object_of_reference, confidence=tolerance)
        if ref_location is not None:
            screen_width, screen_height = pyautogui.size()
            ref_left = int(ref_location.left)
            ref_top = int(ref_location.top)
            ref_width = int(ref_location.width)
            ref_height = int(ref_location.height)

            if direction == 'up':
                region = (0, 0, screen_width, ref_top + ref_height)
            elif direction == 'down':
                region = (0, ref_top, screen_width, screen_height - ref_top)
            elif direction == 'left':
                region = (0, 0, ref_left + ref_width, screen_height)
            elif direction == 'right':
                region = (ref_left, 0, screen_width - ref_left, screen_height)
            else:
                logger.error(f"未知的方向: {direction}")
                raise ValueError(f"未知的方向: {direction}")

            # 确保 region 中的值为整数且宽/高为正
            region = (int(region[0]), int(region[1]), max(int(region[2]), 0), max(int(region[3]), 0))
            logger.info(f"参考对象位置: left={ref_left}, top={ref_top}, width={ref_width}, height={ref_height}")
            logger.info(f"计算后的搜索区域: {region}")
        else:
            logger.error(f"无法找到参考对象: {object_of_reference}")
            raise Exception(f"无法找到参考对象: {object_of_reference}")

    while time.time() - start_time < timeout:
        try:
            screenshot = pyautogui.screenshot(region=region)
            screenshot_np = np.array(screenshot)

            # 过滤颜色
            if if_filter_color:
                filtered_screenshot = filter_color(screenshot, target_avg_color)
                filtered_screenshot_np = np.array(filtered_screenshot)
            else:
                filtered_screenshot_np = screenshot_np

            locations = list(pyautogui.locateAllOnScreen(image_path, confidence=tolerance, region=region))
            if locations and len(locations) >= instance:
                if click_index is not None:
                    if click_index <= len(locations):
                        location = locations[click_index - 1]  # 获取第 click_index 个匹配的图片
                    else:
                        logger.error(f"click_index超出范围: {click_index}")
                        continue
                elif use_random == 1:
                    location = random.choice(locations)  # 随机选择一个匹配的图片
                elif preference:
                    if preference == "top":
                        location = min(locations, key=lambda loc: loc.top)
                    elif preference == "bottom":
                        location = max(locations, key=lambda loc: loc.top)
                    elif preference == "left":
                        location = min(locations, key=lambda loc: loc.left)
                    elif preference == "right":
                        location = max(locations, key=lambda loc: loc.left)
                    else:
                        location = locations[instance - 1]
                else:
                    location = locations[instance - 1]  # 默认取第 instance 个匹配的图片

                # 根据 type 参数计算点击的中心坐标
                if type == "bottom":
                    center_x = int(location.left + location.width / 2)
                    center_y = int(location.top + location.height)
                elif type == "right":
                    center_x = int(location.left + location.width)
                    center_y = int(location.top + location.height / 2)
                elif type == "left":
                    center_x = int(location.left)
                    center_y = int(location.top + location.height / 2)
                else:
                    center_x = int(location.left + location.width / 2)
                    center_y = int(location.top + location.height / 2)

                try:
                    coordinate = (center_x, center_y)
                    click_delay = 0.2  # 点击延时（秒）
                    logger.debug(f"目标坐标: {coordinate}，点击延时: {click_delay}秒")

                    click_executed = False
                    # 优先使用 pyautogui 执行点击操作
                    try:
                        if if_click_right == 1:
                            logger.info(f"使用 pyautogui 执行右键点击，目标坐标: {coordinate}")
                            pyautogui.rightClick(center_x, center_y)
                        else:
                            logger.info(f"使用 pyautogui 执行左键点击，目标坐标: {coordinate}，点击次数: {times}")
                            pyautogui.click(center_x, center_y, clicks=times)
                        time.sleep(click_delay)
                        logger.info("通过 pyautogui 成功发送鼠标点击消息")
                        click_executed = True
                    except Exception as pyautogui_exception:
                        logger.error(f"pyautogui 点击失败: {pyautogui_exception}，尝试使用 win32gui")
                        try:
                            # 备选方案：使用 win32gui 发送点击消息
                            hwnd = win32gui.WindowFromPoint(coordinate)
                            logger.debug(f"调用 win32gui.WindowFromPoint 后获得的窗口句柄: {hwnd}")
                            if hwnd:
                                window_title = win32gui.GetWindowText(hwnd)
                                logger.info(f"成功获取窗口句柄: {hwnd}，对应窗口标题: {window_title}")
                                lParam = (coordinate[1] << 16) | (coordinate[0] & 0xFFFF)
                                logger.debug(f"计算得到 lParam = {lParam} (x: {coordinate[0]}, y: {coordinate[1]})")
                                if if_click_right == 1:
                                    WM_RBUTTONDOWN = 0x0204
                                    WM_RBUTTONUP = 0x0205
                                    logger.info(f"发送 WM_RBUTTONDOWN 消息到窗口句柄 {hwnd}")
                                    win32gui.PostMessage(hwnd, WM_RBUTTONDOWN, 0x00000002, lParam)
                                    time.sleep(click_delay)
                                    logger.info(f"发送 WM_RBUTTONUP 消息到窗口句柄 {hwnd}")
                                    win32gui.PostMessage(hwnd, WM_RBUTTONUP, 0, lParam)
                                else:
                                    WM_LBUTTONDOWN = 0x0201
                                    WM_LBUTTONUP = 0x0202
                                    logger.info(f"发送 WM_LBUTTONDOWN 消息到窗口句柄 {hwnd}")
                                    win32gui.PostMessage(hwnd, WM_LBUTTONDOWN, 0x00000001, lParam)
                                    time.sleep(click_delay)
                                    logger.info(f"发送 WM_LBUTTONUP 消息到窗口句柄 {hwnd}")
                                    win32gui.PostMessage(hwnd, WM_LBUTTONUP, 0, lParam)
                                time.sleep(click_delay)
                                logger.info("通过 win32gui 成功发送鼠标点击消息")
                                click_executed = True
                            else:
                                logger.warning(f"无法通过 win32gui 获取有效窗口句柄，点击失败，目标坐标: {coordinate}")
                        except Exception as win32_exc:
                            logger.error(f"win32gui 点击失败: {win32_exc}")

                    if click_executed:
                        logger.info(f"点击图片成功，图片路径为 {image_path}，坐标为 {coordinate}")
                        clicked = True  # 标记点击操作已成功
                        break
                except Exception as e:
                    logger.error(f"异常：点击操作失败，错误详情: {e}")
                    # 点击失败时不终止循环，继续尝试
            else:
                logger.info(f"按钮/标识识别失败: {image_path}")
        except Exception as e:
            logger.error(f"按钮/标识识别失败: {e}")
        time.sleep(1)
    if not clicked:
        logger.error(f"按钮/标识识别失败: 在{timeout}秒内，图片路径为{image_path}")
        raise Exception(f"按钮/标识识别失败: 在{timeout}秒内，图片路径为{image_path}")

def search_symbol(symbol, timeout=10, region=None, tolerance=0.9, if_filter_color=False, object_of_reference=None, direction=None):
    start_time = time.time()
    symbol = image_fit_screen(symbol)
    logger.info(f"在{timeout}秒内识别图片，要求准确度系数为{tolerance}，图片路径为{symbol}")
    
    # 读取目标图像并计算其平均颜色
    target_image = cv2.imread(symbol)
    target_avg_color = cv2.mean(target_image)[:3]
    
    # 如果提供了参考对象和方向，计算新的搜索区域
    if object_of_reference and direction:
        # 首先定位参考对象
        ref_location = pyautogui.locateOnScreen(object_of_reference, confidence=tolerance)
        if ref_location is not None:
            screen_width, screen_height = pyautogui.size()
            # 将 ref_location 的属性转换为整数
            ref_left = int(ref_location.left)
            ref_top = int(ref_location.top)
            ref_width = int(ref_location.width)
            ref_height = int(ref_location.height)

            if direction == 'up':
                region = (0, 0, screen_width, ref_top)
            elif direction == 'down':
                region = (0, ref_top + ref_height, screen_width, screen_height - (ref_top + ref_height))
            elif direction == 'left':
                region = (0, 0, ref_left, screen_height)
            elif direction == 'right':
                region = (ref_left + ref_width, 0, screen_width - (ref_left + ref_width), screen_height)
            else:
                logger.error(f"未知的方向: {direction}")
                raise ValueError(f"未知的方向: {direction}")

            # 确保 region 中的值都是整数，并且宽度和高度为正数
            region = (
                int(region[0]),
                int(region[1]),
                max(int(region[2]), 0),
                max(int(region[3]), 0)
            )
            logger.info(f"参考对象位置: left={ref_left}, top={ref_top}, width={ref_width}, height={ref_height}")
            logger.info(f"计算后的搜索区域: {region}")
        else:
            logger.error(f"无法找到参考对象: {object_of_reference}")
            raise Exception(f"无法找到参考对象: {object_of_reference}")
    
    if timeout != 0:
        while time.time() - start_time < timeout:
            try:
                screenshot = pyautogui.screenshot(region=region)
                screenshot_np = np.array(screenshot)
                screenshot_avg_color = cv2.mean(screenshot_np)[:3]
                
                # 过滤颜色
                if if_filter_color:
                    filtered_screenshot = filter_color(screenshot, target_avg_color)
                    filtered_screenshot_np = np.array(filtered_screenshot)
                else:
                    filtered_screenshot_np = screenshot_np
                
                location = pyautogui.locateOnScreen(symbol, region=region, confidence=tolerance)
                if location is not None:
                    center_x = location.left + location.width // 2
                    center_y = location.top + location.height // 2
                    logger.info(f"已确认{symbol}存在，坐标为({center_x}, {center_y})")
                    time.sleep(0.5)
                    return True
            except pyautogui.ImageNotFoundException:
                pass
            except Exception as e:
                logger.error(f"发生异常: {e}")
                raise Exception(f"发生异常: {e}")
        return False
    else:
        try:
            while True:
                screenshot = pyautogui.screenshot(region=region)
                screenshot_np = np.array(screenshot)
                screenshot_avg_color = cv2.mean(screenshot_np)[:3]
                
                # 过滤颜色
                if if_filter_color:
                    filtered_screenshot = filter_color(screenshot, target_avg_color)
                    filtered_screenshot_np = np.array(filtered_screenshot)
                else:
                    filtered_screenshot_np = screenshot_np
                
                location = pyautogui.locateOnScreen(symbol, region=region, confidence=tolerance)
                if location is not None:
                    center_x = location.left + location.width // 2
                    center_y = location.top + location.height // 2
                    logger.info(f"已确认{symbol}存在，坐标为({center_x}, {center_y})")
                    time.sleep(0.5)
                    return True
        except pyautogui.ImageNotFoundException:
            return False
        except Exception as e:
            logger.error(f"发生异常: {e}")
            raise Exception(f"发生异常: {e}")

def search_symbol_erroring(symbol, timeout=10, region=None, tolerance=0.9, if_filter_color=False, object_of_reference=None, direction=None):
    start_time = time.time()
    symbol = image_fit_screen(symbol)
    logger.info(f"在{timeout}秒内识别图片，要求准确度系数为{tolerance}，图片路径为{symbol}，识别不到自动报错")
    
    # 读取目标图像并计算其平均颜色
    target_image = cv2.imread(symbol)
    target_avg_color = cv2.mean(target_image)[:3]
    
    # 如果提供了参考对象和方向，计算新的搜索区域
    if object_of_reference and direction:
        ref_location = pyautogui.locateOnScreen(object_of_reference, confidence=tolerance)
        if ref_location is not None:
            screen_width, screen_height = pyautogui.size()
            if direction == 'up':
                region = (0, 0, screen_width, ref_location.top)
            elif direction == 'down':
                region = (0, ref_location.top + ref_location.height, screen_width, screen_height - (ref_location.top + ref_location.height))
            elif direction == 'left':
                region = (0, 0, ref_location.left, screen_height)
            elif direction == 'right':
                region = (ref_location.left + ref_location.width, 0, screen_width - (ref_location.left + ref_location.width), screen_height)
            else:
                logger.error(f"未知的方向: {direction}")
                raise ValueError(f"未知的方向: {direction}")
            logger.info(f"计算后的搜索区域: {region}")
        else:
            logger.error(f"无法找到参考对象: {object_of_reference}")
            raise Exception(f"无法找到参考对象: {object_of_reference}")
    
    if timeout != 0:
        while time.time() - start_time < timeout:
            try:
                screenshot = pyautogui.screenshot(region=region)
                screenshot_np = np.array(screenshot)
                screenshot_avg_color = cv2.mean(screenshot_np)[:3]
                
                # 过滤颜色
                if if_filter_color:
                    filtered_screenshot = filter_color(screenshot, target_avg_color)
                    filtered_screenshot_np = np.array(filtered_screenshot)
                else:
                    filtered_screenshot_np = screenshot_np
                
                location = pyautogui.locateOnScreen(symbol, region=region, confidence=tolerance)
                if location is not None:
                    center_x = location.left + location.width // 2
                    center_y = location.top + location.height // 2
                    logger.info(f"已确认图片存在，坐标为({center_x}, {center_y})")
                    time.sleep(0.5)
                    return True
            except pyautogui.ImageNotFoundException:
                pass
            except Exception as e:
                logger.error(f"发生异常: {e}")
                raise Exception(f"发生异常: {e}")
        # 如果超时后还没有找到符号，抛出超时异常
        logger.error(f"超时: 没找到图片,图片路径为{symbol}")
        raise Exception(f"超时: 没找到图片,图片路径为{symbol}")
    else:
        try:
            while True:
                screenshot = pyautogui.screenshot(region=region)
                screenshot_np = np.array(screenshot)
                screenshot_avg_color = cv2.mean(screenshot_np)[:3]
                
                # 过滤颜色
                if if_filter_color:
                    filtered_screenshot = filter_color(screenshot, target_avg_color)
                    filtered_screenshot_np = np.array(filtered_screenshot)
                else:
                    filtered_screenshot_np = screenshot_np
                
                location = pyautogui.locateOnScreen(symbol, region=region, confidence=tolerance)
                if location is not None:
                    center_x = location.left + location.width // 2
                    center_y = location.top + location.height // 2
                    logger.info(f"已确认图片存在，识别到的坐标为({center_x}, {center_y})")
                    time.sleep(0.5)
                    return True
        except pyautogui.ImageNotFoundException:
            raise Exception(f"没找到图片: 图片路径为{symbol}")
        except Exception as e:
            logger.error(f"发生异常: {e}")
            raise Exception(f"发生异常: {e}")

def count_symbol_on_region(symbol, object_of_reference=None, direction=None, region=None, confidence=0.7):
    """
    统计屏幕上某个标识的数量

    :param symbol: 要查找的标识图片
    :param object_of_reference: 参考对象的图片路径
    :param direction: 查找方向，可选值为 'up', 'down', 'left', 'right'
    :param region: 要查找的屏幕区域，格式为(x, y, width, height)，默认为None表示全屏查找
    :param confidence: 查找的置信度，默认为0.9
    :return: 返回找到的标识数量
    """
    try:
        # 如果提供了参考对象和方向，计算新的搜索区域
        if object_of_reference and direction:
            # 首先定位参考对象
            ref_location = pyautogui.locateOnScreen(object_of_reference, confidence=confidence)
            if ref_location is not None:
                screen_width, screen_height = pyautogui.size()
                # 将 ref_location 的属性转换为整数
                ref_left = int(ref_location.left)
                ref_top = int(ref_location.top)
                ref_width = int(ref_location.width)
                ref_height = int(ref_location.height)

                if direction == 'up':
                    region = (0, 0, screen_width, ref_top)
                elif direction == 'down':
                    region = (0, ref_top + ref_height, screen_width, screen_height - (ref_top + ref_height))
                elif direction == 'left':
                    region = (0, 0, ref_left, screen_height)
                elif direction == 'right':
                    region = (ref_left + ref_width, 0, screen_width - (ref_left + ref_width), screen_height)
                else:
                    logger.error(f"未知的方向: {direction}")
                    raise ValueError(f"未知的方向: {direction}")

                # 确保 region 中的值都是整数，并且宽度和高度为正数
                region = (
                    int(region[0]),
                    int(region[1]),
                    max(int(region[2]), 0),
                    max(int(region[3]), 0)
                )
                logger.info(f"参考对象位置: left={ref_left}, top={ref_top}, width={ref_width}, height={ref_height}")
                logger.info(f"计算后的搜索区域: {region}")
            else:
                logger.error(f"无法找到参考对象: {object_of_reference}")
                raise Exception(f"无法找到参考对象: {object_of_reference}")

        locations = list(pyautogui.locateAllOnScreen(symbol, region=region, confidence=confidence))
        count = len(locations)
        logger.info(f"在屏幕上找到 {count} 个 {symbol}")
        return count
    except pyautogui.ImageNotFoundException:
        logger.error(f"未找到图片: {symbol}")
        return 0
    except Exception as e:
        logger.error(f"count_symbol_on_region函数发生异常: {e}")
        return 0




# =========================装饰器=========================
# 用于记录每个用例函数的最终状态
test_case_status = {}

def screenshot_error_to_excel(max_attempts=2, running_event=None):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None

            for attempt in range(max_attempts):
                if running_event and not running_event.is_set():
                    break
                try:
                    logger.info(f"开始执行 {func.__name__} (尝试 {attempt + 1}/{max_attempts})")
                    auto_close_msgbox(f"开始执行 {func.__name__}", "信息", 5000, topmost=True)
                    
                    result = func(*args, **kwargs)
                    
                    if running_event and not running_event.is_set():
                        raise Exception("Execution stopped by user")
                    
                    logger.info(f"{func.__name__} 执行完毕")
                    auto_close_msgbox(f"{func.__name__} 执行完毕", "信息", 5000, topmost=True)
                    
                    test_case_status[func.__name__] = "success"  # 标记为成功
                    return result, None
                except Exception as e:
                    last_exception = e
                    logger.error(f"{func.__name__} 执行出错 (尝试 {attempt + 1}/{max_attempts}): {e}")
                    current_function_name = func.__name__
                    path = os.getcwd()
                    logger.debug(f"excel保存路径: {path}")
                    screenshot_to_excel(current_function_name, e)
                    
                    if attempt < max_attempts - 1:
                        logger.warning(f"准备重试 {func.__name__}")
                        pyautogui.hotkey("ctrl","d")
                        time.sleep(1)
                        start_time = time.time()
                        processes_to_kill = ["AOI.exe","SPCViewMain.exe", "DVPro.UI.exe"]
                        for proc in psutil.process_iter(['pid', 'name']):
                            if any(proc_name in proc.info['name'] for proc_name in processes_to_kill):
                                try:
                                    parent_proc = psutil.Process(proc.info['pid'])
                                    children = parent_proc.children(recursive=True)
                                    for child in children:
                                        child.kill()
                                    parent_proc.kill()
                                    parent_proc.wait()
                                except Exception as kill_error:
                                    logger.error(f"无法杀掉进程 {proc.info['name']} (PID: {proc.info['pid']}): {kill_error}")
                        elapsed_time = time.time() - start_time
                        logger.info(f"杀掉进程总用时: {elapsed_time} 秒")
                        
                        aoi_still_running = any("AOI" in p.info['name'] for p in psutil.process_iter(['name']))
                        if not aoi_still_running:
                            logger.info("所有AOI进程已关闭，准备重试")
                        else:
                            logger.error("AOI进程关闭失败，但仍将尝试重试")
                    else:
                        logger.error(f"{func.__name__} 达到最大尝试次数 ({max_attempts})")
                        auto_close_msgbox(f"方法名: {func.__name__}, 异常信息: {str(e)}", "错误信息", 5000, topmost=True)
                        test_case_status[func.__name__] = "failed"  # 标记为失败
            return None, last_exception
        return wrapper
    return decorator
def auto_close_msgbox(message, title, timeout=10000, topmost=False):
    def show_message():
        root = tk.Tk()
        root.withdraw()
        top = tk.Toplevel(root)
        top.title(title)
        label = tk.Label(top, text=message)
        label.pack(padx=20, pady=20)
        if topmost:
            top.attributes("-topmost", True)
        top.after(timeout, top.destroy)
        root.after(timeout + 100, root.destroy)  # 确保在弹框销毁后销毁root
        root.mainloop()

    threading.Thread(target=show_message).start()
# 报错时截图存至excel，后关闭aoi
def screenshot_to_excel(test_case_name, exception):
    logger.info("开始截图异常情况")
    # 获取当前工作目录
    current_dir = os.getcwd()
    log_dir = os.path.join(current_dir, "log")
    logger.info("创建log目录")
    os.makedirs(log_dir, exist_ok=True)  # 确保log目录存在
    logger.info("创建完成")

    try:
        # 截图
        screenshot_file = os.path.join(log_dir, "temp_screenshot.png")
        if os.path.exists(screenshot_file):
            os.remove(screenshot_file)  # 删除已经存在的screenshot_file
        screenshot = ImageGrab.grab()
        logger.info("开始保存截图")
        screenshot = screenshot.resize((screenshot.width // 2, screenshot.height // 2))  # 缩小图片尺寸以减少文件大小
        screenshot.save(screenshot_file, optimize=True, quality=85)  # 保存截图为文件，优化并压缩质量
        logger.info("保存截图完成")

        # 定义Excel文件路径
        excel_path = os.path.join(log_dir, "test_results.xlsx")

        # 检查Excel文件是否存在，如果不存在则创建
        if not os.path.exists(excel_path):
            logger.info("创建excel")
            wb = Workbook()
        else:
            logger.info("加载excel")
            wb = load_workbook(excel_path)
        ws = wb.active

        # 确定下一个空白行
        row = ws.max_row + 1
        logger.info("写入数据")

        # 写入数据
        ws[f"A{row}"] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f"B{row}"] = test_case_name
        ws[f"C{row}"] = str(exception)
        logger.info("插入图片")

        # 将截图插入到Excel
        img = ExcelImage(screenshot_file)
        img.anchor = f"D{row}"  # 设置图片的锚点
        ws.add_image(img)
        logger.info("数据处理完毕，开始保存excel")
        wb.save(excel_path)
        logger.info(f"Excel文件保存在: {excel_path}")
        logger.info("保存完毕")
    except Exception as e:
        logger.error(f"在保存截图和数据到Excel时发生错误: {e}")
    finally:
        if os.path.exists(screenshot_file):
            os.remove(screenshot_file)  # 清理临时文件


# ====================业务处理========================
# 执行方法结束后打开循环运行
def delete_bad_mark():
    try:
        check_and_launch_aoi()
        open_program(if_specific=True)
        while click_color(region=config.COMPONENT_OPERATION_REGION, target_color_rgb=(133,43,222), offset=False):
            time.sleep(0.2)
            pyautogui.rightClick()
            if search_symbol(config.DELETE_BAD_MARK, 2, tolerance=0.7):
                click_by_png(config.DELETE_BAD_MARK, tolerance=0.7)
            else:
                break
            if search_symbol(config.QUESTION_MARK, 2, tolerance=0.7):
                pyautogui.press("enter")
                time.sleep(0.1)
    except Exception as e:
        logger.error(f"删除坏板标记时发生错误: {e}")
    click_by_png(config.SAVE)
    if search_symbol(config.QUESTION_MARK):
        pyautogui.press("enter")
    time.sleep(3)
    close_aoi()
def loop(job_path,job_name):
    check_and_launch_aoi()
    check_loop()
    open_program(if_specific=True, job_path=job_path, job_name=job_name)
    click_by_png(config.PLAY, 2)
    for _ in range (3):
        time.sleep(5)
        pyautogui.press("enter")
        time.sleep(5)

def check_test_time():
    right_side_text = get_text_by_text("测试时", "right")
    if not right_side_text:
        return False
    corrected_text = right_side_text.replace(';', ':')
    logger.info(f"识别到测试时右侧的内容为: {corrected_text}")
    try:
        detected_time = datetime.datetime.strptime(corrected_text, "%Y/%m/%d %H:%M:%S")
    except ValueError:
        raise Exception("识别到的时间格式不正确")
    if (datetime.datetime.now() - detected_time).total_seconds() <= 300:
        logger.info("识别到的时间在五分钟内")
        return True
    else:
        logger.warning("识别到的时间不在五分钟内")
        return False
    
# 检测rv中是否有五分钟内的新数据
def check_new_data_in_rv(if_new_data = None):
    # 先选择未复核
    if search_symbol(config.RV_NO_CHECKED_NO_CHOSED, 2):
        click_by_png(config.RV_NO_CHECKED_NO_CHOSED, timeout=2)
    # 空的的话就跳过
    if search_symbol(config.RV_PCB_LIST_EMPTY, 2):
        click_by_png(config.RV_CHECKED_NO_CHOSED, timeout=2)
    else:
        # 否则 拉到最下面 点击最后一条 查看右上角job信息
        data_coordinate = random.choice(list(pyautogui.locateAllOnScreen(config.RV_PCB_LIST_LANE, confidence=0.8)))
        scroll_down(data_coordinate)
        click_by_png(config.RV_PCB_LIST_LANE, 2, 5, preference="down",tolerance=0.7)
        if_have_new_data_in_no_checked = check_test_time(True)
        if if_have_new_data_in_no_checked:
            if if_new_data:
                logger.info("rv中正确出现了一笔五分钟内的新数据")
            else:
                raise Exception("rv中出现了一笔五分钟内的新数据")

    # 进已复核再确认一下 拉到最下面 查看右上角job信息
    click_by_png(config.RV_CHECKED_NO_CHOSED, timeout=2)
    if_have_new_data_in_checked = None
    if search_symbol(config.RV_PCB_LIST_EMPTY, 2):
        pass
    else:
        data_coordinate = random.choice(list(pyautogui.locateAllOnScreen(config.RV_PCB_LIST_LANE, confidence=0.8)))
        scroll_down(data_coordinate)
        click_by_png(config.RV_PCB_LIST_LANE, 2, 5, preference="down",tolerance=0.7)

        if_have_new_data_in_checked = check_test_time(True)
        if if_have_new_data_in_checked:
            if if_new_data:
                logger.info("rv中正确出现了一笔五分钟内的新数据")
            else:
                raise Exception("rv中出现了一笔五分钟内的新数据")
                
    if not if_have_new_data_in_checked and not if_have_new_data_in_no_checked:
        if if_new_data:
            raise Exception("rv未出现五分钟内的新数据")


# 找有对应窗口的元件 并确认含待料的情况
def find_component_window(window_names, limit_time=300, image=None):
    start_time = time.time()
    retry_attempts = 5
    # 边在左侧点击元件，边往下找
    while True:
        before_screenshot = pyautogui.screenshot(region=config.BOARD_COMPONENTS_REGION)
        pyautogui.press("i")
        time.sleep(3)
        after_screenshot = pyautogui.screenshot(region=config.BOARD_COMPONENTS_REGION)
        
        if before_screenshot == after_screenshot:
            retry_attempts -= 1
            if retry_attempts == 0:
                logger.error(f"程式列表到底了都没找到含对应窗口的元件，窗口图片路径为{window_names}")
                raise Exception(f"程式列表到底了都没找到含对应窗口的元件，窗口图片路径为{window_names}")
        else:
            retry_attempts = 5  # 重置试错机会
            # 连续0.2s不相同时重置试错机会
            start_compare_time = time.time()
            while time.time() - start_compare_time < 3:
                before_screenshot = pyautogui.screenshot(region=config.BOARD_COMPONENTS_REGION)
                time.sleep(0.2)
                after_screenshot = pyautogui.screenshot(region=config.BOARD_COMPONENTS_REGION)
                if before_screenshot != after_screenshot:
                    retry_attempts = 5
                    break

        if window_names is None:
            if search_symbol(config.COMPONENT_WINDOW_EMPTY, 3, region=config.COMPONENT_WINDOW_REGION, tolerance=0.7):
                break
        elif window_names == 1:
            if not search_symbol(config.COMPONENT_WINDOW_EMPTY, 3, region=config.COMPONENT_WINDOW_REGION, tolerance=0.7):
                break
        else:
            # 判断window_names是一个还是多个
            if isinstance(window_names, str):
                window_names = [window_names]
            
            # 判断是否存在对应窗口
            for window_name in window_names:
                if search_symbol(window_name, timeout=2,  region=config.COMPONENT_WINDOW_REGION, if_filter_color=True):
                    click_by_png(window_name, if_filter_color=True, region=config.COMPONENT_WINDOW_REGION)
                    time.sleep(7)
                    if image:
                        if search_symbol(image, timeout=2, if_filter_color=True, region=config.PALETTE_REGION):
                            return
                    else:
                        return

        # 超过指定时间没找到的话 raise
        if time.time() - start_time > limit_time:
            raise Exception(f"超过{limit_time}秒未找到窗口: 图片路径为{window_names}")
def make_package():
    logger.info("开始制造封装")
    if search_symbol(config.TOOL_DARK, timeout=2):
        click_by_png(config.TOOL_DARK)
    click_by_png(config.PACKAGE_TYPE_MANAGE)
    pyautogui.keyDown('shift')
    pyautogui.click(730, 375)
    pyautogui.keyUp('shift')

    click_by_png(config.EDIT_PACKAGE_TYPE)
    write_text((900,525),"test")
    click_by_png(config.YES)
    write_text((820,290),'test')
    click_by_png(config.QUERY)
    # 读取封装a
    pyautogui.click((980,360))
    click_by_png(config.COPY_PART_NO_NAME)
    package_a = pyperclip.paste()
    # 读取封装b
    pyautogui.click((980,375))
    click_by_png(config.COPY_PART_NO_NAME)
    package_b = pyperclip.paste()
    time.sleep(1)

    click_by_png(config.CLOSE_BUTTON)
    logger.info("封装制造完毕")
    click_by_png(config.EDIT_DARK)

    return package_a,package_b


# 整版界面处理元件
def board_component_process(menu_choice):
    target_color = (0, 255, 0)
    while True:
        # 截取区域内的屏幕图像
        screenshot = pyautogui.screenshot(region=config.COMPONENT_REGION)
        screenshot_np = np.array(screenshot)
        # 寻找颜色匹配的像素点
        matches = np.where(np.all(screenshot_np == target_color, axis=-1))
        if matches[0].size > 0:
            # 获取所有匹配的点的坐标
            coordinates = list(zip(matches[1], matches[0]))
            # 按照x坐标从大到小排序（从右到左）
            coordinates.sort(reverse=True, key=lambda coord: coord[0])
            for x, y in coordinates:
                # 转换为区域内的相对坐标
                x += config.COMPONENT_REGION[0]
                y += config.COMPONENT_REGION[1]
                # 右键点击
                pyautogui.rightClick(x, y)
                # 搜索符号
                if search_symbol(menu_choice, timeout=10, tolerance=0.7):
                    click_by_png(menu_choice, timeout=10, tolerance=0.7)
                    break
                else:
                    pyautogui.click((400, 5))
            break

# 添加待料
def add_waiting_material():
    for _ in range(2):
        click_by_png(config.ADD_STANDARD_IMAGE)
        click_by_png(config.YES)
        if search_symbol(config.ADD_IMAGE_CLOSE, 5):
            click_by_png(config.ADD_IMAGE_CLOSE)
        time.sleep(5)


# 画检测窗口
# def add_window(button="w"):
#     time.sleep(1.5)
#     crosshair_center = get_crosshair_center()
#     if crosshair_center is None:
#         logger.info("未找到准星中心")
#         return

#     nearby_point = (crosshair_center[0] + 3, crosshair_center[1] + 3)
#     logger.info(nearby_point)
#     # 绘制框框（获取准星旁边的颜色，扩大到颜色分界处，截取坐标）
#     target_color = pyautogui.screenshot(region=config.COMPONENT_REGION).getpixel(nearby_point)
#     # 转换颜色到HSV
#     target_color_hsv = cv2.cvtColor(np.uint8([[target_color]]), cv2.COLOR_RGB2HSV)[0][0]
#     # 定义颜色的HSV范围，初始范围
#     hue_variation = 5
#     saturation_variation = 10
#     value_variation = 10
#     found = False
#     last_top_left = None
#     last_bottom_right = None

#     while not found:
#         lower_bound = np.array([target_color_hsv[0] - hue_variation, target_color_hsv[1] - saturation_variation,
#                                 target_color_hsv[2] - value_variation])
#         upper_bound = np.array([target_color_hsv[0] + hue_variation, target_color_hsv[1] + saturation_variation,
#                                 target_color_hsv[2] + value_variation])
#         # 截取config.COMPONENT_REGION内的屏幕图像
#         screenshot = pyautogui.screenshot(region=config.COMPONENT_REGION)
#         screenshot_np = np.array(screenshot)
#         screenshot_hsv = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2HSV)
#         # 创建颜色掩码
#         mask = cv2.inRange(screenshot_hsv, lower_bound, upper_bound)

#         # 使用形态学操作增强边界
#         kernel = np.ones((5, 5), np.uint8)
#         mask = cv2.dilate(mask, kernel, iterations=2)
#         mask = cv2.erode(mask, kernel, iterations=2)

#         # 使用边缘检测
#         edges = cv2.Canny(mask, 100, 200)

#         # 寻找连贯区域的轮廓
#         contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
#         target_contour = None
#         for contour in contours:
#             if cv2.pointPolygonTest(contour, nearby_point, False) >= 0:
#                 target_contour = contour
#                 found = True
#                 break
#             # 如果找到了符合条件的连贯区域
#         if target_contour is not None:
#             x, y, w, h = cv2.boundingRect(target_contour)
#             # 扩大区域
#             expand_margin = 10
#             top_left = (x - expand_margin, y - expand_margin)
#             bottom_right = (x + w + expand_margin, y + h + expand_margin)
#             last_top_left = top_left
#             last_bottom_right = bottom_right
#             logger.info("找到疑似cad区域，左上及右下坐标如下" + str(top_left) + "," + str(bottom_right))
#         else:
#             # 增加HSV范围并重试
#             hue_variation += 1
#             saturation_variation += 2
#             value_variation += 2
#             if hue_variation > 180 or saturation_variation > 255 or value_variation > 255:
#                 logger.info("未识别出cad区域，可能准心不在cad内")
#                 raise Exception("未识别出cad区域，可能准心不在cad内")
                

#     if last_top_left and last_bottom_right:
#         # 使用pyautogui模拟鼠标拖动
#         pyautogui.press(button)
#         pyautogui.moveTo(last_top_left, duration=1)
#         pyautogui.mouseDown()
#         pyautogui.moveTo(last_bottom_right, duration=1)
#         pyautogui.mouseUp()
#         logger.info("cad描边完毕")
#     else:
#         logger.info("未找到任何区域")

def add_window(button="w"):
    time.sleep(2.5)
    crosshair_center = get_crosshair_center()
    if crosshair_center is None:
        logger.error("未找到准星中心，使用默认坐标(935,445)")
        crosshair_center = (935, 445)

    # 初始化区域大小和扩展步长
    region_size = 50  # 扩大初始搜索区域到50像素
    expansion_step = 5
    found = False
    top_left = bottom_right = None
    start_time = time.time()  # 记录开始时间

    while not found and time.time() - start_time < 30:
        logger.info("正在寻找区域......")
        # 计算当前区域
        region = (crosshair_center[0] - region_size, crosshair_center[1] - region_size, 2 * region_size, 2 * region_size)
        screenshot = pyautogui.screenshot(region=region)
        screenshot_np = np.array(screenshot)
        screenshot_hsv = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2HSV)

        # 将图像数据转换为二维数组
        pixel_colors = screenshot_hsv.reshape((-1, 3))

        # 使用K-means聚类找到最常见的颜色块
        kmeans = KMeans(n_clusters=5, random_state=0, n_init=10).fit(pixel_colors)
        unique, counts = np.unique(kmeans.labels_, return_counts=True)
        dominant_cluster = unique[np.argmax(counts)]
        dominant_color_hsv = kmeans.cluster_centers_[dominant_cluster]

        # 定义颜色的HSV范围
        hue_variation = 10
        saturation_variation = 40
        value_variation = 40
        lower_bound = np.array([dominant_color_hsv[0] - hue_variation, dominant_color_hsv[1] - saturation_variation,
                                dominant_color_hsv[2] - value_variation], dtype=np.uint8)
        upper_bound = np.array([dominant_color_hsv[0] + hue_variation, dominant_color_hsv[1] + saturation_variation,
                                dominant_color_hsv[2] + value_variation], dtype=np.uint8)

        # 创建颜色掩码
        mask = cv2.inRange(screenshot_hsv, lower_bound, upper_bound)

        # 使用形态学操作增强边界
        kernel = np.ones((5, 5), np.uint8)
        mask = cv2.dilate(mask, kernel, iterations=2)
        mask = cv2.erode(mask, kernel, iterations=2)

        # 使用边缘检测
        edges = cv2.Canny(mask, 100, 200)

        # 寻找连贯区域的轮廓
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if contours:
            # 找到最大的轮廓
            largest_contour = max(contours, key=cv2.contourArea)
            x, y, w, h = cv2.boundingRect(largest_contour)
            # 检查准星是否在轮廓内
            if (x <= region_size and x + w >= region_size and y <= region_size and y + h >= region_size):
                top_left = (x + region[0], y + region[1])  # 转换为原始图像的坐标
                bottom_right = (x + w + region[0], y + h + region[1])
                found = True
                logger.info(f"找到疑似cad区域，左上及右下坐标如下: {top_left}, {bottom_right}")
            else:
                region_size += expansion_step  # 扩大搜索区域
                if region_size > 500:  # 设置最大搜索限制
                    logger.info("扩展到最大范围后未找到任何区域")
                    break
        else:
            region_size += expansion_step  # 扩大搜索区域
            if region_size > 500:  # 设置最大搜索限制
                logger.info("扩展到最大范围后未找到任何区域")
                break

    if found:
        # 使用pyautogui模拟鼠标拖动
        if button:
            pyautogui.press(button)
        pyautogui.moveTo(top_left, duration=1)
        pyautogui.mouseDown()
        pyautogui.moveTo(bottom_right, duration=1)
        pyautogui.mouseUp()
        logger.info("cad描边完毕")
    else:
        logger.info("未找到任何区域,开始瞎画")
        if button:
            pyautogui.press(button)
        pyautogui.moveTo(872 + random.randint(-5, 5), 362 + random.randint(-5, 5), duration=1)
        pyautogui.mouseDown()
        pyautogui.moveTo(1000 + random.randint(-5, 5), 528 + random.randint(-5, 5), duration=1)
        pyautogui.mouseUp()
        logger.info("瞎画完成")
# 打开程式(随机) 
def open_program(program_type=0, if_recent=True, if_specific=False, job_path=None, job_name=None):
    """
    打开程式
    :param program_type: 0 - 默认, 1 - non-compressed, 2 - compressed
    """
    # 先检测有没有什么傻逼窗口 关掉
    if search_symbol(config.NO, 1.5):
        click_by_png(config.NO, tolerance=0.9)
    pyautogui.press('enter')
    time.sleep(1)
    # 再检测有没有狗日的登录框

    login_symbols = [
        (config.USER_LOGIN_LIGHT, 0.95),
        (config.USER_LOGIN_DARK, 0.95),
        (config.USER_LOGIN_CHINESE_LIGHT, 0.8),
        (config.USER_LOGIN_CHINESE_DARK, 0.8)
    ]
    if any(search_symbol(symbol, 1.5, tolerance=tol) for symbol, tol in login_symbols):
        write_text((865,500),"000")
        time.sleep(0.5)
        pyautogui.press('enter')
        if search_symbol(config.INVALID_PASSWORD, 1.5, tolerance=0.95):
            # 关闭密码错误窗口
            pyautogui.press('enter')
            time.sleep(0.5)
            # 输入下个密码
            write_text((865,500),"sinictek")
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(1)
            # 关闭登录成功窗口
            pyautogui.press('enter')
        else:
            # 关闭登录成功窗口
            pyautogui.press('enter')
    # 现在才开始打开程式
    click_by_png(config.OPEN_PROGRAM)
    # 先检测有没有什么傻逼窗口 关掉
    if search_symbol(config.NO, 1.5):
        click_by_png(config.NO, tolerance=0.9)
    directory = r"D:\EYAOI\JOB"
    time.sleep(3)
    pyautogui.press("enter")
    if if_specific and job_path is None and job_name is None:
        pyperclip.copy(directory)
        time.sleep(0.1)
        logger.info(f"路径为{directory},剪切板内容: {pyperclip.paste()}")
        pyautogui.hotkey('ctrl', 'v')
    elif if_specific and job_path is not None and job_name is not None:
        parent_directory = os.path.dirname(job_path)
        if parent_directory == r"D:\EYAOI":
            parent_directory = r"D:\EYAOI\JOB"
        logger.info(f"打开程式窗口选择文件夹窗口写入job_path的上一级目录: {parent_directory}, 写入job_name: {job_name}")
        pyperclip.copy(parent_directory)
        time.sleep(0.1)
        logger.info(f"剪切板内容: {pyperclip.paste()}")
        pyautogui.hotkey('ctrl', 'v')
        # write_text((650,220),parent_directory)
    else:
        logger.info(f"打开程式窗口选择文件夹窗口写入job_path: {directory}")
        # write_text((650,220),directory)
        pyperclip.copy(directory)
        time.sleep(0.1)
        logger.info(f"剪切板内容: {pyperclip.paste()}")
        pyautogui.hotkey('ctrl', 'v')
    time.sleep(2)
    pyautogui.press("enter")
    click_by_png(config.SELECT_FOLDER,tolerance=0.6)
    if not if_recent:
        click_by_png(config.OFFSET_LEFT_1,tolerance=0.98, type="left")

    time.sleep(2)
    if if_specific and job_path is not None and job_name is not None:
        write_text((570,220),job_name)
    elif if_specific and job_path is None and job_name is None:
        write_text((570,220),"全算法")
    # 双击程式
    if program_type == 0:
        symbols = [config.OPEN_PROGRAM_PLUS, config.OPEN_PROGRAM_CURSOR]
        found = False
        for symbol in symbols:
            if search_symbol(symbol, 3, tolerance=0.7):
                click_by_png(symbol, 2, tolerance=0.7)
                found = True
                break
        if not found:
            raise Exception("打开程式时未找到程式")

    if program_type == 1:
        plus_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_PLUS))
        cursor_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_CURSOR))
        for pos in plus_positions + cursor_positions:
            pyautogui.click(pos)
            time.sleep(1)
        click_by_png(config.OPEN_PROGRAM_PLUS, 2)
        if search_symbol(config.OPEN_PROGRAM_EMPTY, 3):
            raise Exception("该程式无job")
    if program_type == 2:
        plus_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_PLUS))
        cursor_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_CURSOR))
        for pos in plus_positions + cursor_positions:
            pyautogui.click(pos)
            time.sleep(1)
        click_by_png(config.OPEN_PROGRAM_CURSOR, 2)
        if search_symbol(config.OPEN_PROGRAM_EMPTY, 3):
            raise Exception("该程式无job")
    if search_symbol(config.YES, 2):
        click_by_png(config.YES)
    if search_symbol(config.OPEN_PROGRAM_SWITCH, 2):
        check_checkbox_status_before_text("允许程式切换",True)
        pyautogui.press('enter')
    while search_symbol(config.PROGRAM_LOADING, 5):
        time.sleep(5)
    # cnm的傻逼岗位 傻逼软件一堆窗口
    minimize_service_process_manager()


# 获取最近编辑的一个程式
def get_topest_program():
    cursor_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_CURSOR))
    plus_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_PLUS))
    all_positions = cursor_positions + plus_positions
    if all_positions:
        topest = sorted(all_positions, key=lambda pos: pos.top)[0]
    else:
        logger.error("未找到任何cursor或plus的坐标")
        raise Exception("未找到任何cursor或plus的坐标")

    return topest


# 确保在编辑界面
def ensure_in_edit_mode():
    login_symbols = [
        (config.USER_LOGIN_LIGHT, 0.95),
        (config.USER_LOGIN_DARK, 0.95),
        (config.USER_LOGIN_CHINESE_LIGHT, 0.8),
        (config.USER_LOGIN_CHINESE_DARK, 0.8)
    ]
    if any(search_symbol(symbol, 1.5, tolerance=tol) for symbol, tol in login_symbols):
        write_text((865,500),"000")
        time.sleep(0.5)
        pyautogui.press('enter')
        if search_symbol(config.INVALID_PASSWORD, 3):
            pyautogui.press('enter')
            write_text((865,500),"sinictek")
            time.sleep(0.5)
            pyautogui.press('enter')                        
    # 先尝试打开程式元件栏
    found_light = search_symbol(config.PROGRAM_COMPONENT_LIGHT, 0)
    found_dark = search_symbol(config.PROGRAM_COMPONENT_DARK, 0)

    if found_light or found_dark:
        if found_dark:
            click_by_png(config.PROGRAM_COMPONENT_DARK)
        if not click_component():
            # 如果点击元件失败，则执行以下代码
            open_program()
            if search_symbol(config.BOARD_AUTO, 50):
                if search_symbol(config.PROGRAM_COMPONENT_DARK, 30):
                    time.sleep(3)
                    click_by_png(config.PROGRAM_COMPONENT_DARK)
            click_component()
    else:
        open_program()
        if search_symbol(config.BOARD_AUTO, 50):
            if search_symbol(config.PROGRAM_COMPONENT_DARK, 30):
                time.sleep(3)
                click_by_png(config.PROGRAM_COMPONENT_DARK)
        click_component()
    # 防止有是否保存的提示
    if search_symbol(config.NO, 2):
        click_by_png(config.NO)
    time.sleep(7)
    click_by_png(config.EDIT_DARK)
    time.sleep(3)

# 确保打开了有多个拼版的job
def ensure_multiple_collages():
    click_by_png(config.OPEN_PROGRAM)
    time.sleep(1)
    # 搜索并点击屏幕上所有＋
    plus_positions = list(pyautogui.locateAllOnScreen(config.OPEN_PROGRAM_PLUS))
    for pos in plus_positions:
        pyautogui.click(pos)
        time.sleep(3)
        # read_text_ocr识别率太低了 改用图片
        if not search_symbol(config.OPEN_PROGRAM_SINGLE_BOARD, 5, config.PROGRAM_INFORMATION_REGION, 0.99):
            click_by_png(config.OPEN_PROGRAM_LOAD_1)
            click_by_png(config.YES)
            return
    logger.error("未找到多个拼版的job")
    raise Exception("未找到多个拼版的job")


# 点击元件
def click_component(click_index=None):
    try:
        # 几种元件里点一个
        components = [config.NO_CHECKED_COMPONENT, config.CHECKED_COMPONENT, config.PASS_COMPONENT, config.NO_PASS_COMPONENT]
        # random.shuffle(components)  # 随机排序元件列表
        for component in components:
            if search_symbol(component, 2, tolerance=0.95, region=config.BOARD_INFORMATION_REGION):
                time.sleep(1)
                click_by_png(component, 2, use_random=1, tolerance=0.8, region=config.BOARD_INFORMATION_REGION, click_index=click_index)
                time.sleep(5)
                return True
        return False
    except Exception as e:
        logger.error(f"点击元件时出错: {e}")
        return False

# 检测该料号的窗口算法参数都已编辑完成，相同封装类型的其他料号的元件的窗口算法参数编辑情况        
def check_same_package_same_param(if_same_param):
    try:
        time.sleep(3)
        # 先截图已选择元件的对应区域
        logger.debug("点击测试按钮")
        click_by_png(config.TEST, 2, region=config.BOARD_INFORMATION_REGION)
        time.sleep(5)
        # 截图PackageType区域和算法参数区域 前者判断封装类型，后者判断参数相同
        logger.debug("截图PackageType区域和算法参数区域")
        package_type_image_old = pyautogui.screenshot(region=config.PACKAGE_TYPE_REGION)
        alg_param_image_old = pyautogui.screenshot(region=config.ALG_PARAM_REGION)
        
        # 遍历所有元件类型
        components = [config.NO_CHECKED_COMPONENT, config.CHECKED_COMPONENT, config.PASS_COMPONENT, config.NO_PASS_COMPONENT]
        for component in components:
            # 找到屏幕上所有该类型元件的位置
            logger.debug(f"查找元件类型: {component}")
            try:
                component_positions = list(pyautogui.locateAllOnScreen(component, region=config.BOARD_INFORMATION_REGION))
            except Exception as e:
                logger.warning(f"查找元件类型 {component} 时出错: {e}")
                continue
            if not component_positions:
                raise Exception("未找到任何元件")
            for pos in component_positions:
                logger.debug(f"双击元件位置: {pos}")
                pyautogui.doubleClick(pos)
                time.sleep(5)  # 等待界面响应
                
                # 检查是否是相同封装类型的元件 相同的话参数是否一致 不相同的话参数不一致
                logger.debug("检查封装类型")
                package_type_image_new = pyautogui.screenshot(region=config.PACKAGE_TYPE_REGION)
                if package_type_image_old == package_type_image_new:
                    # 检查算法参数是否一致
                    logger.debug("检查算法参数是否一致")
                    alg_param_image_new = pyautogui.screenshot(region=config.ALG_PARAM_REGION)
                    if alg_param_image_old == alg_param_image_new:
                        # 参数一致
                        if if_same_param:
                            logger.info("参数已编辑")
                        else:
                            raise Exception("参数已编辑，但预期为未编辑")
                    else:
                        # 参数不一致
                        if not if_same_param:
                            logger.info("参数未编辑")
                        else:
                            raise Exception("参数未编辑，但预期为已编辑")
                else:
                    alg_param_image_new = pyautogui.screenshot(region=config.ALG_PARAM_REGION)
                    if alg_param_image_old == alg_param_image_new:
                        raise Exception("不同封装类型算法参数一致")
    except Exception as e:
        logger.error(f"检查相同封装类型参数时出错: {e}")
        raise
# 检测屏幕区域内是否存在指定颜色
def check_color_in_region(target_color=(220, 20, 60), region=(862, 344, 1010 - 862, 543 - 344)):
    """
    检测屏幕指定区域内是否存在指定颜色的像素点。如果region为坐标点，则检测点周围10个像素范围内是否存在指定颜色。

    :param target_color: 要检测的颜色，格式为(R, G, B)
    :param region: 要检测的屏幕区域，格式为(x, y, width, height))或坐标点(x, y)或单个坐标点(x, y)
    :return: 如果找到指定颜色，返回True，否则返回False
    """
    if isinstance(region, tuple) and len(region) == 2:
        # 如果region为坐标点，则检测点周围10个像素范围内是否存在指定颜色
        logger.info(f"检测区域为坐标点: {region}")
        x, y = region
        for dx in range(-10, 11):
            for dy in range(-10, 11):
                try:
                    # 确保 x + dx 和 y + dy 是整数
                    x_coord = int(x + dx)
                    y_coord = int(y + dy)
                    pixel_color = pyautogui.pixel(x_coord, y_coord)
                    if pixel_color == target_color:
                        return True
                except Exception as e:
                    logger.warning(f"获取像素颜色时出错: {e}")
                    continue
        return False
    elif isinstance(region, tuple) and len(region) == 4:
        # 捕获指定区域的屏幕截图
        logger.info(f"检测区域为区域: {region}")
        screenshot = pyautogui.screenshot(region=region)

        # 遍历区域内的所有像素
        for x in range(region[2]):
            for y in range(region[3]):
                pixel_color = screenshot.getpixel((x, y))
                if pixel_color == target_color:
                    return True
        return False
    else:
        raise ValueError("region参数格式不正确，应为(x, y)、(x, y, width, height)")
    
# ===================================设置快捷键====================================
def shortcut_key_online_parameter_display_sync_package():
    # UI设置--快捷键设置--元件编辑，【在线调参显示同步封装】设置快捷键K
    if search_symbol(config.SETTING_DARK, 1):
        click_by_png(config.SETTING_DARK)
    else:
        if not search_symbol(config.SETTING_LIGHT, 3):
            bring_aoi_to_top()
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_SHORTCUT_KEY_SETTING)
    time.sleep(2)

    scroll_down((380, 250), config.SHORTCUT_KEY_COMPONENT_EDIT_REGION)
    write_text_textbox(config.PARAM_ONLINE_PARAMETER_DISPLAY_SYNC_PACKAGE, write_content="K", if_select_all=False)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# ====================================设置勾选====================================

# 参数设置--数据导出配置--在线调参，【保留最后PCB板数】设置N
def param_keep_the_last_pcb_number():
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    # 注意 轨道1/2不选的话会有问题
    if search_symbol(config.PARAM_TRACK_1, 2) and search_symbol(config.PARAM_TRACK_2, 2):
        click_by_png(config.PARAM_TRACK_1)

    write_text_textbox(config.PARAM_KEEP_THE_LAST_PCB_NUMBER, write_content="N")
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 参数设置--数据导出配置--在线调参，勾选【Good元件】、【NG元件】，或勾选【所有】
def param_good_and_ng_component_limits(good_limit, ng_limit):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)

    # 先确定轨道1/2选了
    if search_symbol(config.PARAM_TRACK_1, 2) and search_symbol(config.PARAM_TRACK_2, 2):
        click_by_png(config.PARAM_TRACK_1)

    # 再确定没有勾选所有
    if search_symbol(config.PARAM_ONLINE_ALL_YES, 2):
        click_by_png(config.PARAM_ONLINE_ALL_YES,type="bottom")

    if good_limit is not None:
        if search_symbol(config.PARAM_GOOD_COMPONENT_LIMIT, 2):
            click_by_png(config.PARAM_GOOD_COMPONENT_LIMIT)
        while search_symbol(config.PARAM_AMOUNT_LIMIT_ALL_YES, 2, tolerance=0.95):
            click_by_png(config.PARAM_AMOUNT_LIMIT_ALL_YES,timeout=2,tolerance=0.95,type="right")
        write_text((1275, 730), good_limit)
    if ng_limit is not None:
        if search_symbol(config.PARAM_NG_COMPONENT_LIMIT, 2):
            click_by_png(config.PARAM_NG_COMPONENT_LIMIT)
        while search_symbol(config.PARAM_AMOUNT_LIMIT_ALL_YES, 2, tolerance=0.95):
            click_by_png(config.PARAM_AMOUNT_LIMIT_ALL_YES,timeout=2,tolerance=0.95,type="right")
        write_text((1275, 755), ng_limit)
    # 无传参则勾所有
    if good_limit is None and ng_limit is None:
        if search_symbol(config.PARAM_GOOD_COMPONENT_LIMIT, 2):
            click_by_png(config.PARAM_GOOD_COMPONENT_LIMIT)
        if search_symbol(config.PARAM_NG_COMPONENT_LIMIT, 2):
            click_by_png(config.PARAM_NG_COMPONENT_LIMIT)
        while search_symbol(config.PARAM_AMOUNT_LIMIT_ALL_NO, 2, tolerance=0.95):
            click_by_png(config.PARAM_AMOUNT_LIMIT_ALL_NO,timeout=2,tolerance=0.95,type="right")

    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

def check_loop():
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(3)
    check_checkbox_status_before_text("打开左右循环", True)
    check_checkbox_status_before_text("重新载板", True)
    check_checkbox_status_before_text("左进右出", False)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】--【硬件设置】--【数据导出配置】--【数据导出配置】--勾选【输出测试数据】
def check_export_test_data(if_checked):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    check_checkbox_status_before_text("输出测试数据", if_checked)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# 【设置】--【硬件设置】--【数据导出配置】--【数据导出配置】--勾选【允许离线测试整板发送数据】
def check_offline_send_data(if_checked):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    check_checkbox_status_before_text("允许离线测试整板发送数据", if_checked)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


# 勾选共享元件库路径
def check_share_lib_path(if_checked):
    # 点开设置--硬件设置--数据导出配置--元件库配置
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    # 确保共享元件库路径为勾选状态 （用的坐标）
    is_checked((902, 291), (914, 303), if_checked, 1)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 勾选保存程式默认导出到公共库+勾选自动保存【设置】-【UI配置】-【程序设置】
def check_default_export_auto_save(if_default_export, if_auto_save):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(1)
    is_checked((659, 699), (671, 711), if_default_export)
    is_checked((659, 491), (671, 503), if_auto_save)
    time.sleep(0.5)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# 参数配置--UI配置--软件界面：不选【自动加载程式】 TODO 这边用图的 后面需改
def check_auto_load_program(if_auto_load_program):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(1)
    if if_auto_load_program:
        if search_symbol(config.SETTING_UI_AUTO_LOAD_PROGRAM_NO, 2):
            click_by_png(config.SETTING_UI_AUTO_LOAD_PROGRAM_NO)
    else:
        if search_symbol(config.SETTING_UI_AUTO_LOAD_PROGRAM_YES, 2):
            click_by_png(config.SETTING_UI_AUTO_LOAD_PROGRAM_YES)
    time.sleep(0.5)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


        



# 【设置】-【硬件设置】-【UI配置】-【图像设置】搜索范围xy最大扩展： 最小设置0.8，搜索范围扩展x y:设置100%
def check_xy_max_extension():
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(1)
    write_text((530, 90), "100")
    write_text((530, 115), "100")
    write_text((530, 145), "0.8")
    write_text((530, 170), "0.8")
    time.sleep(0.5)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

#【设置】-【硬件设置】-【UI配置】-【软件界面】-不勾【不允许黏贴元件到元件】
def check_not_allow_paste_component_to_component(if_not_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    is_checked((1268, 846), (1280, 1005), if_not_allow)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

#【设置】-【硬件设置】-【UI配置】-【软件界面】-不勾【不允许黏贴元件到空白处】
def check_not_allow_paste_component_to_blank(if_not_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    is_checked((1268, 872), (1280, 884), if_not_allow)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 勾选 允许跨元件复制
def check_cross_component_copy():
    # 参数配置-UI配置-软件界面：选择【允许跨元件复制】
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    is_checked((1268, 993), (1280, 1005), True)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


# 允许同步相同的封装,  默认同步封装
def check_not_sync_same_and_default_package(if_not_sync_same_package, if_default_sync_package):
    # 参数配置--UI配置--程序设置
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    is_checked((659, 825), (671, 837), if_not_sync_same_package)
    is_checked((659, 849), (671, 861), if_default_sync_package)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

def check_not_sync_same_package(if_not_sync_same_package):
    # 参数配置--UI配置--程序设置
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    is_checked((659, 825), (671, 837), if_not_sync_same_package)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 参数配置--流程配置--缺陷视图：【打开DV复判模式】、【DV自动确认】按钮
def check_dv(if_open_dv_mode=None, if_auto_check_dv=None):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_PROCESS_SETTING)
    time.sleep(2)
    if if_open_dv_mode is not None:
        check_checkbox_status_before_text("打开DV复判模式", if_open_dv_mode)
    if if_auto_check_dv is not None:
        check_checkbox_status_before_text("DV自动确认", if_auto_check_dv)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)



# 参数配置--硬件设置--数据导出配置--元件库--导入筛选限制
def check_import_filtering_restriction(percent):
    if search_symbol(config.SETTING_DARK, 3):
        click_by_png(config.SETTING_DARK)
    else:
        search_symbol_erroring(config.SETTING_LIGHT, 1)
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    write_text((1340,225), percent)
    time.sleep(1.5)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# 【设置】--【硬件设置】--【数据导出配置】--【数据导出配置】--勾选【输出路径】--更改该路径
def check_output_path(if_check,path):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    check_checkbox_status_before_text("输出路径",if_check,"right",200)
    # 去你妈的 坐标就坐标吧 能实现就行
    write_text((520,175), path)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】--【硬件设置】--【数据导出配置】--【数据导出配置】--勾选【使用日期文件夹】
def check_use_date_folder(if_use):
    """
    if_use为0 不勾选 1 选择YYYY  2 YYYY-MM 3 YYYY-MM-DD
    """
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    
    if if_use == 0:
        check_checkbox_status_before_text("使用日期文件夹", False)
    elif 1<= if_use <= 3:
        check_checkbox_status_before_text("使用日期文件夹", True)
        click_by_png(config.OFFSET_USE_DATE_FOLDER,type="right")
        try:
            location = pyautogui.locateOnScreen(config.DATE_FOLDER_LIST, confidence=0.5, minSearchTime=2)
        except pyautogui.ImageNotFoundException:
            logger.error("未找到日期文件夹下拉框")
            raise ValueError("未找到日期文件夹下拉框")
            
        if location:
                x, y, width, height = location
                section_height = height // 3
                click_y = y + section_height * (if_use + 0.5)
                click_x = x + width // 2
                pyautogui.click(click_x, click_y)
    else:
        raise ValueError("if_use 必须在 0 到 3 之间")
    
    time.sleep(1)
    click_by_png(config.APPLY, timeout=1.5, tolerance=0.75)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        click_by_png(config.CLOSE, 2, timeout=1.5, tolerance=0.85)
    else:
        click_by_png(config.CLOSE, 2, timeout=1.5, tolerance=0.85)
    time.sleep(0.5)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)



# 数据导出配置--数据导出配置--输出数据延时--下拉框选择倒计时的时间【x】
def check_output_data_delay(delay_time = 0):
    if search_symbol(config.SETTING_DARK, 3):
        click_by_png(config.SETTING_DARK)
    else:
        search_symbol_erroring(config.SETTING_LIGHT, 1)
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    click_by_png(config.OUTPUT_DELAY_TIME, type="left")
    if 0 <= delay_time <= 5:
        try:
            location = pyautogui.locateOnScreen(config.DELAY_TIME_LIST, confidence=0.65, minSearchTime=2)
        except pyautogui.ImageNotFoundException:
            logger.error("未找到倒计时下拉框")
            raise ValueError("未找到倒计时下拉框")

    if location:
        if 0 <= delay_time <= 5:
            x, y, width, height = location
            section_height = height // 6
            click_y = y + section_height * (delay_time + 0.5)
            click_x = x + width // 2
            pyautogui.click(click_x, click_y)
        else:
            logger.error("未找到倒计时下拉框")
            raise ValueError("未找到倒计时下拉框")
    else:
        raise ValueError("delay_time 必须在 0 到 5 之间")

    time.sleep(1.5)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


# 【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导出库后刷新树结构
def check_refresh_tree(if_fresh):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((1182, 123), (1194, 135), if_fresh)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# UI设置，演算法配置-点击开发者选项，输入密码：devsinictekaoi
def check_open_developer_options(type=None):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_DEVELOPER_OPTIONS)
    time.sleep(2)
    write_text((900,530),"devsinictekaoi")
    pyautogui.press('enter')

    if type == "save_3d_data_yes":
        if search_symbol(config.DEVELOPER_OPTIONS_SAVE_3D_DATA_NO, 3):
            click_by_png(config.DEVELOPER_OPTIONS_SAVE_3D_DATA_NO)
    elif type == "save_3d_data_no":
        if search_symbol(config.DEVELOPER_OPTIONS_SAVE_3D_DATA_YES, 3):
            click_by_png(config.DEVELOPER_OPTIONS_SAVE_3D_DATA_YES)
    elif type == "save_check_data_yes":
        if search_symbol(config.DEVELOPER_OPTIONS_SAVE_CHECK_DATA_NO, 3):
            click_by_png(config.DEVELOPER_OPTIONS_SAVE_CHECK_DATA_NO)
    elif type == "save_check_data_no":
        if search_symbol(config.DEVELOPER_OPTIONS_SAVE_CHECK_DATA_YES, 3):
            click_by_png(config.DEVELOPER_OPTIONS_SAVE_CHECK_DATA_YES)

    click_by_ocr("是")
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 参数配置--演算法配置--关联子框检测模式
def check_patent_not_NG(type):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    # 不计算
    if type == 1:
        pyautogui.click((935, 180))
        time.sleep(0.5)
        pyautogui.click((935, 200))
    # 继续计算
    if type == 2:
        pyautogui.click((935, 180))
        time.sleep(0.5)
        pyautogui.click((935, 215))
    # 继续关联
    if type == 3:
        pyautogui.click((935, 180))
        time.sleep(0.5)
        pyautogui.click((935, 230))
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 设置-硬件设置-演算法配置-勾选保存DJB文件
def check_save_djb(if_save_djb):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    check_checkbox_status_before_text("保存DJB文件", if_save_djb)

    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 设置-硬件设置-演算法配置-勾选所有算法（关闭所有算法）
def check_close_all_algs(if_close_all_algs = True, if_close_color_analysis = False):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    if if_close_all_algs:
        check_checkbox_status_before_text("所有算法",True)
        check_checkbox_status_before_text("基准点匹配",False)
        check_checkbox_status_before_text("条码检测",False)
        # 勾选的话 good的可以正常显示
        check_checkbox_status_before_text("颜色分析",if_close_color_analysis)
    else:
        check_checkbox_status_before_text("所有算法",True)
        pyautogui.moveTo(config.CENTRE)
        check_checkbox_status_before_text("所有算法",False)

    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】-【硬件设置】-【UI配置】-【软件界面】-不勾【自动选择窗口】
def check_auto_choose_window(if_auto_choose):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    is_checked((1268, 921), (1280, 933), if_auto_choose)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


def check_export_ok(if_export_ok=None, if_export_all_ok=None):
    if if_export_ok is None and if_export_all_ok is None:
        raise ValueError("check_export_ok参数不能都为空")
    # 参数配置--UI配置--程序设置
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    if if_export_ok is not None:
        if if_export_ok:
            if search_symbol(config.SETTING_UI_EXPORT_OK_NO, 2):
                click_by_png(config.SETTING_UI_EXPORT_OK_NO)
        else:
            if search_symbol(config.SETTING_UI_EXPORT_OK_YES, 2):
                click_by_png(config.SETTING_UI_EXPORT_OK_YES)
    if if_export_all_ok is not None:
        if if_export_all_ok:
            if search_symbol(config.SETTING_UI_EXPORT_ALL_OK_NO, 2):
                click_by_png(config.SETTING_UI_EXPORT_ALL_OK_NO)
        else:
            if search_symbol(config.SETTING_UI_EXPORT_ALL_OK_YES, 2):
                click_by_png(config.SETTING_UI_EXPORT_ALL_OK_YES)   
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    click_by_png(config.CLOSE, 2, timeout=1.5, tolerance=0.85)
    time.sleep(1)
    # pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        # pyautogui.press('enter')
        click_by_ocr("是")
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    else:
        click_by_ocr("关闭")
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 导出相同料号新增序列号 【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导出相同料号新增序列号
def check_export_pn_add_sn(if_export_pn_add_sn):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((902, 145), (914, 157), if_export_pn_add_sn)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

def check_allow_preview(if_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((1182, 145), (1194, 157), if_allow)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

#【设置】-【硬件设置】-【UI配置】-【图像设置】-【支持回退操作】
def check_allow_fallback(if_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    is_checked((333, 541), (345, 553), if_allow)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
    


# 【设置】-【硬件设置】-【UI配置】-【软件界面】-【允许跨元件复制】
def check_allow_copy_cross_component(if_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_UI_SETTING)
    time.sleep(2)
    is_checked((1268, 993), (1280, 1005), if_allow)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# 勾选【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导出代料影像数，输入1
def check_export_wm_img_1():
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    write_text((1070, 225), '1')
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 勾选【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导出代料影像数，输入1，勾选所有
def check_export_wm_img_1_all(if_export_wm_img_1_all):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    write_text((1070, 225), '1')
    time.sleep(1)
    is_checked((1131, 221), (1143, 233), if_export_wm_img_1_all)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)


# 导入库同步封装类型 【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导入库同步封装类型
def check_import_sync_package(if_import_sync_package):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((1182, 168), (1194, 180), if_import_sync_package)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 勾选【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--选择料号，输入天数1天
def check_pn_1_day():
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((1182, 193), (1194, 205), True)
    time.sleep(1)
    write_text((1320, 200), '1')
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 勾选【设置】--【演算法配置】--【Component Lib Inspect Setting]--导入元件库删除OCV待料 
def check_import_delete_ocv_wm(if_delete):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    is_checked((700, 243), (712, 255), if_delete)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】--【演算法配置】--【演算法配置】--允许cad teach
def check_allow_cad_teach(if_allow):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_ALGORITHM_SETTING)
    time.sleep(2)
    is_checked((349, 374), (361, 386), if_allow)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)
# 【设置】--【数据导出配置】--【元件库设置】--导入更新元件高度
def check_import_update_height(if_update):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((902, 244), (914, 256), if_update)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】--【数据导出配置】--【元件库设置】--导入更新元件x/y
def check_import_component_xy(if_update):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((902, 266), (914, 278), if_update)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【元件信息】--标准xyh前面的勾选框 注意：在编辑界面才可使用
def check_standard_xyh(if_checked):
    pyautogui.rightClick(config.CENTRE)
    time.sleep(1)
    click_by_png(config.COMPONENT_INFORMATION)
    time.sleep(1)
    is_checked((375, 574), (387, 586), if_checked)
    time.sleep(1)
    click_by_png(config.PROGRAM_ATTRIBUTE_CLOSE)
    time.sleep(1)

# 【设置】--【硬件设置】--【数据导出配置】--【元件库设置】--导出图像到元件库
def check_export_image_libs(if_export):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((902, 123), (914, 157), if_export)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

# 【设置】--【数据导出配置】--【元件库设置】--过滤辅助窗口（3d基准面、板弯补偿算法）
def filter_auxiliary_window(if_filter):
    click_by_ocr("设置")
    click_by_png(config.PARAM_HARDWARE_SETTING)
    time.sleep(2)
    click_by_png(config.PARAM_DATA_EXPORT_SETTING)
    time.sleep(2)
    is_checked((902, 194), (914, 206), if_filter)
    time.sleep(1)
    time.sleep(1)
    click_by_png(config.APPLY,timeout=1.5, tolerance=0.85)
    time.sleep(1)
    if search_symbol(config.NO, 1.5, tolerance=0.75):
        time.sleep(3)
        click_by_ocr("关闭")
    else:
        time.sleep(3)
        click_by_ocr("关闭")
    time.sleep(1)
    pyautogui.press('enter')
    if search_symbol(config.QUESTION_MARK, 2,tolerance=0.7):
        pyautogui.press('enter')
    if search_symbol(config.WARNING, 2):
        pyautogui.press('enter')
    start_time = time.time()
    while search_symbol(config.PARAM_SETTING_TOPIC, 2):
        if time.time() - start_time > 30:
            raise Exception("超过半分钟了设置界面仍未关闭")
        time.sleep(1.5)

def modify_component():
    time.sleep(2)
    pyautogui.press("b")
    time.sleep(5)
    pyautogui.keyDown('left')
    time.sleep(2)
    pyautogui.keyUp('left')

# ================================================文件处理=====================================
# 检查文件夹及其子文件夹中是否有文件在时间限制内创建或修改过
def check_new_data(path, name=None, minutes=5):
    recent_creations = []
    all_files_and_dirs = []

    if name:
        # 如果name非空，检查指定文件或文件夹
        target_path = os.path.join(path, name)
        if os.path.exists(target_path):
            ctime = os.path.getctime(target_path)
            mtime = os.path.getmtime(target_path)
            logger.debug(f"检测到文件或文件夹：{target_path}，创建时间：{ctime}，修改时间：{mtime}")
            all_files_and_dirs.append(target_path)
            if time.time() - ctime < minutes * 60 or time.time() - mtime < minutes * 60:
                recent_creations.append(target_path)
    else:
        # 如果name为空，检查所有文件和子文件夹
        for root, dirs, files in os.walk(path):
            for item in files:
                full_path = os.path.join(root, item)
                ctime = os.path.getctime(full_path)
                mtime = os.path.getmtime(full_path)
                logger.debug(f"检测到文件：{full_path}，创建时间：{ctime}，修改时间：{mtime}")
                all_files_and_dirs.append(full_path)
                if time.time() - ctime < minutes * 60 or time.time() - mtime < minutes * 60:
                    recent_creations.append(full_path)

    logger.info(f"路径 {path} 下的所有文件: {all_files_and_dirs}")
    if recent_creations:
        logger.info(f"最近创建或修改的文件: {recent_creations}")
        return True
    else:
        logger.info("没有最近创建或修改的文件")
        return False
def check_data_amount(path):
    a = 0
    for filename in os.listdir(path):
        a += 1
    return a


# 将剪切板内的内容与文件夹内内容作对比
def check_amount_content(coordinate, path):
    # 读内容至剪切板
    read_text_choosed(coordinate[0], coordinate[1])
    clipboard_content = pyperclip.paste()
    lines = clipboard_content.split('\n')

    if not lines:
        raise Exception("剪切板内容为空")

    # 提取第一行的数字
    try:
        expected_count = int(lines[0].strip())
    except Exception:
        raise Exception("第一行不包含有效的数字")

    # 获取文件夹中的所有文件名
    files = os.listdir(path)
    # 过滤掉Default文件夹
    files = [file for file in files if "Default" not in file]
    # 检查每一行内容是否存在于文件夹中的某个文件名中(需要忽略Default文件夹)
    matched_files = []
    for line in lines[1:]:  # 从第二行开始，因为第一行是数字
        found = False
        for file in files:
            if line in file:
                matched_files.append(file)
                found = True
                break
        if not found:
            print(f"没有找到匹配的文件：{line}")

    # 检查文件总数是否与第一行数字一致
    actual_count = len(matched_files)

    if actual_count == expected_count:
        print("文件数量匹配成功")
    else:
        print(f"文件数量不匹配：期望 {expected_count}, 实际 {actual_count}")


ALL_COMPONENTS = []


# 获取程式元件列表
def get_component_list():
    global ALL_COMPONENTS
    time.sleep(3)
    # 注意首次加载板的时候出现的和后续加载时的面板不同
    # 首次加载应该识别程式元件 点击后打开
    # 点击程式元件面板
    search_symbol(config.PROGRAM_COMPONENT_DARK, 60)
    time.sleep(5)
    click_by_controls(config.PROGRAM_COMPONENT_DARK, 2)
    time.sleep(0.3)
    confidence_level = 0.9
    try:
        # 识别未检测的元件坐标并保存，标记为no_checked
        no_checked_components = list(
            pyautogui.locateAllOnScreen(config.NO_CHECKED_COMPONENT, confidence=confidence_level))
        no_checked_positions = [{'x': pos.left, 'y': pos.top, 'status': 'no_checked', 'seen': False} for pos in
                                no_checked_components]
    except pyautogui.ImageNotFoundException:
        no_checked_positions = []
        logger.info("未检测的元件图像未找到。")

    try:
        # 识别已检测的元件坐标并保存，标记为checked
        checked_components = list(pyautogui.locateAllOnScreen(config.CHECKED_COMPONENT, confidence=confidence_level))
        checked_positions = [{'x': pos.left, 'y': pos.top, 'status': 'checked', 'seen': False} for pos in
                             checked_components]
    except pyautogui.ImageNotFoundException:
        checked_positions = []
        logger.info("已检测的元件图像未找到。")
    # 更新全局变量all_components
    ALL_COMPONENTS = no_checked_positions + checked_positions


# 获取选择框
def get_choose_box():
    logger.info("开始获取选择框")
    # 定位中心点附近的黄色方块
    center_x, center_y = 935, 445
    search_region = (center_x - 393, center_y - 148, 786, 296)  # 更新搜索区域
    screenshot = pyautogui.screenshot(region=search_region)
    image = np.array(screenshot)
    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
    hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # 定义黄色的HSV范围
    lower_yellow = np.array([30, 255, 255])
    upper_yellow = np.array([30, 255, 255])

    # 根据阈值构建掩模
    mask = cv2.inRange(hsv_image, lower_yellow, upper_yellow)

    # 寻找轮廓
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # 过滤小轮廓
    yellow_blocks = [cv2.boundingRect(cnt) for cnt in contours if cv2.contourArea(cnt) > 6]
    logger.info(yellow_blocks)
    # 检查找到的方块数量
    if len(yellow_blocks) == 8:
        # 按照距离中心点的距离排序
        yellow_blocks.sort(key=lambda pos: (pos[0] + pos[2] // 2 - center_x + search_region[0]) ** 2 + (
                pos[1] + pos[3] // 2 - center_y + search_region[1]) ** 2)
        logger.info(yellow_blocks)
        # 选择第5个最近的方块
        target_block = yellow_blocks[4]  # 选择第五个最近的方块，索引为4
        target_x = target_block[0] + target_block[2] // 2 + search_region[0]
        target_y = target_block[1] + target_block[3] // 2 + search_region[1]
        screen_width, screen_height = pyautogui.size()
        screen_target_x = target_x * screen_width / 1920
        screen_target_y = target_y * screen_height / 1080
        target = (screen_target_x, screen_target_y)
        logger.debug(target)
        return True, target
    elif len(yellow_blocks) == 4:
        # 随机选择一个方块
        target_block = random.choice(yellow_blocks)
        target_x = target_block[0] + target_block[2] // 2 + search_region[0]
        target_y = target_block[1] + target_block[3] // 2 + search_region[1]
        screen_width, screen_height = pyautogui.size()
        screen_target_x = target_x * screen_width / 1920
        screen_target_y = target_y * screen_height / 1080
        target = (screen_target_x, screen_target_y)
        logger.debug(target)
        return True, target
    elif len(yellow_blocks) > 0:
        return True, None
    else:
        return False, None

# 获取算法框中心点
def get_frame_center(color):
    # 截取整个屏幕
    screenshot = pyautogui.screenshot()
    screenshot = np.array(screenshot)
    
    # 精确颜色匹配（可以增加一点点容差进行调试）
    tolerance = 5
    lower_color = np.array([max(0, color[0] - tolerance), max(0, color[1] - tolerance), max(0, color[2] - tolerance)])
    upper_color = np.array([min(255, color[0] + tolerance), min(255, color[1] + tolerance), min(255, color[2] + tolerance)])
    
    # 根据阈值构建掩模
    mask = cv2.inRange(screenshot, lower_color, upper_color)
    
    # 寻找轮廓
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # 过滤小轮廓
    blocks = [cv2.boundingRect(cnt) for cnt in contours if cv2.contourArea(cnt) > 6]
    
    if blocks:
        # 选择第一个找到的方块
        target_block = blocks[0]
        target_x = target_block[0] + target_block[2] // 2
        target_y = target_block[1] + target_block[3] // 2
        center_point = (target_x, target_y)
        logger.info(f"中心点坐标: {center_point}")
        return center_point
    else:
        logger.info("未找到匹配的算法框中点")
        return None


# 扩大算法框
def expand_choose_box(point=config.CENTRE):
    # 获取选择框
    success, target = get_choose_box()
    if success and target is not None:
        # 左键点击按住方块中心点
        pyautogui.mouseDown(target)
        time.sleep(0.5)
        # 计算拖拽方向，远离中心点
        direction_x = target[0] - point[0]
        direction_y = target[1] - point[1]
        move_x = 50 if direction_x >= 0 else -50
        move_y = 50 if direction_y >= 0 else -50
        # 往远离中心点的方向拖拽50单位
        pyautogui.moveRel(move_x, move_y, duration=1)
        time.sleep(0.5)
        # 松开鼠标
        pyautogui.mouseUp()
    else:
        logger.debug(f"Success: {success}, Target: {target}")
        logger.error("未能成功获取选择框，随机选择一个黄色点拖曳")
        # 选择region.COMPONENT_OPERATION_REGION内任一rgb为（255，255，0)的点
        found_point = None
        for rgb in [(255, 255, 0), (242, 242, 23), (244, 244, 23),(251,249,0)]:
            found_point = get_color_direction_coordinate(rgb, config.COMPONENT_OPERATION_REGION, 'left')
            if found_point is not None:
                break
        if found_point is not None:
            # 左键点击按住找到的黄色或类似黄色的点
            pyautogui.mouseDown(found_point)
            time.sleep(0.5)
            # 计算拖拽方向，远离中心点
            direction_x = found_point[0] - point[0]
            direction_y = found_point[1] - point[1]
            move_x = 50 if direction_x >= 0 else -50
            move_y = 50 if direction_y >= 0 else -50
            # 往远离中心点的方向拖拽50单位
            pyautogui.moveRel(move_x, move_y, duration=1)
            time.sleep(0.5)
            # 松开鼠标
            pyautogui.mouseUp()
        else:
            logger.error("未能找到黄色或类似黄色的点")
# 获取color颜色在某个区域最靠某个方向的像素点的基于全屏的坐标点
def get_color_direction_coordinate(color, region, direction):
    # 截取指定区域的屏幕 (保留RGB格式)
    screenshot = pyautogui.screenshot(region=region)
    screenshot = np.array(screenshot)
    
    # 精确颜色匹配（可以增加一点点容差进行调试）
    lower_color = np.array([color[0] - 1, color[1] - 1, color[2] - 1])
    upper_color = np.array([color[0] + 1, color[1] + 1, color[2] + 1])
    
    # 生成颜色掩码
    mask = cv2.inRange(screenshot, lower_color, upper_color)
    
    # 获取非零像素点的坐标
    coords = cv2.findNonZero(mask)
    
    if coords is None:
        logger.error("未找到指定颜色的像素点")
        return None
    
    # 根据方向选择最靠近的像素点
    if direction == 'left':
        target_coord = min(coords, key=lambda x: x[0][0])
    elif direction == 'right':
        target_coord = max(coords, key=lambda x: x[0][0])
    elif direction == 'up':
        target_coord = min(coords, key=lambda x: x[0][1])
    elif direction == 'down':
        target_coord = max(coords, key=lambda x: x[0][1])
    else:
        raise Exception("get_color_direction_coordinate传参有问题")
    
    # 返回全屏的坐标 (直接加上区域的起始坐标)
    screen_x = target_coord[0][0] + region[0]
    screen_y = target_coord[0][1] + region[1]
    
    return screen_x, screen_y
# 计算该区域内颜色为rgb的像素点的数量
def get_color_in_region(color, region):
    # 截取指定区域的屏幕
    screenshot = pyautogui.screenshot(region=region)
    screenshot = np.array(screenshot)
    
    # 不需要颜色空间转换，直接处理RGB颜色
    # screenshot = cv2.cvtColor(screenshot, cv2.COLOR_RGB2BGR) # 可以省略这一行
    
    # 设置颜色范围，增加小容差 ±1
    lower_color = np.array([color[0] - 1, color[1] - 1, color[2] - 1])
    upper_color = np.array([color[0] + 1, color[1] + 1, color[2] + 1])
    
    # 生成颜色掩码，只有匹配的像素值会被设置为255，其他为0
    mask = cv2.inRange(screenshot, lower_color, upper_color)
    
    # 计算匹配颜色的像素点数量
    pixel_count = cv2.countNonZero(mask)
    
    return pixel_count

# 识别屏幕上是否存在指定文字
def get_text_coordinate(target_text, tolerance=0.66):
    try:
        # 截取屏幕
        screenshot = pyautogui.screenshot()

        # 将截图转换为OpenCV格式
        screenshot_cv = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)


        # 识别文字
        result = paddleOCR.ocr(screenshot_cv, cls=True)

        # 检查是否包含目标文字
        for line in result:
            for (bbox, (text, prob)) in line:
                similarity = SequenceMatcher(None, target_text, text).ratio()
                if similarity >= tolerance:
                    x = int(bbox[0][0])  # 获取文字的左上角x坐标
                    y = int(bbox[0][1])  # 获取文字的左上角y坐标
                    width = int(bbox[2][0] - bbox[0][0])  # 计算文字的宽度
                    height = int(bbox[2][1] - bbox[0][1])  # 计算文字的高度
                    logger.info(f"识别到与{target_text}相似的文字：{text}，坐标和尺寸：({x}, {y}, {width}, {height})")
                    return (x, y, width, height)  # 返回识别到的文字坐标和尺寸

        logger.info(f"未识别到与{target_text}相似的文字")
        return None
    except Exception as e:
        logger.error(f"发生异常: {e}")
        return None

# 获得文字右侧的文字
def get_text_by_text(text, direction="right", region=None):
    # 如果指定了区域，则截取该区域的屏幕，否则截取全屏
    if region:
        test_time = pyautogui.screenshot(region=region)
    else:
        test_time = pyautogui.screenshot()
    
    test_time_cv = cv2.cvtColor(np.array(test_time), cv2.COLOR_RGB2BGR)
    
    # 使用PaddleOCR进行文字识别
    result = paddleOCR.ocr(test_time_cv, cls=True)
    
    for line in result:
        for (bbox, (detected_text, prob)) in line:
            if text in detected_text:
                # 在图像上绘制识别到的文字区域
                cv2.rectangle(test_time_cv, bbox[0], bbox[2], (0, 255, 0), 2)
                cv2.putText(test_time_cv, detected_text, (bbox[0][0], bbox[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)

                if direction == "right":
                    target_x = bbox[1][0]  # 获取目标文字右侧的x坐标
                    for (other_bbox, (other_text, other_prob)) in line:
                        if other_bbox[0][0] > target_x and abs(other_bbox[0][1] - bbox[0][1]) < 10:
                            # 在图像上绘制相邻文字区域
                            cv2.rectangle(test_time_cv, other_bbox[0], other_bbox[2], (255, 0, 0), 2)
                            cv2.putText(test_time_cv, other_text, (other_bbox[0][0], other_bbox[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)
                            # 显示图像
                            cv2.imshow("Detected Text", test_time_cv)
                            cv2.waitKey(0)
                            cv2.destroyAllWindows()
                            return other_text.strip()
                elif direction == "left":
                    target_x = bbox[0][0]  # 获取目标文字左侧的x坐标
                    for (other_bbox, (other_text, other_prob)) in line:
                        if other_bbox[1][0] < target_x and abs(other_bbox[0][1] - bbox[0][1]) < 10:
                            # 在图像上绘制相邻文字区域
                            cv2.rectangle(test_time_cv, other_bbox[0], other_bbox[2], (255, 0, 0), 2)
                            cv2.putText(test_time_cv, other_text, (other_bbox[0][0], other_bbox[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)
                            # 显示图像
                            cv2.imshow("Detected Text", test_time_cv)
                            cv2.waitKey(0)
                            cv2.destroyAllWindows()
                            return other_text.strip()
                elif direction == "up":
                    target_y = bbox[0][1]  # 获取目标文字上方的y坐标
                    for (other_bbox, (other_text, other_prob)) in line:
                        if other_bbox[2][1] < target_y and abs(other_bbox[0][0] - bbox[0][0]) < 10:
                            # 在图像上绘制相邻文字区域
                            cv2.rectangle(test_time_cv, other_bbox[0], other_bbox[2], (255, 0, 0), 2)
                            cv2.putText(test_time_cv, other_text, (other_bbox[0][0], other_bbox[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)
                            # 显示图像
                            cv2.imshow("Detected Text", test_time_cv)
                            cv2.waitKey(0)
                            cv2.destroyAllWindows()
                            return other_text.strip()
                elif direction == "down":
                    target_y = bbox[2][1]  # 获取目标文字下方的y坐标
                    for (other_bbox, (other_text, other_prob)) in line:
                        if other_bbox[0][1] > target_y and abs(other_bbox[0][0] - bbox[0][0]) < 10:
                            # 在图像上绘制相邻文字区域
                            cv2.rectangle(test_time_cv, other_bbox[0], other_bbox[2], (255, 0, 0), 2)
                            cv2.putText(test_time_cv, other_text, (other_bbox[0][0], other_bbox[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)
                            # 显示图像
                            cv2.imshow("Detected Text", test_time_cv)
                            cv2.waitKey(0)
                            cv2.destroyAllWindows()
                            return other_text.strip()
    raise Exception(f"未能识别到{direction}方向的文字")
# 矫正识别结果的错别字，并处理冒号后的文本
def correct_typos(text):
    typo_dict = {
        "所有耸": "所有算法",
        "检测窗口": "检测窗口",
        "未完待续": "未完待续",
        "错件": "错件"
    }
    for typo, correct in typo_dict.items():
        similarity = similar(text, typo)
        if similarity >= 0.66:
            text = text.replace(typo, correct)
    return text

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()
# 检查文字前面的勾选框状态
def check_checkbox_status_before_text(target_text, if_check=True, direction="left", range=25, mark_rgb=(107, 107, 107), frame_rgb=(51, 51, 51), similarity_threshold=0.66):
    logger.info("开始检查文字前的勾选框状态")
    logger.info(f"目标文字: {target_text}, 方向: {direction}, 范围: {range}, 目标颜色: {mark_rgb}, 框架颜色: {frame_rgb}")

    try:
        # 如果target_text含英文的话，similarity_threshold可以-0.15
        if re.search('[a-zA-Z]', target_text):
            similarity_threshold -= 0.25

        # 截取屏幕
        screenshot = pyautogui.screenshot()

        # 将截图转换为OpenCV格式
        screenshot_cv = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)

        # 使用PaddleOCR识别文字
        result = paddleOCR.ocr(screenshot_cv, cls=True)  # 获取详细信息以获取位置

        # 检查是否包含目标文字
        for line in result:
            for (bbox, (text, _)) in line:
                text = correct_typos(text)  # 矫正识别结果的错别字
                similarity = SequenceMatcher(None, target_text, text).ratio()
                if similarity >= similarity_threshold:
                    logger.info(f"识别到与{target_text}相似的文字：{text}，相似度: {similarity}")

                    # 获取文字位置
                    (top_left, top_right, bottom_right, bottom_left) = bbox
                    logger.info(f"文字位置: {bbox}")
                    if direction == "left":
                        x_start, x_end = int(top_left[0] - int(range) + 22), int(top_left[0] + 22)
                        y_start, y_end = int(top_left[1]), int(bottom_left[1])
                    elif direction == "right":
                        x_start, x_end = int(top_right[0] - 22), int(top_right[0] + int(range) - 22)
                        y_start, y_end = int(top_right[1]), int(bottom_right[1])
                    # 只要是蓝色就行
                    color_region = screenshot_cv[y_start:y_end, x_start:x_end]
                    color_found = False
                    logger.info(f"开始在区域 ({x_start}, {y_start}) 到 ({x_end}, {y_end}) 检测蓝色")

                    # 识别蓝色
                    lower_blue = np.array([100, 0, 0])
                    upper_blue = np.array([255, 100, 100])
                    blue_mask = cv2.inRange(color_region, lower_blue, upper_blue)

                    if np.any(blue_mask):
                        logger.info(f"在方向{direction}识别到蓝色，区域: ({x_start}, {y_start}) 到 ({x_end}, {y_end})")
                        color_found = True
                        # 比对和if_check是否一致
                        if (if_check and np.count_nonzero(blue_mask) > 1) or (not if_check and np.count_nonzero(blue_mask) <= 1):
                            # 单击找到的颜色像素点
                            click_x, click_y = np.mean(np.argwhere(blue_mask), axis=0).astype(int)
                            click_x += x_start
                            click_y += y_start
                            logger.info(f"点击位置: ({click_x}, {click_y})")
                            pyautogui.click(x=click_x, y=click_y)
                    if not color_found:
                        logger.info(f"未识别到勾选框的蓝色勾选状态，开始改用框颜色和mark颜色进行检测")

                        logger.info(f"开始检查框架颜色")
                        frame_region = screenshot_cv[y_start:y_end, x_start:x_end]
                        # 支持多种frame_rgb颜色
                        frame_rgbs = [frame_rgb, (51,51,51),(54,54,54),(60,59,57),(66,66,66),(53, 52, 55), (132, 132, 132), (138, 138, 138)]
                        frame_present = False
                        for frgb in frame_rgbs:
                            lower_bound = np.array(frgb) - 5
                            upper_bound = np.array(frgb) + 5
                            mask = cv2.inRange(frame_region, lower_bound, upper_bound)
                            logger.info(f"在区域 ({x_start}, {y_start}) 到 ({x_end}, {y_end}) 识别颜色范围: {lower_bound} 到 {upper_bound}")
                            if np.any(mask):
                                frame_present = True
                                break
                        logger.debug(f"框架颜色存在状态: {frame_present}")

                        if frame_present:
                            # 检查勾选框内部的√mark颜色是否存在
                            mark_region = screenshot_cv[y_start:y_end, x_start:x_end]
                            # 支持多种mark_rgb颜色
                            mark_rgbs = [mark_rgb, (72, 72, 72), (108, 108, 108)]
                            total_mark_count = 0
                            for m_rgb in mark_rgbs:
                                mark_mask = cv2.inRange(mark_region, np.array(m_rgb) - 5, np.array(m_rgb) + 5)  # 缩小颜色范围
                                mark_count = cv2.countNonZero(mark_mask)
                                total_mark_count += mark_count
                                logger.debug(f"勾选框内部的√mark{m_rgb}颜色像素数量: {mark_count}")

                            # 设置一个合理的阈值来判断是否存在
                            if (if_check and total_mark_count > 6) or (not if_check and total_mark_count <= 1):
                                logger.info(f"识别到与{target_text}相似的文字且确认勾选框状态与预期一致，返回True")
                                return True

                            logger.info(f"识别到与{target_text}相似的文字且确认勾选框状态与预期不一致，将在指定区域点击")
                            # 点击操作
                            click_x = (x_start + x_end) // 2
                            click_y = (y_start + y_end) // 2
                            logger.info(f"点击位置: ({click_x}, {click_y})")
                            pyautogui.click(x=click_x, y=click_y)
                            return False
                    else:
                        logger.error(f"在方向{direction}未找到目标颜色的方框")
                elif similarity > 0.5:
                    logger.info(f"识别到与{target_text}相似度大于0.5的文字：{text}，相似度: {similarity}")
    except Exception as e:
        logger.error(f"确认文字前勾选框状态方法发生异常: {e}")
    return False

# 调整将CAD框随机变大，再变小
def adjust_cad_frame():
    search_symbol(config.EDIT_BACK, 10)
    time.sleep(2)
    # 选中框框快捷键
    pyautogui.press('b')
    logger.info("选中检测框")
    time.sleep(1.5)
    success, point = get_choose_box()
    logger.info(point)
    # 识别到完整的选择框
    if success:
        if point is not None:
            logger.info("move")
            x, y = point
            logger.info(x, y)
            pyautogui.moveTo(x, y, duration=0.5)
            pyautogui.mouseDown()
            # 计算区域边界
            left_bound, top_bound, width, height = 540, 150, 1327 - 540, 738 - 150
            right_bound = left_bound + width
            bottom_bound = top_bound + height

            # 计算到边界的最小距离
            min_dist_to_left = point[0] - left_bound
            min_dist_to_right = right_bound - point[0]
            min_dist_to_top = point[1] - top_bound
            min_dist_to_bottom = bottom_bound - point[1]

            # 选择最小距离并计算随机移动距离
            min_dist = min(min_dist_to_left, min_dist_to_right, min_dist_to_top, min_dist_to_bottom)
            move_dist = random.randint(1, min_dist)

            # 确定移动方向
            if min_dist == min_dist_to_left:
                move_x = -move_dist
                move_y = 0
            elif min_dist == min_dist_to_right:
                move_x = move_dist
                move_y = 0
            elif min_dist == min_dist_to_top:
                move_x = 0
                move_y = -move_dist
            else:
                move_x = 0
                move_y = move_dist

            # 执行扩大选择框操作
            pyautogui.moveRel(move_x, move_y, duration=2)
            time.sleep(0.5)
            pyautogui.mouseUp()
            # 最大点
            big_x, big_y = pyautogui.position()
            time.sleep(1)
            # 执行缩小选择框操作
            center_x, center_y = 935, 445
            pyautogui.moveTo(big_x, big_y)
            pyautogui.mouseDown()
            move_to_center_x = random.randint(0, abs(center_x - big_x))
            move_to_center_y = random.randint(0, abs(center_y - big_y))
            if center_x < big_x:
                move_to_center_x = -move_to_center_x
            if center_y < big_y:
                move_to_center_y = -move_to_center_y
            target = (move_to_center_x, move_to_center_y)
            pyautogui.moveRel(target[0], target[1], duration=1)
            pyautogui.mouseUp()
        # TODO 识别到不完整的选择框 估计是选择框出界了
        else:
            pass
    else:
        sys.exit("未识别到选择框")


def random_choose_light():
    # 绝对坐标列表（1920x1080分辨率）
    absolute_coords = [
        (555, 315), (555, 450), (555, 580), (555, 715),
        (680, 315), (680, 450), (680, 580),
        (820, 315), (820, 450), (820, 580),
    ]
    # 获取当前屏幕分辨率
    screen_width, screen_height = pyautogui.size()
    # 将绝对坐标转换为相对坐标
    relative_coords = [
        (int(x * screen_width / 1920), int(y * screen_height / 1080)) for x, y in absolute_coords
    ]
    # 随机选择一个坐标
    chosen_coord = random.choice(relative_coords)
    # 单击选中的坐标
    pyautogui.click(chosen_coord)

def random_open_program():
    # 先检测有没有什么傻逼窗口 关掉
    if search_symbol(config.NO, 1.5):
        click_by_png(config.NO, tolerance=0.9)
    pyautogui.press('enter')
    click_by_png(config.OPEN_PROGRAM)
    if search_symbol(config.NO, 3):
        click_by_png(config.NO, tolerance=0.9)
    time.sleep(3)
    options = [config.OPEN_PROGRAM_CURSOR, config.OPEN_PROGRAM_PLUS]
    found = False
    for option in random.sample(options, len(options)):  # 随机排序尝试
        if search_symbol(option, 2):
            click_by_png(option, 2)
            found = True
            break
    if not found:
        raise Exception("打开程式界面没有搜索到程式")
    time.sleep(1)
    click_by_png(config.YES)
    while search_symbol(config.PROGRAM_LOADING):
        time.sleep(5)
    time.sleep(3)

# 算法参数面板随机改变值(缩小即无效)
def random_change_param(if_test=0):
    # 分为两种改随机的方式:
    # 下拉框为面积内随机选（点击点，输入各个下拉框面积，下拉随机值（限定范围），随机选面积内一点点击，但麻烦）


    # 先点击算法 根据算法点不同的点
    # 使用for循环点击各个点，输入0-500随机数字
    alg_modes = {
        config.CW_COLOR_ANALYSISING: config.CW_COLOR_ANALYSISING_POINTS,
        config.CW_COLOR_MATCHING: config.CW_COLOR_MATCHING_POINTS,
        config.CW_IMAGE_MATCHING: config.CW_IMAGE_MATCHING_POINTS,
        config.CW_COLOR_AREA: config.CW_COLOR_AREA_POINTS,
        config.CW_SQUARE_POSITIONING: config.CW_SQUARE_POSITIONING_POINTS,
        config.CW_PIN_SIMILARITY_MATCHING: config.CW_PIN_SIMILARITY_MATCHING_POINTS,
    }

    found = False
    for mode, points in alg_modes.items():
        if search_symbol(mode, region=config.COMPONENT_WINDOW_REGION):
            click_by_png(mode, region=config.COMPONENT_WINDOW_REGION)
            time.sleep(8)
            found = True
            break

    if not found:
        return False
    # 修改参数
    for point in points:
        pyautogui.doubleClick(point)
        random_number = random.randint(0, 500)
        pyautogui.typewrite(str(random_number))
        time.sleep(0.5)
        pyautogui.press('enter')
        time.sleep(0.5)
        pyautogui.press('enter')
    # 点击测试窗口/元件/分组/整版
    start_time = time.time()
    test_config = {
        1: (config.TEST_WINDOW, 3),
        2: (config.TEST_COMPONENT, 10),
        3: (config.TEST_GROUP, 20),
        4: (config.TEST_BOARD, 20)
    }

    if if_test in test_config:
        click_by_png(test_config[if_test][0], test_config[if_test][1])
        if if_test > 1:  # 对于2, 3, 4的情况需要进行循环检测
            while True:
                if not search_symbol(config.TESTING_COMPONENT, 5, tolerance=0.8):
                    break
                if time.time() - start_time > 600:  # 超过十分钟
                    raise Exception("程序一直处于测试元件中")
                elif time.time() - start_time > 300:  # 超过五分钟
                    for proc in psutil.process_iter(['name']):
                        if "Sinic-Tek" in proc.info['name'] and "NEW" in proc.info['name']:
                            proc.terminate()  # 关闭进程
                time.sleep(5)
        time.sleep(5)
    else:
        return
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.press('enter')
    return True

    
# 添加标准影像的参数面板
def random_change_image_param():
    points = config.IMAGE_PARAM_POINTS
    for point in points:
        time.sleep(1.5)
        pyautogui.doubleClick(point)
        time.sleep(0.5)
        random_number = random.randint(0, 80)
        pyautogui.typewrite(str(random_number))
        time.sleep(0.5)
        pyautogui.press('enter')

# 随机修改rgb值
def random_change_rgb():
    #对(835,825) (835,870) (835,915) (835,950)
    points = [(835, 825), (835, 870), (835, 915), (835, 950)]
    for point in points:
        pyautogui.click(point)
        time.sleep(20)
        pyautogui.hotkey('ctrl', 'a')
        random_number = random.randint(0, 255)
        time.sleep(10)
        pyautogui.typewrite(str(random_number))
    pyautogui.press('enter')
    time.sleep(10)


# ==========================工具类======================
# =========================excel=======================
def adjust_image_size(image_path, max_width, max_height):
    try:
        with PILImage.open(image_path) as img:
            original_width, original_height = img.size
            ratio = min(max_width / original_width, max_height / original_height)
            new_width = int(original_width * ratio)
            new_height = int(original_height * ratio)
            return new_width, new_height
    except Exception as e:
        logger.error(f"Error adjusting image size for {image_path}: {e}")
        return None, None

def insert_image_limited(ws, cell, image_path, max_width=75, max_height=75):
    if os.path.exists(image_path):
        new_width, new_height = adjust_image_size(image_path, max_width, max_height)
        if new_width is None or new_height is None:
            cell.value = "错误"
            return
        img = ExcelImage(image_path)
        img.width = new_width
        img.height = new_height
        img.anchor = cell.coordinate
        ws.add_image(img)
        cell.value = ""
    else:
        cell.value = "空"
        logger.debug(f"图片路径不存在: {image_path}")

# 图片适应屏幕分辨率
def image_fit_screen(image_path):
    # 获取当前屏幕分辨率
    screen_width, screen_height = pyautogui.size()
    if screen_width == 1920 and screen_height == 1080:
        return image_path
    else:
        logger.error("屏幕宽度: {}, 屏幕高度: {}", screen_width, screen_height)
        # 打开图像文件
        img = PILImage.open(image_path)
        # 调整图像大小到当前屏幕分辨率
        img = img.resize((int(img.width * screen_width / 1920), int(img.height * screen_height / 1080)),
                        PILImage.Resampling.LANCZOS)
        # 确定项目根目录
        project_root = os.path.dirname(os.path.abspath(__file__))  # 假设此脚本位于项目根目录
        temp_dir = os.path.join(project_root, 'temp')
        # 如果temp文件夹不存在，则创建它
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        # 构建临时文件路径
        filename = os.path.basename(image_path)  # 获取原始图像的文件名
        temp_image_path = os.path.join(temp_dir, f"temp_{filename}")  # 在temp目录下创建新的文件名
        temp_image_path = os.path.normpath(temp_image_path)
        img.save(temp_image_path)
        return temp_image_path

def points_are_similar(points1, points2, max_error=10):
    if len(points1) != len(points2):
        return False
    for p1, p2 in zip(points1, points2):
        if abs(p1[0] - p2[0]) > max_error or abs(p1[1] - p2[1]) > max_error:
            return False
    return True
# 确保指定位置内是该文本
def text_in_bbox(text, bbox):
    # 点击输入框
    center_x = (bbox[0] + bbox[2]) / 2
    center_y = (bbox[1] + bbox[3]) / 2
    pyautogui.click(center_x, center_y)
    time.sleep(0.5)  # 等待点击生效

    # 全选并复制
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')

    # 读取剪贴板内容
    clipboard_text = pyperclip.paste()

    # 检查剪贴板内容是否包含预期文本
    if text not in clipboard_text:
        pyautogui.typewrite(text)
        pyautogui.press('enter')


def compare_images(img1, img2):
    return ImageChops.difference(img1, img2).getbbox() is None


# ========================处理数据=============================
# 坐标系调整
def adjust_coordinates(top_left, bottom_right, base_resolution=(1920, 1080)):
    # 获取当前屏幕分辨率
    monitor = get_monitors()[0]
    current_resolution = (monitor.width, monitor.height)

    # 计算坐标调整比例
    x_ratio = current_resolution[0] / base_resolution[0]
    y_ratio = current_resolution[1] / base_resolution[1]

    # 调整坐标
    adjusted_top_left = (int(top_left[0] * x_ratio), int(top_left[1] * y_ratio))
    adjusted_bottom_right = (int(bottom_right[0] * x_ratio), int(bottom_right[1] * y_ratio))

    return adjusted_top_left, adjusted_bottom_right

# 检查A、B或C是否大部分存在于all_content中，支持substring内可能有错别字
def contains(substring, full_string, threshold=0.8):
    # 去除换行符并分割成多行
    substring_lines = substring.replace('\n', ' ').split()
    full_string = full_string.replace('\n', ' ')
    
    for line in substring_lines:
        # 先找到最相近的那一段
        best_match = max(full_string.split(), key=lambda x: SequenceMatcher(None, line, x).ratio())
        matcher = SequenceMatcher(None, line, best_match)
        match_ratio = matcher.ratio()
        logger.debug(f"Comparing line: {line} with best_match: {best_match}, match ratio: {match_ratio}")
        if match_ratio >= threshold:
            return True
    
    # 如果没有一行匹配度超过阈值，检查所有行的平均匹配度
    total_ratio = sum(SequenceMatcher(None, line, full_string).ratio() for line in substring_lines) / len(substring_lines)
    logger.debug(f"Total match ratio: {total_ratio}")
    return total_ratio >= threshold

def get_center_coordinates(coord1, coord2):
    """
    计算两个坐标点的中心坐标。
    :param coord1: 第一个坐标点 (x1, y1)
    :param coord2: 第二个坐标点 (x2, y2)
    :return: 中心坐标 (x, y)
    """
    x_center = (coord1[0] + coord2[0]) // 2
    y_center = (coord1[1] + coord2[1]) // 2
    return (x_center, y_center)
def get_frame_points(region, color):
    try:
        logger.info(f"开始截取区域: {region}")
        screenshot = pyautogui.screenshot(region=region)
        screenshot_np = np.array(screenshot)
        screenshot_hsv = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2HSV)

        # 将 RGB 颜色转换为 HSV 颜色
        color_hsv = cv2.cvtColor(np.uint8([[color]]), cv2.COLOR_RGB2HSV)[0][0]
        logger.info(f"目标颜色的HSV值: {color_hsv}")

        # 定义颜色的HSV范围，只找color这个rgb
        lower_bound = np.array([color_hsv[0] - 5, color_hsv[1] - 5, color_hsv[2] - 5])
        upper_bound = np.array([color_hsv[0] + 5, color_hsv[1] + 5, color_hsv[2] + 5])
        logger.info(f"颜色范围: {lower_bound} - {upper_bound}")

        # 创建颜色掩码
        mask = cv2.inRange(screenshot_hsv, lower_bound, upper_bound)

        # 使用形态学操作增强边界
        kernel = np.ones((5, 5), np.uint8)
        mask = cv2.dilate(mask, kernel, iterations=5)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=5)

        # 寻找连贯区域的轮廓
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            raise Exception("未找到指定颜色的框")

        # 调试：显示检测到的轮廓
        # contour_img = cv2.drawContours(screenshot_np.copy(), contours, -1, (0, 255, 0), 2)
        # cv2.imshow("Contours", contour_img)
        # cv2.waitKey(0)

        # 假设只有一个框，获取其四个对角点
        logger.info(f"检测到的框的数量: {len(contours)}")
        contour = max(contours, key=cv2.contourArea)

        # 使用多边形逼近算法
        epsilon = 0.02 * cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, epsilon, True)

        # 获取轮廓的顶点 
        points = approx.reshape(-1, 2)
        points_fullscreen = [(point[0] + region[0], point[1] + region[1]) for point in points]
        logger.debug(points_fullscreen)
        if len(points_fullscreen) == 8:
            merged_points = []
            pairs = [(0, 7), (1, 2), (3, 4), (5, 6)]
            for i, j in pairs:
                x_center = (points_fullscreen[i][0] + points_fullscreen[j][0]) // 2
                y_center = (points_fullscreen[i][1] + points_fullscreen[j][1]) // 2
                merged_points.append((x_center, y_center))
            points_fullscreen = merged_points
        
        # 在屏幕上显示各个点的位置
        if len(points_fullscreen) == 4:
            for point in points_fullscreen:
                cv2.circle(screenshot_np, point, 5, (0, 0, 255), -1)
                cv2.putText(screenshot_np, f"{point}", (point[0] + 10, point[1]), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
            
            # 调整窗口大小以显示所有点
            height, width, _ = screenshot_np.shape
            # cv2.namedWindow("AAA", cv2.WINDOW_NORMAL)
            # cv2.resizeWindow("AAA", width, height)
            # cv2.imshow("AAA", screenshot_np)
            # cv2.waitKey(0)
        else:
            logger.error("未能检测到四个顶点")
        return points_fullscreen
    except Exception as e:
        logger.error(f"获取框的坐标时发生错误: {e}")
        raise

def write_text(coordinate, text, if_select_all=True, if_press_enter=False):
    pyautogui.click(coordinate)
    time.sleep(0.5)
    if if_select_all:   
        pyautogui.click(coordinate, clicks=2, interval=0.1)
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.5)
    
    # 确保输出内容和text大小写一致
    if text.isupper():
        text_to_type = text.upper()
    elif text.islower():
        text_to_type = text.lower()
    else:
        text_to_type = text  # 保持原样

    # 输入文本
    logger.info(f"开始输入文本: {text_to_type}")
    # pyautogui.typewrite(text_to_type)  #输入中文会有问题
    pyperclip.copy(text_to_type)
    pyautogui.hotkey('ctrl', 'v')
    
    if if_press_enter:
        pyautogui.press('enter')

def write_text_textbox(image_path=None, ocr_text=None, write_content=None, tolerance=0.75, color=(255, 255, 255), direction="right", locate_direction="left", if_select_all=True, if_press_enter=False):
    try:
        if image_path:
            logger.debug(f"开始处理图片: {image_path}, 方向: {direction}, 颜色: {color}, 识别方向: {locate_direction}")
            time.sleep(2)
            
            # 获取图片在屏幕上的位置
            symbol = image_fit_screen(image_path)
            logger.info(f"开始寻找 {symbol}")
            
            try:
                locations = list(pyautogui.locateAllOnScreen(symbol, confidence=tolerance))
                if locations:
                    if locate_direction == "left":
                        location = min(locations, key=lambda loc: loc.left)
                    elif locate_direction == "right":
                        location = max(locations, key=lambda loc: loc.left)
                    elif locate_direction == "top":
                        location = min(locations, key=lambda loc: loc.top)
                    elif locate_direction == "bottom":
                        location = max(locations, key=lambda loc: loc.top)
                    logger.debug(f"图片位置: {location}")
                else:
                    logger.error("未找到指定的图片")
                    raise Exception("未找到指定的图片")
            except pyautogui.ImageNotFoundException:
                logger.error("未找到指定的图片")
                raise Exception("未找到指定的图片")
            
            except Exception as e:
                logger.error(f"发生异常: {e}")
                raise Exception(f"发生异常: {e}")
        
            # 获取颜色为 color 的区域，基于 location 的中心点，按方向搜索
            x, y, width, height = map(int, [location.left, location.top, location.width, location.height])
        
        if ocr_text:
            x, y, width, height = get_text_coordinate(ocr_text, tolerance)
            logger.info(f"识别文字的x,y,width,height: ({x}, {y}, {width}, {height})")
        
        # 根据方向设置搜索区域，确保范围合适
        if direction == "right":
            # 向右：从 ocr_text 的右边界开始，限制为上下边界范围
            search_x = x + width
            logger.info(f"方向: 向右, 从 ocr_text 的右边界开始, 限制为上下边界范围, 坐标: ({search_x}, {y}, {1920 - search_x}, {height})")
            screenshot = pyautogui.screenshot(region=(int(search_x), int(y), int(1920 - search_x), int(height)))
            logger.info(f"轮廓左上角和右下角全屏坐标: ({search_x}, {y}) 到 ({1920}, {y + height})")
        elif direction == "left":
            # 向左：从 ocr_text 的左边界开始，限制为上下边界范围
            search_x = 0
            logger.info(f"方向: 向左, 从 ocr_text 的左边界开始, 限制为上下边界范围, 坐标: ({search_x}, {y}, {x}, {height})")
            screenshot = pyautogui.screenshot(region=(int(search_x), int(y), int(x), int(height)))
            logger.info(f"轮廓左上角和右下角全屏坐标: ({search_x}, {y}) 到 ({x}, {y + height})")
        elif direction == "above":
            # 向上：从 ocr_text 的上边界开始，限制为左右边界范围
            search_y = 0
            logger.info(f"方向: 向上, 从 ocr_text 的上边界开始, 限制为左右边界范围, 坐标: ({x}, {search_y}, {width}, {y})")
            screenshot = pyautogui.screenshot(region=(int(x), int(search_y), int(width), int(y)))
            logger.info(f"轮廓左上角和右下角全屏坐标: ({x}, {search_y}) 到 ({x + width}, {y})")
        elif direction == "below":
            # 向下：从 ocr_text 的下边界开始，限制为左右边界范围
            search_y = y + height
            logger.info(f"方向: 向下, 从 ocr_text 的下边界开始, 限制为左右边界范围, 坐标: ({x}, {search_y}, {width}, {1080 - search_y})")
            screenshot = pyautogui.screenshot(region=(int(x), int(search_y), int(width), int(1080 - search_y)))
            logger.info(f"轮廓左上角和右下角全屏坐标: ({x}, {search_y}) 到 ({x + width}, 1080)")
        else:
            logger.error("无效的方向参数")
            raise Exception("无效的方向参数")

        # 转换截图为 numpy 数组并进行颜色匹配
        screenshot_np = np.array(screenshot)
        lower_bound = np.array(color) - 0
        upper_bound = np.array(color) + 0
        mask = cv2.inRange(screenshot_np, lower_bound, upper_bound)
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            logger.error("未找到文本框")
            raise Exception("未找到文本框")
        logger.debug(f"找到的轮廓数量: {len(contours)}")

        # 过滤掉较小的轮廓
        min_contour_area = 80  # 设置最小轮廓面积阈值
        contours = [contour for contour in contours if cv2.contourArea(contour) > min_contour_area]
        if not contours:
            logger.error("未找到符合条件的颜色区域(文本框)")
            raise Exception("未找到符合条件的颜色区域(文本框)")
        logger.debug(f"过滤后的轮廓数量: {len(contours)}")

        # # 可视化调试：在屏幕上显示所有识别到的轮廓
        # debug_image = screenshot_np.copy()
        # cv2.drawContours(debug_image, contours, -1, (0, 255, 0), 2)
        # cv2.imshow("All Contours", debug_image)
        # cv2.waitKey(0)

        # # 保存识别到轮廓后的截图
        # cv2.imwrite("contours_screenshot.png", debug_image)
        # logger.info("识别到轮廓后的截图已保存为 contours_screenshot.png")

        # 获取离标识最近的颜色为 color 的区域
        def get_distance_to_center(contour):
            M = cv2.moments(contour)
            if M["m00"] == 0:
                return float('inf')
            contour_center_x = int(M["m10"] / M["m00"])
            contour_center_y = int(M["m01"] / M["m00"])
            return contour_center_x, contour_center_y

        # 记录所有轮廓的中心点和距离标识中心点的距离
        for contour in contours:
            contour_center_x, contour_center_y = get_distance_to_center(contour)
            if direction in ["right", "left"]:
                screen_contour_center_x = search_x + contour_center_x
            else:
                screen_contour_center_x = x + contour_center_x
            if direction in ["above", "below"]:
                screen_contour_center_y = search_y + contour_center_y
            else:
                screen_contour_center_y = y + contour_center_y
            logger.debug(f"文本框中心点: ({screen_contour_center_x}, {screen_contour_center_y})")

        # 查找离中心点300范围内的颜色区域
        if direction in ["right", "left"]:
            valid_contours = [c for c in contours if abs(get_distance_to_center(c)[0]) <= 300]
            if not valid_contours:
                raise Exception("未找到300范围内的颜色区域")
            if direction == "right":
                closest_contour = min(valid_contours, key=lambda c: get_distance_to_center(c)[0])
            else: 
                closest_contour = max(valid_contours, key=lambda c: get_distance_to_center(c)[0])
        elif direction in ["above", "below"]:
            valid_contours = [c for c in contours if abs(get_distance_to_center(c)[1]) <= 300]
            if not valid_contours:
                raise Exception("未找到300范围内的颜色区域")
            if direction == "above":
                closest_contour = max(valid_contours, key=lambda c: get_distance_to_center(c)[1])
            else:
                closest_contour = min(valid_contours, key=lambda c: get_distance_to_center(c)[1])

        # 获取轮廓B的中心点
        M = cv2.moments(closest_contour)
        if M["m00"] != 0:
            contour_center_x = int(M["m10"] / M["m00"])
            contour_center_y = int(M["m01"] / M["m00"])
            
            # 根据相对位置计算在全屏中的坐标
            if direction in ["right", "left"]:
                full_screen_center_x = search_x + contour_center_x
            else:
                full_screen_center_x = x + contour_center_x
            if direction in ["above", "below"]:
                full_screen_center_y = search_y + contour_center_y
            else:
                full_screen_center_y = y + contour_center_y
            logger.info(f"识别到的文本框中心点的全屏坐标: ({full_screen_center_x}, {full_screen_center_y})")

            # # 可视化调试：在屏幕上显示选择的文本框
            # debug_image_selected = screenshot_np.copy()
            # cv2.drawContours(debug_image_selected, [closest_contour], -1, (255, 0, 0), 2)
            # cv2.imshow("Selected Contour", debug_image_selected)
            # cv2.waitKey(0)

            # 检查距离是否超过300
            if direction in ["right", "left"]:
                distance = abs(contour_center_x)
            else:
                distance = abs(contour_center_y)

            if distance > 300:
                detected_color = screenshot_np[contour_center_y, contour_center_x]
                logger.error(f"找到的区域与标识距离超过300, 坐标: ({full_screen_center_x}, {full_screen_center_y}), 颜色: {detected_color}")
                raise Exception("找到的区域与标识距离超过300")

            if write_content:
                # 将文本写入该区域B的中心点
                write_text((full_screen_center_x, full_screen_center_y), write_content, if_select_all, if_press_enter)
                logger.info(f"写入完毕，内容为：{write_content}")
            else:
                # 如果没有text的话 就单击该坐标
                pyautogui.click(full_screen_center_x, full_screen_center_y)
                logger.info(f"没有提供文本，单击坐标: ({full_screen_center_x}, {full_screen_center_y})")
        else:
            raise Exception("未找到合适的颜色区域")
    except Exception as e:
        logger.error(f"写文字到指定文本框时发生错误: {e}")


def read_text(*args):
    # 检查参数数量和类型
    if len(args) == 1 and isinstance(args[0], tuple):
        x, y = args[0]
    elif len(args) == 2:
        x, y = args
    else:
        raise ValueError("read_text参数必须是一个包含两个元素的元组，或者两个单独的数值。")

    # 获取当前屏幕分辨率
    monitor = get_monitors()[0]
    current_resolution = (monitor.width, monitor.height)

    # 计算坐标调整比例
    x_ratio = current_resolution[0] / 1920
    y_ratio = current_resolution[1] / 1080

    # 调整坐标
    adjusted_x = int(x * x_ratio)
    adjusted_y = int(y * y_ratio)

    # 单击调整后的坐标
    pyautogui.click(adjusted_x, adjusted_y)

    # 全选并复制
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('ctrl', 'c')

    # 读取剪贴板内容
    clipboard_text = pyperclip.paste()

    return clipboard_text


# 由于部分控件无法直接ctrl+a 被迫用这种方式去读取内容
def read_text_choosed(x, y):
    # 左键按住x，y的坐标点
    pyautogui.moveTo(x, y)
    pyautogui.mouseDown()

    # 鼠标拖到指定位置后松开
    pyautogui.moveTo(1900, 1025)
    pyautogui.mouseUp()

    # 复制到剪切板
    pyautogui.hotkey('ctrl', 'c')


# 点击选中颜色
def click_color(times=1, region=None, target_color_rgb=(239, 240, 242), right_click=0, direction=None, offset=True, if_center=False):
    # 截取区域内的屏幕图像
    screenshot = pyautogui.screenshot(region=region)
    screenshot_np = np.array(screenshot)
    # 寻找颜色匹配的像素点，允许RGB上下偏差5
    lower_bound = np.array([c - 15 for c in target_color_rgb])
    upper_bound = np.array([c + 15 for c in target_color_rgb])
    matches = np.where(np.all((screenshot_np >= lower_bound) & (screenshot_np <= upper_bound), axis=-1))
    if matches[0].size > 0:
        if if_center:
            # 计算所有匹配点的中点
            min_x, max_x = np.min(matches[1]), np.max(matches[1])
            min_y, max_y = np.min(matches[0]), np.max(matches[0])
            region_center_x = (min_x + max_x) // 2
            region_center_y = (min_y + max_y) // 2
            region_center_x += region[0]
            region_center_y += region[1]
            logger.info(f"区域中心点: ({region_center_x}, {region_center_y})")
            distances = np.sqrt((matches[1] - region_center_x) ** 2 + (matches[0] - region_center_y) ** 2)
            index = np.argmin(distances)
            # 获取的x y为内边缘的点
            x, y = matches[1][index], matches[0][index]
            # 转换为全屏坐标
            x += region[0]
            y += region[1]
            logger.info(f"原始坐标: ({x}, {y})")
            # 向区域中心偏移1单位
            x += 1 if x < region_center_x else -1
            y += 1 if y < region_center_y else -1
            logger.info(f"新坐标: ({x}, {y})")
        elif direction:
            if direction == 'up':
                index = np.argmin(matches[0])
            elif direction == 'down':
                index = np.argmax(matches[0])
            elif direction == 'left':
                index = np.argmin(matches[1])
            elif direction == 'right':
                index = np.argmax(matches[1])
            x, y = matches[1][index], matches[0][index]
            # 转换为全屏坐标
            x += region[0]
            y += region[1]
        else:
            # 随机获取一个匹配的点的坐标
            random_index = random.randint(0, len(matches[0]) - 1)
            x, y = matches[1][random_index], matches[0][random_index]
            # 转换为全屏坐标
            x += region[0]
            y += region[1]
        
        logger.info("找到目标点"+str(x)+","+str(y))
        # 如果times为0，则只移动鼠标不点击
        if times == 0:
            pyautogui.moveTo(x, y)
            return True
        # 点击指定次数
        for _ in range(times):
            # 获取目标点周围的随机一个像素点
            if offset:
                offset_x = random.randint(-1, 1)
                offset_y = random.randint(-1, 1)
            else:
                offset_x = 0
                offset_y = 0
            click_x = x + offset_x
            click_y = y + offset_y
            if right_click == 1:
                pyautogui.rightClick(click_x, click_y)
            else:
                pyautogui.click(click_x, click_y)
            logger.info(f"点击坐标: ({click_x}, {click_y})")
        return True
    logger.info("未找到匹配的像素点")
    return False
# 计算坐标点附近range像素点范围内 颜色为color的像素点个数
def count_color_in_range(coordinate, range, color):
    screenshot = pyautogui.screenshot()
    screenshot_np = np.array(screenshot)
    start_x, start_y = coordinate[0] - range, coordinate[1] - range
    end_x, end_y = coordinate[0] + range, coordinate[1] + range
    matches = np.where(np.all(screenshot_np[start_y:end_y, start_x:end_x] == color, axis=-1))
    logger.info(f"找到的像素点数量: {len(matches[0])}")
    return len(matches[0])
# 获取区域内颜色占比
def get_color_ratio_in_region(region, target_color_rgb):
    screenshot = pyautogui.screenshot(region=region)
    screenshot_np = np.array(screenshot)
    total_pixels = screenshot_np.shape[0] * screenshot_np.shape[1]
    matches = np.where(np.all(screenshot_np == target_color_rgb, axis=-1))
    color_pixels = len(matches[0])
    color_ratio = color_pixels / total_pixels
    logger.info(f"区域内颜色占比: {color_ratio:.2%}")
    return color_ratio


# 获取鼠标指针处的颜色
def get_mouse_color(mouse_coordinate):
    screenshot = pyautogui.screenshot()
    screenshot_np = np.array(screenshot)
    target_color_rgb = screenshot_np[mouse_coordinate[1], mouse_coordinate[0]]
    return target_color_rgb

# 比较两张图片是否相同（转为灰度图像）
def compare_images(image1, image2):
    # 将截图转换为灰度图像
    image1_gray = image1.convert('L')
    image2_gray = image2.convert('L')

    # 使用ImageChops.difference来比较图像差异
    diff = ImageChops.difference(image1_gray, image2_gray)

    # 如果差异图像的极值为0，说明两张图像相同
    return diff.getbbox() is None


def read_text_ocr(top_left_point, bottom_right_point):
    logger.info("开始识别文字")
    # 截取屏幕上指定区域的图像，并打印截图区域
    bbox = (top_left_point[0], top_left_point[1], bottom_right_point[0], bottom_right_point[1])
    screenshot = ImageGrab.grab(bbox=bbox)

    # 使用PaddleOCR进行文字识别
    screenshot_np = np.array(screenshot)
    screenshot_cv = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2BGR)
    results = paddleOCR.ocr(screenshot_cv, cls=True)  # 使用PaddleOCR进行识别

    # 安全地提取文本，考虑到可能的不同元组长度
    recognized_text = ''.join([text for line in results for (bbox, (text, prob)) in line])

    # 使用correct_typos函数矫正识别结果
    corrected_text = correct_typos(recognized_text)
    logger.info(f"识别的文字: {corrected_text}")

    return corrected_text


# 在区域内下滚，直到滚到底（鼠标位置，识别区域）
def scroll_down(coordinate, region=None, scroll_amount=None, wait_time=0.1):
    if scroll_amount is not None:
        # 将scroll_amount转换为整数
        scroll_amount = int(scroll_amount)
        
        # 如果有传入入参，只能一百一百的滚动
        if scroll_amount % 100 != 0:
            raise ValueError("scroll_amount必须是100的倍数")
        pyautogui.moveTo(coordinate)
        for _ in range(abs(scroll_amount) // 100):
            pyautogui.scroll(100 if scroll_amount > 0 else -100)
            time.sleep(wait_time)
        return
    else:
        scroll_amount = -100  # 修改默认滚动量为-100
        pyautogui.moveTo(coordinate)
        retry_count = 0  # 初始化重试计数器
        while True:
            # 截取初始区域截图，如果region为None则截图全屏
            region_last_time = pyautogui.screenshot(region=region) if region else pyautogui.screenshot()
            # 下滚指定单位
            pyautogui.scroll(scroll_amount)
            time.sleep(wait_time)  # 等待滚动完成
            
            # 截取当前区域截图，如果region为None则截图全屏
            region_current_time = pyautogui.screenshot(region=region) if region else pyautogui.screenshot()
            if compare_images(region_current_time, region_last_time):
                retry_count += 1
                if retry_count >= 5:  # 如果连续5次截图相同，则认为滚动到底
                    return True
            else:
                retry_count = 0  # 重置重试计数器
                region_last_time = region_current_time




# 下拉框，选择所需要的图/文字(参数名意思：图 0 /文字 1 图片路径/识别的文字  区域)
def drop_down_box_search(type, content, region, click_times=1, timeout=6, similarity_threshold=0.8):
    # 根据类型选择操作
    if type == 0:
        # 搜索并点击图片
        start_time = time.time()
        while time.time() - start_time < timeout:
            found = search_symbol(content, region=region, timeout=1)
            if found:
                click_by_png(content, region=region, times=click_times)
                pyautogui.press("enter")
                return True
            else:
                # 鼠标移到region最右侧中间
                right_edge_x = region[0] + region[2]
                middle_y = region[1] + region[3] / 2
                pyautogui.moveTo(right_edge_x, middle_y)
                pyautogui.scroll(-200)  # 再下滚
        return False
    elif type == 1:
        # 使用PaddleOCR搜索并点击文字
        start_time = time.time()
        while time.time() - start_time < timeout:
            # 截取指定区域的屏幕
            screenshot = pyautogui.screenshot(region=region)
            results = paddleOCR.ocr(np.array(screenshot), cls=True)
            # 遍历识别结果，寻找目标文字
            for line in results:
                for word_info in line:
                    bbox, (text_found, prob) = word_info
                    if similar(content, text_found) >= similarity_threshold:
                        top_left = bbox[0]
                        bottom_right = bbox[2]
                        x = (top_left[0] + bottom_right[0]) / 2
                        y = (top_left[1] + bottom_right[1]) / 2

                        full_x = region[0] + x
                        full_y = region[1] + y

                        for _ in range(click_times):
                            pyautogui.click(full_x, full_y)
                        pyautogui.press("enter")
                        return True
            # 如果没有找到匹配的文字，将鼠标移到区域的最右侧并向下滚动200单位
            right_edge_x = region[0] + region[2]
            middle_y = region[1] + region[3] / 2
            pyautogui.moveTo(right_edge_x, middle_y)
            pyautogui.scroll(-200)
        return False

# =====================================文件处理=====================================
# 删除文件夹下所有文件
def delete_documents(path):
    for filename in os.listdir(path):
        file_path = os.path.join(path, filename)
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
        elif os.path.isdir(file_path):
            shutil.rmtree(file_path)
def bring_aoi_to_top(exe_name="AOI"):
    """
    尝试将指定的 exe 进程窗口置顶。

    参数：
        exe_name: 目标进程的名称，默认值为 "AOI.exe"

    返回：
        True 表示成功置顶，False 表示未找到或操作失败。
    """

    logger.info(f"开始执行 bring_aoi_to_top，目标进程: {exe_name}")

    def get_all_hwnd(hwnd, mouse):
        if (win32gui.IsWindow(hwnd) and
            win32gui.IsWindowEnabled(hwnd) and
            win32gui.IsWindowVisible(hwnd)):
            hwnd_map.update({hwnd: win32gui.GetWindowText(hwnd)})

    hwnd_map = {}
    win32gui.EnumWindows(get_all_hwnd, 0)

    for h, t in hwnd_map.items():
        if t and exe_name in t:
            try:
                logger.info(f"找到目标窗口: {t}")
                win32gui.BringWindowToTop(h)
                shell = win32com.client.Dispatch("WScript.Shell")
                shell.SendKeys('%')
                win32gui.SetForegroundWindow(h)
                win32gui.ShowWindow(h, win32con.SW_RESTORE)
                logger.info(f"已将 {exe_name} 的窗口置顶")
                return True
            except Exception as e:
                logger.error(f"无法将窗口置顶: {e}")
                return False

    logger.warning(f"未找到 {exe_name} 进程的窗口")
    return False