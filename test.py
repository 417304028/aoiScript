import pyautogui
import time
import utils, config
import re
import os
import traceback
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference, Series

directory = r'd:\work\temp'
excel_path = os.path.join(directory, f"汇总.xlsx")

def parse_log_file(log_file_path, log_type):
    data = {}
    
    # 根据日志类型初始化不同的数据结构
    if log_type == 'solve_simpCalTime':
        data = {
            '运行JOB名称': [],
            '训练时间-TrainTime(S)': [],
            '计算时间-AlgCalTime(S)': [],  # 调整顺序：计算时间放在第三列
            '拍照时间-CapFovTime(S)': [],  # 拍照时间放在第四列
            '总运行时间-AlgRunTime(S)': [],
            'ThdS': [],  # 添加ThdS字段用于匹配
        }
    elif log_type == 'solve_threadTimeStats':
        data = {
            '运行JOB名称': [],  # 对应日志中的JobName字段
            '在线离线': [],     # 对应日志中的PCBInspect_Mode字段
            '测试版本': [],     # 对应日志中的AppVersion字段
            '线程数': [],       # 统计该job对应有几个Thd_x线程
            '测试元件数量': [], # 所有线程处理的元件总数，即CompNum总和
            '计算时间最低值': [], # 所有线程中计算时间的最小值
            '计算时间最高值': [], # 所有线程中计算时间的最大值
            '计算时间平均值(去掉最高和最低值取平均)': [], # (所有计算时间-最高值-最低值)/(线程数-2)
            '所有PCBID计算时间': [], # 所有线程的计算时间列表或ThdNum值
            'ThdS': [],         # 线程开始时间
            '训练时间-TrainTime(S)': [], # 从solve_simpCalTime日志获取的训练时间
            '计算时间-AlgCalTime(S)': [], # 从solve_simpCalTime日志获取的计算时间
            '拍照时间-CapFovTime(S)': [], # 从solve_simpCalTime日志获取的拍照时间
            '总运行时间-AlgRunTime(S)': [], # 从solve_simpCalTime日志获取的总运行时间
        }
    elif log_type == 'solve_alg3DCrr':
        data = {
            '表格1-汇总按PCBID和窗口': [],  # 汇总按PCBID和窗口生成所有数据格式
            '表格2-按窗口生成所有没有CrrAfter': [],  # 按窗口生成所有没有CrrAfter数据格式
            '表格3-按照array级窗口合并': [],  # 去掉PCBID，按照array级窗口合并
            '表格4-按照元件级窗口合并': []   # 去掉PCBID和Array，按照元件级窗口合并
        }
        # 用于存储合并数据的字典
        array_data = {}  # 用于表格3
        part_data = {}   # 用于表格4
        
        # 用于存储没有CrrAfter的PCBID和窗口
        no_after_data = {}  # 格式: {(PCBID, ArrayID, PartDesignate, AlgID): True}
    elif log_type == 'solve_algInspection':
        data = {
            '表格1-汇总按PCBID和窗口': [],  # 根据PCBID和WinName分组
            '表格2-按窗口生成所有没有After': [],  # 判断该行是否不存在After
            '表格3-按照array级窗口合并': [],  # 去除PCBID，按array合并
            '表格4-按照元件级窗口合并': [],  # 去掉PCBID和Array，按PartDes合并
            '表格5-不含After': [],  # 不含After的数据
            '表格6-含After': []  # 含After的数据
        }
        
        # 用于存储合并数据的字典
        array_data = {}  # 用于表格3
        part_data = {}   # 用于表格4
        before_after_pairs = {}
        # 用于存储表格6数据的字典
        after_data = {}  # 格式: {(ArrayNo, PartDes, AlgID, WinName): {PCBID: After值}}
        pcbid_set = set()  # 存储所有出现的PCBID
    with open(log_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        
        if log_type == 'solve_simpCalTime':
            i = 0
            while i < len(lines) - 4:  # 确保有足够的行来处理一条完整记录
                pcbid_match = re.search(r'PCBID:(\d+)', lines[i])
                if pcbid_match:
                    pcbid = pcbid_match.group(1)
                    
                    # 提取时间戳，用于与线程日志匹配
                    timestamp = None
                    if len(pcbid) >= 14:  # 确保PCBID格式为YYYYMMDDHHMMSS
                        timestamp = f"{pcbid[8:10]}:{pcbid[10:12]}:{pcbid[12:14]}"
                    
                    # 检查接下来的4行是否包含所需的信息
                    train_match = re.search(r'TrainTime\(S\):([0-9.]+)', lines[i+1])
                    cap_match = re.search(r'CapFovTime\(S\):([0-9.]+)', lines[i+2])
                    alg_match = re.search(r'AlgCalTime\(S\):([0-9.]+)', lines[i+3])
                    run_match = re.search(r'AlgRunTime\(S\):([0-9.]+)', lines[i+4])
                    
                    if train_match and cap_match and alg_match and run_match:
                        data['运行JOB名称'].append(pcbid)
                        data['训练时间-TrainTime(S)'].append(float(train_match.group(1)))
                        data['计算时间-AlgCalTime(S)'].append(float(alg_match.group(1)))
                        data['拍照时间-CapFovTime(S)'].append(float(cap_match.group(1)))
                        data['总运行时间-AlgRunTime(S)'].append(float(run_match.group(1)))
                        data['ThdS'].append(timestamp)  # 保存时间戳用于匹配
                        
                    i += 5  # 跳过已处理的5行
                else:
                    i += 1  # 如果没有找到PCBID，前进一行
            
        elif log_type == 'solve_threadTimeStats':
            # 提取PCBID和PCBInspect_Mode信息
            pcbid_match = re.search(r'PCBID_(\d+)', ''.join(lines))
            mode_match = re.search(r'PCBInspect_Mode:(\w+)', ''.join(lines))
            
            # 修改正则表达式以正确提取JobName，不包含AppVersion
            job_match = re.search(r'JobName=([^\s]+(?:\s+[^\s]+)*?)(?:\s+AppVersion=|$)', ''.join(lines))
            version_match = re.search(r'AppVersion=([^\s]+)', ''.join(lines))
            
            pcbid = pcbid_match.group(1) if pcbid_match else "未知"
            pcb_mode = mode_match.group(1) if mode_match else "未知"
            job_name = job_match.group(1).strip() if job_match else pcbid  # 如果没有JobName，使用PCBID
            app_version = version_match.group(1) if version_match else "未知"
            
            # 打印提取的JobName和AppVersion，用于调试
            print(f"提取的JobName: '{job_name}'")
            print(f"提取的AppVersion: '{app_version}'")
            
            # 提取线程数和组件数量
            thread_count = 0
            comp_nums = []
            calc_times = []
            thd_s = ""
            
            for line in lines:
                match = re.search(r'Thd_(\d+):ThdS:(\d{1,2}:\d{1,2}:\d{1,2}),CalS:(\d{1,2}:\d{1,2}:\d{1,2}),CalE:(\d{1,2}:\d{1,2}:\d{1,2}),Time:\s*(\d+)\s*/\s*(\d+).*?CalWinTime:(\d+),CompNum:(\d+)', line)
                if match:
                    thread_count = max(thread_count, int(match.group(1)) + 1)  # 线程从0开始计数
                    thd_s = match.group(2)
                    calc_time = int(match.group(6))  # 使用 / 后面的值
                    comp_num = int(match.group(8))
                    
                    calc_times.append(calc_time)
                    comp_nums.append(comp_num)
            
            # 提取ThdNum值
            thdnum_match = re.search(r'ThdNum:(\d+)', ''.join(lines))
            thdnum = thdnum_match.group(1) if thdnum_match else ""
            
            # 计算统计值
            if calc_times:
                min_time = min(calc_times)
                max_time = max(calc_times)
                
                # 计算平均值(去掉最高和最低值)
                if len(calc_times) > 2:
                    # 排除最高值和最低值后计算平均值
                    sorted_times = sorted(calc_times)
                    avg_time = sum(sorted_times[1:-1]) / (len(sorted_times) - 2)
                else:
                    # 如果线程数不足3个，则直接计算平均值
                    avg_time = sum(calc_times) / len(calc_times)
                
                # 格式化所有PCBID计算时间为逗号分隔的列表
                all_times_str = ', '.join([str(t) for t in calc_times])
                
                # 填充数据
                data['运行JOB名称'].append(job_name)
                data['在线离线'].append(pcb_mode)
                data['测试版本'].append(app_version)
                data['线程数'].append(thread_count)
                data['测试元件数量'].append(sum(comp_nums))
                data['计算时间最低值'].append(min_time)
                data['计算时间最高值'].append(max_time)
                data['计算时间平均值(去掉最高和最低值取平均)'].append(round(avg_time, 2))
                data['所有PCBID计算时间'].append(all_times_str)
                data['ThdS'].append(thd_s)
                
                # 这些字段可能需要从其他日志中获取，暂时留空
                data['训练时间-TrainTime(S)'].append("")
                data['计算时间-AlgCalTime(S)'].append("")
                data['拍照时间-CapFovTime(S)'].append("")
                data['总运行时间-AlgRunTime(S)'].append("")
            
        elif log_type == 'solve_alg3DCrr':
            for line in lines:
                # 匹配包含CrrAfter或CrrBefore的行
                match = re.search(r'PCBID_(\w+)_Tdx_(\d+),.*?ArrayID_(\w+)_PartDesignate_(\w+)_AlgID_(\w+)_OperSignal_(\w+)_ChipType_(\w+)，BWinCount：(\d+):(?:CrrBefore|CrrAfter):(\w+)(?:;iRet:\d+,3DTime:(\d+))?', line)
                if match:
                    pcbid = f"PCBID_{match.group(1)}_Tdx_{match.group(2)}"
                    array_id = match.group(3)
                    part_designate = match.group(4)
                    alg_id = match.group(5)
                    oper_signal = match.group(6)
                    chip_type = match.group(7)
                    bwin_count = int(match.group(8))
                    crr_status = match.group(9)  # True或False
                    three_d_time = int(match.group(10)) if match.group(10) else 0
                    
                    # 判断是CrrBefore还是CrrAfter
                    is_after = "CrrAfter" in line
                    crr_value = 0  # 默认值
                    
                    # 如果是CrrAfter行，提取3DTime作为CrrAfter值
                    if is_after:
                        crr_value = three_d_time
                    
                    # 表格1：汇总所有数据
                    entry = {
                        'PCBID_No': pcbid,
                        'OperSignal': oper_signal,
                        'ArrayID': array_id,
                        'ChipType': chip_type,
                        'PartDesignate': part_designate,
                        'AlgID': alg_id,
                        'BWinCount': bwin_count,
                        'CrrBefore': 0,  # 默认值，可能需要从其他行获取
                        'CrrAfter': crr_value if is_after else None
                    }
                    
                    # 只有包含完整信息的行才添加到表格1
                    if is_after:
                        data['表格1-汇总按PCBID和窗口'].append(entry)
                    
                    # 表格2：没有CrrAfter的数据
                    # 修改这部分逻辑，正确处理没有CrrAfter的情况
                    key = (pcbid, array_id, part_designate, alg_id)
                    
                    if not is_after and crr_status == "False":
                        # 记录没有CrrAfter的窗口
                        no_after_data[key] = {
                            'PCBID_No': pcbid,
                            'OperSignal': oper_signal,
                            'ArrayID': array_id,
                            'ChipType': chip_type,
                            'PartDesignate': part_designate,
                            'AlgID': alg_id,
                            'BWinCount': bwin_count,
                            'CrrStatus': crr_status
                        }
                    elif is_after:
                        # 如果找到了CrrAfter，从no_after_data中移除
                        if key in no_after_data:
                            del no_after_data[key]
                    
                    # 为表格3和表格4准备数据
                    if is_after:
                        # 表格3：按array合并
                        array_key = f"{array_id}_{chip_type}_{part_designate}_{alg_id}"
                        if array_key not in array_data:
                            array_data[array_key] = {
                                'ArrayID': array_id,
                                'ChipType': chip_type,
                                'PartDesignate': part_designate,
                                'AlgID': alg_id,
                                'BWinCount': bwin_count,
                                'low': float('inf'),  # 初始化为无穷大
                                'high': 0,
                                'average': 0,
                                'values': [],
                                '所有CrrAfter值': []
                            }
                        
                        # 更新最低值、最高值和所有值
                        array_data[array_key]['low'] = min(array_data[array_key]['low'], crr_value)
                        array_data[array_key]['high'] = max(array_data[array_key]['high'], crr_value)
                        array_data[array_key]['values'].append(crr_value)
                        array_data[array_key]['所有CrrAfter值'].append(str(crr_value))
                        
                        # 表格4：按元件合并
                        part_key = f"{part_designate}_{alg_id}"
                        if part_key not in part_data:
                            part_data[part_key] = {
                                'PartDesignate': part_designate,
                                'AlgID': alg_id,
                                'BWinCount': bwin_count,
                                'low': float('inf'),
                                'high': 0,
                                'average': 0,
                                'values': [],
                                '所有CrrAfter值': []
                            }
                        
                        # 更新最低值、最高值和所有值
                        part_data[part_key]['low'] = min(part_data[part_key]['low'], crr_value)
                        part_data[part_key]['high'] = max(part_data[part_key]['high'], crr_value)
                        part_data[part_key]['values'].append(crr_value)
                        part_data[part_key]['所有CrrAfter值'].append(str(crr_value))
            
            # 将没有CrrAfter的数据添加到表格2
            for key, entry in no_after_data.items():
                data['表格2-按窗口生成所有没有CrrAfter'].append(entry)
            
            # 处理表格3的数据
            for key, value in array_data.items():
                if value['values']:
                    value['average'] = sum(value['values']) / len(value['values'])
                    value['所有CrrAfter值'] = ', '.join(value['所有CrrAfter值'])
                    # 删除临时字段
                    del value['values']
                    data['表格3-按照array级窗口合并'].append(value)
            
            # 处理表格4的数据
            for key, value in part_data.items():
                if value['values']:
                    value['average'] = sum(value['values']) / len(value['values'])
                    value['所有CrrAfter值'] = ', '.join(value['所有CrrAfter值'])
                    # 删除临时字段
                    del value['values']
                    data['表格4-按照元件级窗口合并'].append(value)
        elif log_type == 'solve_algInspection':
            # 临时存储Before和After对的字典
            before_after_pairs = {}
            
            for line in lines:
                # 匹配日志行
                # match = re.search(r'(\d{2}:\d{2}:\d{2}\.\d{3}).*?PCBID_(\w+)_Tdx_(\d+),ArrayNo:(\w+)\s+PartDes:(\w+)\s+AlgID:(\w+)\s+WinName:(\w+):(\w+)', line)
                # match = re.search(r'(\d{2}:\d{2}:\d{2}\.\d{3}).*?PCBID_(\w+)_Tdx_(\d+),ArrayNo:(\w+)\s+PartDes:(\w+)\s+AlgID:(\w+)\s+WinName:([\w_]+):(\w+)', line)
                match = re.search(r'(\d{2}:\d{2}:\d{2}\.\d{3}).*?PCBID_(\w+)_Tdx_(\d+),ArrayNo:(\w+)\s+PartDes:(\w+)\s+AlgID:(\w+)\s+WinName:([\w_]+_\d+(?:_GrpIdx_\d+)?):(\w+)', line)
                if match:
                    timestamp = match.group(1)
                    pcbid = f"PCBID_{match.group(2)}_Tdx_{match.group(3)}"
                    array_no = match.group(4)
                    part_des = match.group(5)
                    alg_id = match.group(6)
                    win_name = match.group(7)
                    status = match.group(8)  # Before或After
                    
                    # 提取AlgTime
                    alg_time_match = re.search(r'AlgTime:(\d+)', line)
                    alg_time = int(alg_time_match.group(1)) if alg_time_match else 0
                    
                    # 创建记录
                    record = {
                        'PCBID_No': pcbid,
                        'ArrayNo': array_no,
                        'PartDes': part_des,
                        'AlgID': alg_id,
                        'WinName': win_name,
                        'Status': status,
                        'AlgTime': alg_time
                    }
                    
                    # 表格1：汇总按PCBID和窗口
                    key = f"{pcbid}_{win_name}"
                    if key not in before_after_pairs:
                        before_after_pairs[key] = {'Before': None, 'After': None}

                    before_after_pairs[key][status] = record
                    
                    # 表格2：没有After的数据
                    if status == 'Before':
                        # 暂时添加到表格2，如果后面找到对应的After，再移除
                        data['表格2-按窗口生成所有没有After'].append(record)
                    
                    # 表格3：按array合并
                    array_key = f"{array_no}_{part_des}_{alg_id}_{win_name}"
                    if status == 'After':
                        if array_key not in array_data:
                            array_data[array_key] = {
                                'ArrayNo': array_no,
                                'PartDes': part_des,
                                'AlgID': alg_id,
                                'WinName': win_name,
                                'low': float('inf'),
                                'high': 0,
                                'average': 0,
                                'values': [],
                                '所有after值': []
                            }
                        
                        # 更新最低值、最高值和所有值
                        array_data[array_key]['low'] = min(array_data[array_key]['low'], alg_time)
                        array_data[array_key]['high'] = max(array_data[array_key]['high'], alg_time)
                        array_data[array_key]['values'].append(alg_time)
                        array_data[array_key]['所有after值'].append(str(alg_time))

                        part_key = f"{part_des}_{alg_id}"
                        if part_key not in part_data:
                            part_data[part_key] = {
                                'PartDes': part_des,
                                'AlgID': alg_id,
                                'WinName': win_name,
                                'low': float('inf'),
                                'high': 0,
                                'average': 0,
                                'values': [],
                                '所有after值': []
                            }
                        
                        # 更新最低值、最高值和所有值
                        part_data[part_key]['low'] = min(part_data[part_key]['low'], alg_time)
                        part_data[part_key]['high'] = max(part_data[part_key]['high'], alg_time)
                        part_data[part_key]['values'].append(alg_time)
                        part_data[part_key]['所有after值'].append(str(alg_time))
                    print("匹配成功:", match.groups())
                else:
                    print("匹配失败:", line.strip())
            # 处理表格1：汇总按PCBID和窗口
            for key, pair in before_after_pairs.items():
                if pair['Before'] and pair['After']:
                    # 如果同时有Before和After，添加到表格1
                    combined_record = {
                        'PCBID_No': pair['Before']['PCBID_No'],
                        'ArrayNo': pair['Before']['ArrayNo'],
                        'PartDes': pair['Before']['PartDes'],
                        'AlgID': pair['Before']['AlgID'],
                        'WinName': pair['Before']['WinName'],
                        'Before': pair['Before']['AlgTime'],
                        'After': pair['After']['AlgTime']
                    }
                    data['表格1-汇总按PCBID和窗口'].append(combined_record)
                    
                    # 从表格2中移除有After的Before记录
                    data['表格2-按窗口生成所有没有After'] = [
                        record for record in data['表格2-按窗口生成所有没有After'] 
                        if not (record['PCBID_No'] == pair['Before']['PCBID_No'] and 
                                record['WinName'] == pair['Before']['WinName'])
                    ]
            
            # 处理表格3：按array合并
            for key, value in array_data.items():
                if value['values']:
                    value['average'] = sum(value['values']) / len(value['values'])
                    value['所有after值'] = ', '.join(value['所有after值'])
                    # 删除临时字段
                    del value['values']
                    data['表格3-按照array级窗口合并'].append(value)
            
            # 处理表格4：按元件合并
            for key, value in part_data.items():
                if value['values']:
                    value['average'] = sum(value['values']) / len(value['values'])
                    value['所有after值'] = ', '.join(value['所有after值'])
                    # 删除临时字段
                    del value['values']
                    data['表格4-按照元件级窗口合并'].append(value)

            # 处理表格5：不含After的数据（只保留必要的列）
            for record in data['表格2-按窗口生成所有没有After']:
                simplified_record = {
                    'ArrayNo': record['ArrayNo'],
                    'PartDes': record['PartDes'],
                    'AlgID': record['AlgID'],
                    'WinName': record['WinName']
                }
                data['表格5-不含After'].append(simplified_record)
            
            # 清空原始表格5数据，使用新的简化数据
            data['表格5-不含After'] = data['表格5-不含After']
            
            # 处理表格6：含After的数据
            # 收集所有PCBID
            pcbid_set = set()
            pcbid_data = {}
            
            for line in lines:
                match = re.search(r'PCBID_(\w+)_Tdx_(\d+),ArrayNo:(\w+)\s+PartDes:(\w+)\s+AlgID:(\w+)\s+WinName:(\w+):After', line)
                if match:
                    pcbid = f"PCBID_{match.group(1)}_Tdx_{match.group(2)}"
                    array_no = match.group(3)
                    part_des = match.group(4)
                    alg_id = match.group(5)
                    win_name = match.group(6)
                    
                    # 提取窗口名称中的数字部分
                    win_match = re.search(r'W_(\d+)_GrpIdx_(\d+)', win_name)
                    if win_match:
                        grp_idx = win_match.group(2)
                        formatted_win_name = f"VP_0_GrpIdx_{grp_idx}"
                        
                        # 记录PCBID
                        pcbid_set.add(pcbid)
                        
                        # 创建键
                        key = f"{array_no}_{part_des}_{alg_id}_{formatted_win_name}"
                        
                        if key not in pcbid_data:
                            pcbid_data[key] = {
                                'ArrayNo': array_no,
                                'PartDes': part_des,
                                'AlgID': alg_id,
                                'WinName': formatted_win_name
                            }
                        
                        # 记录PCBID出现次数
                        if pcbid not in pcbid_data[key]:
                            pcbid_data[key][pcbid] = 0
                        pcbid_data[key][pcbid] += 1
            
            # 清空原始表格6数据
            data['表格6-含After'] = []
            
            # 将数据添加到表格6
            for key, record in pcbid_data.items():
                row = {
                    'ArrayNo': record['ArrayNo'],
                    'PartDes': record['PartDes'],
                    'AlgID': record['AlgID'],
                    'WinName': record['WinName']
                }
                
                # 添加每个PCBID的计数
                for pcbid in sorted(pcbid_set):
                    row[f"{pcbid}_After"] = record.get(pcbid, '')
                
                data['表格6-含After'].append(row)

    return data

def save_to_excel(data, excel_path, sheet_name):
    try:
        # 确保Excel文件存在
        if not os.path.exists(excel_path):
            Workbook().save(excel_path)

        # 读取现有数据
        existing_data = {}
        try:
            excel_data = pd.read_excel(excel_path, sheet_name=None)
            for sheet in excel_data:
                existing_data[sheet] = excel_data[sheet]
        except Exception as e:
            print(f"读取现有Excel文件时出错: {e}")

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # 处理数据格式
            if isinstance(data, dict):
                # 处理字典列表
                if all(isinstance(value, list) for value in data.values()) and all(value and isinstance(value[0], dict) for value in data.values() if value):
                    for key, value in data.items():
                        if not value:  # 跳过空列表
                            print(f"跳过空数据: {key}")
                            continue
                            
                        df = pd.DataFrame(value)
                        sub_sheet_name = f"{sheet_name}-{key}"

                        if sub_sheet_name in existing_data:
                            combined_df = pd.concat([existing_data[sub_sheet_name], df], ignore_index=True)
                            writer.book.remove(writer.book[sub_sheet_name])
                            combined_df.to_excel(writer, sheet_name=sub_sheet_name, index=False)
                        else:
                            df.to_excel(writer, sheet_name=sub_sheet_name, index=False)

                        worksheet = writer.sheets[sub_sheet_name]
                        for i, col in enumerate(df.columns, 1):
                            worksheet.column_dimensions[get_column_letter(i)].width = 25
                            for row in range(2, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row, column=i)
                                cell.number_format = '@'
                                cell.value = str(cell.value) if cell.value is not None else ""

                        print(f"已保存 {len(df)} 条记录到工作表: {sub_sheet_name}")
                else:
                    # 检查所有值的长度是否一致
                    lengths = {key: len(value) for key, value in data.items() if not isinstance(value, list)}
                    if len(set(lengths.values())) > 1:
                        print(f"警告: 数据长度不一致: {lengths}")
                        # 找出最短的长度，截断所有数据到这个长度
                        min_length = min(lengths.values())
                        for key in lengths:
                            if lengths[key] > min_length:
                                print(f"截断 {key} 从 {lengths[key]} 到 {min_length}")
                                data[key] = data[key][:min_length]
                    
                    # 尝试创建DataFrame
                    try:
                        df = pd.DataFrame(data)
                    except ValueError as e:
                        print(f"创建DataFrame失败: {e}")
                        # 尝试另一种方法: 逐行构建
                        rows = []
                        keys = list(data.keys())
                        max_length = max(len(data[key]) for key in keys if isinstance(data[key], (list, tuple)))
                        for i in range(max_length):
                            row = {}
                            for key in keys:
                                if isinstance(data[key], (list, tuple)) and i < len(data[key]):
                                    row[key] = data[key][i]
                                elif not isinstance(data[key], (list, tuple)):
                                    row[key] = data[key]
                            rows.append(row)
                        df = pd.DataFrame(rows)
                    
                    if sheet_name in existing_data:
                        combined_df = pd.concat([existing_data[sheet_name], df], ignore_index=True)
                        writer.book.remove(writer.book[sheet_name])
                        combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                    worksheet = writer.sheets[sheet_name]
                    for i, col in enumerate(df.columns, 1):
                        worksheet.column_dimensions[get_column_letter(i)].width = 25
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=i)
                            cell.number_format = '@'
                            cell.value = str(cell.value) if cell.value is not None else ""

                    print(f"已保存 {len(df)} 条记录到工作表: {sheet_name}")

        # 删除默认Sheet
        wb = load_workbook(excel_path)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            wb.save(excel_path)

    except Exception as e:
        print(f"保存Excel时出错: {e}")
        traceback.print_exc()

def get_latest_log_file(directory, pattern):
    files = os.listdir(directory)
    log_files = [f for f in files if re.match(pattern, f)]
    if not log_files:
        return None
    
    # 从文件名中提取日期并按日期排序
    log_files.sort(key=lambda x: datetime.datetime.strptime(re.search(r'\d{4}-\d{2}-\d{2}', x).group(), '%Y-%m-%d'), reverse=True)
    return os.path.join(directory, log_files[0]) if log_files else None

# 添加各种解析函数
def solve_simpCalTime(simp_log_file_path, thread_log_file_path):
    data = parse_log_file(simp_log_file_path, 'solve_simpCalTime')
    
    # 读取线程日志文件以获取JobName
    with open(thread_log_file_path, 'r', encoding='utf-8') as thread_log_file:
        thread_lines = thread_log_file.readlines()
    
    # 遍历数据以查找对应的JobName
    for i, pcbid in enumerate(data['运行JOB名称']):
        timestamp = pcbid_to_timestamp(pcbid)
        job_name = find_job_name(thread_lines, timestamp)
        data['运行JOB名称'][i] = job_name if job_name else pcbid  # 如果找不到JobName，保留PCBID
    
    # 保存数据到Excel
    save_to_excel(data, excel_path, 'AlgSimpCalTime-整板级')
    print(f"数据已保存到: {excel_path}")

def pcbid_to_timestamp(pcbid):
    # 假设PCBID格式为YYYYMMDDHHMMSS
    return f"{pcbid[8:10]}:{pcbid[10:12]}:{pcbid[12:14]}"

def find_job_name(thread_lines, timestamp):
    for line in thread_lines:
        if timestamp in line:
            match = re.search(r'JobName=(.*?)(?:\s|$)', line)
            if match:
                return match.group(1)
    return None

def solve_ThreadTimeStats(thread_log_file_path, simp_log_file_path):
    # 创建一个字典来存储数据
    thread_data = {}
    simp_data = {}
    
    # 解析thread日志
    if os.path.exists(thread_log_file_path):
        with open(thread_log_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            print(f"读取到 {len(lines)} 行线程日志")
            
            current_block = []
            current_timestamp = None
            current_pcbid = None
            
            for i, line in enumerate(lines):
                # 寻找新块的开始（通常是包含PCBID的行）
                pcbid_match = re.search(r'PCBID_(\d+)', line)
                if pcbid_match:
                    # 如果找到新的PCBID，处理之前的块
                    if current_block and current_timestamp:
                        print(f"  保存时间戳 {current_timestamp} 的块，包含 {len(current_block)} 行")
                        thread_data[current_timestamp] = {
                            'lines': current_block,
                            'pcbid': current_pcbid
                        }
                    
                    # 开始新块
                    current_pcbid = pcbid_match.group(1)
                    current_block = [line]
                    print(f"第 {i+1} 行: 找到新的PCBID块: PCBID_{current_pcbid}")
                    
                    # 提取时间戳（第一行的时间）
                    time_match = re.search(r'(\d{2}:\d{2}:\d{2}\.\d{3})', line)
                    if time_match:
                        current_timestamp = time_match.group(1)[:8]  # 只取HH:MM:SS部分
                        print(f"  提取时间戳: {current_timestamp}")
                elif current_timestamp:
                    # 继续添加到当前块
                    current_block.append(line)
            
            # 处理最后一个块
            if current_block and current_timestamp:
                thread_data[current_timestamp] = {
                    'lines': current_block,
                    'pcbid': current_pcbid
                }
                print(f"保存最后一个时间戳 {current_timestamp} 的块，包含 {len(current_block)} 行")
    else:
        print(f"错误: 线程日志文件不存在: {thread_log_file_path}")
    
    print(f"\n解析完成，找到 {len(thread_data)} 个时间戳块")
    print(f"时间戳列表: {list(thread_data.keys())}")
    
    print(f"\n开始解析Simp日志: {simp_log_file_path}")
    
    # 从simp_log_file_path获取训练时间等信息
    if os.path.exists(simp_log_file_path):
        with open(simp_log_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            print(f"读取到 {len(lines)} 行Simp日志")
            
            i = 0
            while i < len(lines):
                pcbid_match = re.search(r'PCBID:(\d+)', lines[i])
                if pcbid_match:
                    pcbid = pcbid_match.group(1)
                    print(f"第 {i+1} 行: 找到PCBID {pcbid}")
                    
                    # 检查是否有足够的行来提取所有信息
                    if i + 4 < len(lines):
                        train_match = re.search(r'TrainTime\(S\):([0-9.]+)', lines[i+1])
                        cap_match = re.search(r'CapFovTime\(S\):([0-9.]+)', lines[i+2])
                        alg_cal_match = re.search(r'AlgCalTime\(S\):([0-9.]+)', lines[i+3])
                        alg_run_match = re.search(r'AlgRunTime\(S\):([0-9.]+)', lines[i+4])
                        
                        if train_match and cap_match and alg_cal_match and alg_run_match:
                            train_time = float(train_match.group(1))
                            cap_time = float(cap_match.group(1))
                            alg_cal_time = float(alg_cal_match.group(1))
                            alg_run_time = float(alg_run_match.group(1))
                            
                            simp_data[pcbid] = {
                                'TrainTime': train_time,
                                'CapFovTime': cap_time,
                                'AlgCalTime': alg_cal_time,
                                'AlgRunTime': alg_run_time
                            }
                            
                            print(f"  提取到训练时间: {train_time}, 拍照时间: {cap_time}, 计算时间: {alg_cal_time}, 运行时间: {alg_run_time}")
                            i += 5  # 跳过已处理的5行
                        else:
                            i += 1
                    else:
                        i += 1
                else:
                    i += 1
    else:
        print(f"错误: Simp日志文件不存在: {simp_log_file_path}")
    
    print(f"\n解析完成，找到 {len(simp_data)} 个Simp数据记录")
    print(f"Simp PCBID列表: {list(simp_data.keys())}")
    
    print(f"\n开始合并数据...")
    
    # 合并数据
    combined_data = {
        '运行JOB名称': [],
        '在线离线': [],
        '测试版本': [],
        '线程数': [],
        '测试元件数量': [],
        '计算时间最低值': [],
        '计算时间最高值': [],
        '计算时间平均值(去掉最高和最低值取平均)': [],
        '所有PCBID计算时间': [],
        'ThdS': [],
        '训练时间-TrainTime(S)': [],
        '计算时间-AlgCalTime(S)': [],
        '拍照时间-CapFovTime(S)': [],
        '总运行时间-AlgRunTime(S)': [],
    }
    
    match_count = 0
    
    # 遍历线程数据
    for timestamp, thread_info in thread_data.items():
        print(f"处理时间戳: {timestamp}")
        
        # 构建格式化的PCBID (YYYYMMDDHHMMSS)
        # 假设当前日期是2025-02-27，从日志文件名获取
        formatted_pcbid = f"20250227{timestamp.replace(':', '')}"
        
        # 查找匹配的Simp数据
        if formatted_pcbid in simp_data:
            print(f"  找到匹配的Simp数据: {formatted_pcbid}")
            match_count += 1
            
            # 从线程日志中提取信息
            thread_block = thread_info['lines']
            job_name = None
            app_version = None
            pcb_mode = None
            thread_times = []
            comp_nums = []
            
            for line in thread_block:
                # 提取JobName和AppVersion
                job_match = re.search(r'JobName=([^\s]+)(?:\s+AppVersion=([^\s]+))?', line)
                if job_match:
                    job_name = job_match.group(1)
                    app_version = job_match.group(2) if job_match.group(2) else app_version
                
                # 提取PCBInspect_Mode
                mode_match = re.search(r'PCBInspect_Mode:(\w+)', line)
                if mode_match:
                    pcb_mode = mode_match.group(1)
                
                # 提取线程时间和组件数量
                thd_match = re.search(r'Thd_\d+:ThdS:[^,]+,CalS:[^,]+,CalE:[^,]+,Time:\s*(\d+)\s*/\s*(\d+).*?CalWinTime:(\d+),CompNum:(\d+)', line)
                if thd_match:
                    thread_times.append(int(thd_match.group(2)))
                    comp_nums.append(int(thd_match.group(4)))
            
            # 计算统计值
            if thread_times:
                min_time = min(thread_times)
                max_time = max(thread_times)
                
                # 计算平均值(去掉最高和最低值)
                if len(thread_times) > 2:
                    sorted_times = sorted(thread_times)
                    avg_time = sum(sorted_times[1:-1]) / (len(sorted_times) - 2)
                else:
                    avg_time = sum(thread_times) / len(thread_times)
                
                # 格式化所有PCBID计算时间为逗号分隔的列表
                all_times_str = ', '.join([str(t) for t in thread_times])
                
                # 填充数据
                combined_data['运行JOB名称'].append(job_name if job_name else "未知")
                combined_data['在线离线'].append(pcb_mode if pcb_mode else "未知")
                combined_data['测试版本'].append(app_version if app_version else "未知")
                combined_data['线程数'].append(len(thread_times))
                combined_data['测试元件数量'].append(sum(comp_nums))
                combined_data['计算时间最低值'].append(min_time)
                combined_data['计算时间最高值'].append(max_time)
                combined_data['计算时间平均值(去掉最高和最低值取平均)'].append(round(avg_time, 2))
                combined_data['所有PCBID计算时间'].append(all_times_str)
                combined_data['ThdS'].append(timestamp)
                
                # 添加从Simp日志获取的数据
                combined_data['训练时间-TrainTime(S)'].append(simp_data[formatted_pcbid]['TrainTime'])
                combined_data['计算时间-AlgCalTime(S)'].append(simp_data[formatted_pcbid]['AlgCalTime'])
                combined_data['拍照时间-CapFovTime(S)'].append(simp_data[formatted_pcbid]['CapFovTime'])
                combined_data['总运行时间-AlgRunTime(S)'].append(simp_data[formatted_pcbid]['AlgRunTime'])
        else:
            print(f"  在Simp数据中未找到匹配的PCBID: {formatted_pcbid}")
    
    print(f"\n合并完成，成功匹配 {match_count}/{len(thread_data)} 个时间戳")
    
    print(f"\n保存数据到Excel...")
    save_to_excel(combined_data, excel_path, 'AlgThreadTimeStats-线程级')
    print(f"数据已保存到: {excel_path}")
    
    # 添加这一行来生成折线图
    add_chart_to_excel(excel_path, 'AlgThreadTimeStats-线程级')
    
    return combined_data

def add_chart_to_excel(excel_path, sheet_name):
    try:
        print(f"开始添加折线图到 {excel_path}, 工作表: {sheet_name}")
        wb = load_workbook(excel_path)
        
        # 检查工作表名称
        print(f"工作表列表: {wb.sheetnames}")
        
        # 尝试找到正确的工作表名称
        actual_sheet_name = None
        for name in wb.sheetnames:
            if sheet_name.lower() in name.lower():
                actual_sheet_name = name
                break
        
        if not actual_sheet_name:
            print(f"找不到包含 {sheet_name} 的工作表")
            return
            
        print(f"使用工作表: {actual_sheet_name}")
        ws = wb[actual_sheet_name]
        
        # 找到最后一行
        last_row = ws.max_row
        print(f"最后一行: {last_row}")
        
        # 找到列索引
        job_col = None
        time_col = None
        thds_col = None
        
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"列 {col}: {header}")
            if header and '运行job名称'.lower() in header.lower():  # 不区分大小写
                job_col = col
            elif header == '所有PCBID计算时间':
                time_col = col
            elif header == 'ThdS':
                thds_col = col
        
        print(f"找到列: job_col={job_col}, time_col={time_col}, thds_col={thds_col}")
        
        if not (job_col and time_col and thds_col):
            print("找不到必要的列")
            return
        
        # 收集数据
        job_data = {}
        for row in range(2, last_row + 1):
            job_name = ws.cell(row=row, column=job_col).value
            thds = ws.cell(row=row, column=thds_col).value
            time_str = ws.cell(row=row, column=time_col).value
            
            if not (job_name and thds and time_str):
                continue
                
            # 解析时间值（取"/"后面的数字）
            times = []
            for time_part in time_str.split(','):
                time_part = time_part.strip()
                if time_part.isdigit():
                    times.append(int(time_part))
            
            if job_name not in job_data:
                job_data[job_name] = []
            
            for time in times:
                job_data[job_name].append((thds, time))
        
        print(f"收集到的JOB数据: {job_data.keys()}")
        
        # 创建新的sheet页用于图表和数据（添加排序逻辑）
        sorted_jobs = sorted(job_data.keys())  # 按字母顺序排序JOB名称

        chart_sheet_name = f"{sheet_name}_图表"
        if chart_sheet_name in wb.sheetnames:
            wb.remove(wb[chart_sheet_name])
        chart_ws = wb.create_sheet(title=chart_sheet_name)

        # 标题行保持不变
        chart_ws.cell(row=1, column=1).value = "JOB名称"
        chart_ws.cell(row=1, column=2).value = "时间戳"
        chart_ws.cell(row=1, column=3).value = "计算时间"

        # 按排序后的顺序写入数据并记录范围
        current_row = 2
        job_ranges = {}

        for job_name in sorted_jobs:
            data_points = job_data[job_name]
            start_row = current_row
            
            for thds, time in data_points:
                chart_ws.cell(row=current_row, column=1).value = job_name
                chart_ws.cell(row=current_row, column=2).value = thds
                chart_ws.cell(row=current_row, column=3).value = time
                current_row += 1
            
            end_row = current_row - 1
            job_ranges[job_name] = (start_row, end_row)

        # 创建图表时使用记录的范围数据
        chart = LineChart()
        chart.title = "按JOB生成测试时间生成折线图"
        chart.style = 13
        chart.height = 15
        chart.width = 20
        chart.y_axis.title = "计算时间"
        chart.x_axis.title = "时间戳"

        for i, job_name in enumerate(sorted_jobs):  # 使用enumerate获取索引
            start_row, end_row = job_ranges[job_name]
            
            # 添加调试信息
            print(f"Processing {job_name}: rows {start_row}-{end_row}")
            
            # Y轴数据（计算时间）
            y_values = Reference(chart_ws, 
                                min_col=3, 
                                min_row=start_row, 
                                max_row=end_row)
            print(f"y_values: {y_values}")
            # X轴数据（时间戳）
            x_values = Reference(chart_ws,
                                min_col=2,
                                min_row=start_row,
                                max_row=end_row)
            print(f"x_values: {x_values}")
            
            # 创建系列（正确参数传递方式）
            series = Series(y_values, title=job_name)  # 第一个参数是y值
            series.xvalues = x_values  # 直接设置x轴数据
            series.idx = i  # 必须显式设置索引
            
            chart.series.append(series)

        # 添加图表到指定位置
        chart_ws.add_chart(chart, "D1")
        
        # 保存Excel文件
        wb.save(excel_path)
        print(f"图表已添加到 {excel_path} 的 {chart_sheet_name} 工作表中")
    except Exception as e:
        import traceback
        print(f"添加折线图时出错: {e}")
        print(traceback.format_exc())

def solve_alg3DCrr(log_file_path):
    print(f"正在处理文件: {log_file_path}")
    data = parse_log_file(log_file_path, 'solve_alg3DCrr')

    # 确保每个表格的数据是字典列表
    for key, value in data.items():
        if value and isinstance(value, list):
            # 如果数据是字典列表，确保每个字典的键是列名
            if isinstance(value[0], dict):
                print(f"表格 {key} 的数据格式正确")
            else:
                print(f"表格 {key} 的数据格式不正确，需要转换为字典列表")
                # 如果数据不是字典列表，可以尝试转换
                data[key] = [{"列名": item} for item in value]
        else:
            print(f"表格 {key} 的数据为空或格式不正确")

    # 保存数据到 Excel
    for key, value in data.items():
        if value:
            sub_sheet_name = f"ALG3DCrr-{key}"
            save_to_excel({key: value}, excel_path, sub_sheet_name)

    print(f"数据已保存到: {excel_path}")

def solve_algInspection(log_file_path):
    print(f"正在处理文件: {log_file_path}")
    data = parse_log_file(log_file_path, 'solve_algInspection')
    
    # 检查解析结果
    print("解析结果:")
    for key, value in data.items():
        if isinstance(value, list):
            print(f"  {key}: {len(value)} 条记录")
        elif isinstance(value, dict):
            print(f"  {key}: {len(value)} 个键值对")
    
    save_to_excel(data, excel_path, 'AlgInspection')
 
    # 如果Excel文件已存在，删除AlgInspection主表格
    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            if 'AlgInspection' in wb.sheetnames:
                wb.remove(wb['AlgInspection'])
                wb.save(excel_path)
                print(f"已删除AlgInspection工作表")
        except Exception as e:
            print(f"删除工作表时出错: {e}")
    
    print(f"数据已保存到: {excel_path}")

def create_line_chart(excel_path):
    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb.active  # 假设数据在活动工作表中

    # 创建折线图对象
    chart = LineChart()
    chart.title = "按JOB生成的测试时间"
    chart.y_axis.title = "计算时间"
    chart.x_axis.title = "时间戳"

    # 获取所有不同的JOB名称
    jobs = pd.Series(ws['A'][1:]).unique()  # 假设JOB名称在第一列

    for job in jobs:
        # 找到该JOB的所有行
        rows = [i for i, cell in enumerate(ws['A'][1:], start=2) if cell.value == job]

        if not rows:
            continue

        # 设置数据范围
        data = Reference(ws, min_col=2, min_row=rows[0], max_row=rows[-1])  # 假设计算时间在第二列
        categories = Reference(ws, min_col=1, min_row=rows[0], max_row=rows[-1])  # 假设时间戳在第一列

        # 创建系列并添加到图表
        series = Series(data, title=job)
        chart.append(series)

    # 将图表添加到工作表
    ws.add_chart(chart, "D1")

    # 保存修改后的文件
    wb.save(excel_path)

if __name__ == "__main__":
    print("程序开始执行")
    
    # 获取日志文件路径
    simp_log_file_path = os.path.join(directory, "AlgSimpCalTime_2025-02-27.log")
    thread_log_file_path = os.path.join(directory, "AlgThreadTimeStats_2025-02-27.log")
    ALG3DCrr_log_file_path = os.path.join(directory, "ALG3DCrr_2025-02-27.log")
    AlgInspection_log_file_path = os.path.join(directory, "AlgInspection_Idx_7_2025-02-27.log")
    # # 删除excel_path
    # if os.path.exists(excel_path):
    #     try:
    #         os.remove(excel_path)
    #         print(f"已删除文件: {excel_path}")
    #     except Exception as e:
    #         print(f"删除文件时出错: {e}")
    # 调用各个功能
    solve_simpCalTime(simp_log_file_path, thread_log_file_path)
    solve_ThreadTimeStats(thread_log_file_path, simp_log_file_path)
    solve_alg3DCrr(ALG3DCrr_log_file_path)
    solve_algInspection(AlgInspection_log_file_path)
    
    # # 确保删除默认的Sheet页
    # try:
    #     if os.path.exists(excel_path):
    #         wb = load_workbook(excel_path)
    #         if 'Sheet' in wb.sheetnames:
    #             wb.remove(wb['Sheet'])
    #             wb.save(excel_path)
    #             print("已删除默认的Sheet页")
    # except Exception as e:
    #     print(f"删除Sheet页时出错: {e}")
    
    print("程序执行完毕")

