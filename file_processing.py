import threading
import time
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image

import utils


class FileProcessingApp:
    def __init__(self, master):
        self.master = master
        self.master.title("文件处理")
        self.master.geometry("400x300")
        self.master.configure(bg="#f0f0f0")
        
        # 使用网格布局优化界面
        self.frame = tk.Frame(master, bg="#f0f0f0")
        self.frame.pack(pady=20)
        
        # 提示文本：选择eval_xlsx的路径
        self.eval_label = tk.Label(self.frame, text="请选择eval_xlsx的路径", font=("Arial", 12), bg="#f0f0f0")
        self.eval_label.grid(row=2, column=0, padx=5, sticky='w')
        
        # 选择路径按钮
        self.create_button = tk.Button(self.frame, text="选择路径", command=self.load_eval_xlsx, font=("Arial", 10), bg="#4CAF50", fg="white", relief="raised")
        self.create_button.grid(row=2, column=1, padx=5, sticky='w')
        
        self.selected_path_label = tk.Label(self.frame, text="", font=("Arial", 10), bg="#f0f0f0")
        self.selected_path_label.grid(row=3, column=0, columnspan=2, pady=10, sticky='w')
        
        # 新增执行按钮
        self.execute_button = tk.Button(self.frame, text="执行操作", command=self.start_processing_thread, font=("Arial", 10), bg="#2196F3", fg="white", relief="raised")
        self.execute_button.grid(row=4, column=0, columnspan=2, pady=10, sticky='w')

    # 选择eval_xlsx的路径
    def load_eval_xlsx(self):
        try:
            # 使用文件对话框选择文件
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
            if file_path:
                self.selected_path_label.config(text=f"已选择文件: {file_path}")
                self.create_button.config(text="文件已选择", bg="#8BC34A")
            else:
                messagebox.showwarning("警告", "未选择任何文件")
        except Exception as e:
            messagebox.showerror("错误", f"选择文件时发生错误: {e}")

    def start_processing_thread(self):
        # 启动新线程来执行耗时操作
        threading.Thread(target=self.execute_processing, daemon=True).start()

    # 执行操作
    def execute_processing(self):
        eval_xlsx_path = self.selected_path_label.cget("text").replace("已选择文件: ", "")
        if not eval_xlsx_path:
            messagebox.showwarning("警告", "请确保已选择文件")
            return
        # 调用 static 函数
        create_static_xlsx(eval_xlsx_path)
        
        # 操作完成后显示提示框
        messagebox.showinfo("完成", "文件处理已完成！")

# 根据eval_xlsx生成static_gd_xxx.xlsx
def create_static_xlsx(eval_xlsx_path, save_dir=Path("./")):
    # 最好能处理多个eval_xlsx
    excel_path_list = [eval_xlsx_path]
    # 获取当前时间字符串
    time_str = time.strftime("%Y-%m-%d", time.localtime())
    # 生成保存路径
    excel_save_path = save_dir.joinpath("static_gd_" + time_str + ".xlsx")
    # 创建ExcelWriter对象
    writer = pd.ExcelWriter(excel_save_path)
    data_static_list = []
    # 初始化数据字典
    data_static_dict = {"数据时间": "", "data": {"误报": {}, "漏报": {}}}
    data_times = []
    # 定义列名
    cols = ["料号", "元件类型", "张数", "检测图", "定位图", "对比图", "新版本是否解决", "新版本解决总数", "问题原因", "解决方案", "预计更新时间", "到现场时间"]
    err_report_save_data = {}
    missed_report_save_data = {}
    # 初始化保存数据字典
    for col in cols:
        err_report_save_data[col] = None
        missed_report_save_data[col] = None
    idx = ["误报", "漏报"]
    # 创建DataFrame对象
    df = pd.DataFrame(index=idx, columns=cols)
    
    for excel_path in excel_path_list:
        # 初始化错误报告和失败报告字典
        err_report_static_dict = {}
        missed_report_static_dict = {}
        # 读取Excel数据
        excel_data = pd.read_excel(str(excel_path))
        rows = excel_data.shape[0]

        for row in range(excel_data.shape[0]):
            # 获取错误类型
            err_type = excel_data.iloc[row]["数据类型"]
            # 获取数据时间
            if "数据时间" in excel_data.iloc[row]:
                data_time = excel_data.iloc[row]["数据时间"]
                data_static_dict["数据时间"] = data_time
            else:
                data_static_dict["数据时间"] = time_str
            # 获取图片路径和组件编号
            insp_img_path = excel_data.iloc[row]["img_path"]
            comp_no = Path(insp_img_path).parent.parent.parent.parent.name
            part_no = Path(insp_img_path).parent.parent.parent.name 
            final_res = excel_data.iloc[row]["final_res"]
            # 处理漏测数据
            if err_type == "漏测":
                if part_no in missed_report_static_dict.keys():
                    missed_report_static_dict[part_no]["元件类型"] = comp_no
                    missed_report_static_dict[part_no]["张数"] += 1
                    if final_res == -1:
                        missed_report_static_dict[part_no]["新版本解决总数"] += 1
                    # missed_report_static_dict[part_no]["检测图"] = excel_data.iloc[row]["eval图片"]
                    # missed_report_static_dict[part_no]["定位图"] = excel_data.iloc[row]["定位图片"]
                    # missed_report_static_dict[part_no]["对比图"] = excel_data.iloc[row]["train图片"]
                else:
                    missed_report_static_dict[part_no] = {}
                    missed_report_static_dict[part_no]["元件类型"] = comp_no
                    missed_report_static_dict[part_no]["张数"] = 1
                    missed_report_static_dict[part_no]["新版本解决总数"] = 0
                    missed_report_static_dict[part_no]["新版本是否解决"] = None
                    if final_res == -1:
                        missed_report_static_dict[part_no]["新版本解决总数"] += 1
                    # missed_report_static_dict[part_no]["检测图"] = excel_data.iloc[row]["img_path"]
                    # missed_report_static_dict[part_no]["定位图"] = excel_data.iloc[row]["定位图片"]
                    # missed_report_static_dict[part_no]["对比图"] = excel_data.iloc[row]["train_img_path"]
            # 处理误报数据
            elif err_type == "误报":
                if part_no in err_report_static_dict.keys():
                    err_report_static_dict[part_no]["元件类型"] = comp_no
                    err_report_static_dict[part_no]["张数"] += 1
                    if final_res == 1:
                        err_report_static_dict[part_no]["新版本解决总数"] += 1
                    # missed_report_static_dict[part_no]["检测图"] = excel_data.iloc[row]["eval图片"]
                    # missed_report_static_dict[part_no]["定位图"] = excel_data.iloc[row]["定位图片"]
                    # missed_report_static_dict[part_no]["对比图"] = excel_data.iloc[row]["train图片"]
                else:
                    err_report_static_dict[part_no] = {}
                    err_report_static_dict[part_no]["元件类型"] = comp_no
                    err_report_static_dict[part_no]["张数"] = 1
                    err_report_static_dict[part_no]["新版本解决总数"] = 0
                    err_report_static_dict[part_no]["新版本是否解决"] = None
                    if final_res == 1:
                        err_report_static_dict[part_no]["新版本解决总数"] += 1
                    # err_report_static_dict[part_no]["检测图"] = excel_data.iloc[row]["img_path"]
                    # err_report_static_dict[part_no]["定位图"] = excel_data.iloc[row]["定位图片"]
                    # err_report_static_dict[part_no]["对比图"] = excel_data.iloc[row]["train_img_path"]
            
        # 更新数据字典
        data_static_dict["data"]["误报"] = err_report_static_dict
        data_static_dict["data"]["漏报"] = missed_report_static_dict

        err_report_static_dict_list = []
        missed_report_static_dict_list = []
        # 处理误报数据
        for k, v in data_static_dict["data"]["误报"].items():
            err_report_save_data["料号"] = k
            err_report_save_data["张数"] = v["张数"]
            # err_report_save_data["检测图"] = v["检测图"]
            # err_report_save_data["定位图"] = v["定位图"]
            # err_report_save_data["对比图"] = v["对比图"]
            err_report_save_data["新版本是否解决"] = v["新版本是否解决"]
            err_report_save_data["新版本解决总数"] = v["新版本解决总数"]
            err_report_save_data["问题原因"] = None
            err_report_save_data["解决方案"] = None
            err_report_save_data["预计更新时间"] = None
            err_report_save_data["到现场时间"] = None
            err_report_static_dict_list.append(err_report_save_data)
        # 处理漏报数据
        for k, v in data_static_dict["data"]["漏报"].items():
            missed_report_save_data["料号"] = k
            missed_report_save_data["张数"] = v["张数"]
            # missed_report_save_data["检测图"] = v["检测图"]
            # missed_report_save_data["定位图"] = v["定位图"]
            # missed_report_save_data["对比图"] = v["对比图"]
            missed_report_save_data["新版本是否解决"] = v["新版本是否解决"]
            missed_report_save_data["新版本解决总数"] = v["新版本解决总数"]
            missed_report_save_data["问题原因"] = None
            missed_report_save_data["解决方案"] = None
            missed_report_save_data["预计更新时间"] = None
            missed_report_save_data["到现场时间"] = None
            missed_report_static_dict_list.append(missed_report_save_data)

    # 将数据写入Excel文件
    df = pd.DataFrame.from_dict(missed_report_static_dict, orient='index', columns=cols)       
    df.to_excel(writer, sheet_name="漏报")
    df = pd.DataFrame.from_dict(err_report_static_dict, orient='index', columns=cols)            
    df.to_excel(writer, sheet_name="误报")
    writer.close()

    # 处理Excel文件并插入图片
    process_excel(excel_save_path, eval_xlsx_path)

def process_excel(file_path, eval_xlsx_path):
    try:
        workbook_static = openpyxl.load_workbook(file_path)
        workbook_eval = openpyxl.load_workbook(eval_xlsx_path)
        sheet_eval = workbook_eval.active

        # 获取列名对应的列索引
        eval_columns = {cell.value: idx for idx, cell in enumerate(sheet_eval[1])}
        static_columns = {cell.value: idx for idx, cell in enumerate(workbook_static['漏报'][1])}

        for sheet_name in ['漏报', '误报']:
            if sheet_name in workbook_static.sheetnames:
                sheet_static = workbook_static[sheet_name]
                for row_static in sheet_static.iter_rows(min_row=2, max_col=1):
                    part_no = row_static[0].value
                    matched = False
                    for row_eval in sheet_eval.iter_rows(min_row=2, max_col=6):
                        img_path = row_eval[eval_columns['img_path']].value
                        if get_part_no(img_path) == part_no:
                            matched = True
                            print(f"匹配料号: {part_no}")  # 调试信息
                            for image in sheet_eval._images:
                                if image.anchor._from.row == row_eval[0].row and image.anchor._from.col in [
                                    eval_columns['eval图片'], eval_columns['定位图片'], eval_columns['train图片']]:
                                    # 根据列名插入图片
                                    target_col = static_columns['检测图'] if image.anchor._from.col == eval_columns['eval图片'] else \
                                                 static_columns['定位图'] if image.anchor._from.col == eval_columns['定位图片'] else \
                                                 static_columns['对比图']
                                    new_img = Image(image.ref)
                                    new_img.anchor = f"{chr(65 + target_col)}{row_static[0].row}"
                                    max_width, max_height = 75, 75
                                    new_width, new_height = utils.adjust_image_size(image.ref, max_width, max_height)
                                    if new_width is not None and new_height is not None:
                                        new_img.width = new_width
                                        new_img.height = new_height
                                    sheet_static.add_image(new_img)
                    if not matched:
                        continue
        workbook_static.save(file_path)
    except Exception as e:
        messagebox.showerror("错误", f"处理Excel文件时发生错误: {e}")


def get_part_no(img_path):
    # 从 img_path 中提取出料号
    parts = img_path.split("\\")
    return parts[-3]  # 假设料号在倒数第三个位置

if __name__ == "__main__":
    root = tk.Tk()
    app = FileProcessingApp(root)
    root.mainloop()
