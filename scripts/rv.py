import csv
import glob
import time
import openpyxl
from openpyxl.styles import Font
import pyperclip
import utils
import config
import re
import os
import datetime
import chardet
import pyautogui
import pandas as pd
from pywinauto import Application, Desktop
from loguru import logger

def front_rv_window():
    try:
        utils.click_by_png(config.RV_AI_TOPIC_DARK)
        find_topic = True
    except Exception:
        find_topic = False
    if not find_topic:
        logger.info("未搜索到标题，改为搜索窗格")
        windows = Desktop(backend="win32").windows()
        window_found = False
        logger.info("开始寻找程序窗口")
        pattern = re.compile(r".*Sinictek-训练.*")
        sinictek_amount = 0
        for w in windows:
            if pattern.match(w.window_text()):
                window_properties = w.get_properties()
                logger.info(f"'Sinictek-训练'窗口存在,详细信息：{window_properties}")
                window_found = True
                sinictek_amount += 1
        logger.info(f"'Sinictek-训练'窗口数量: {sinictek_amount}")
        if window_found:
            app = Application().connect(title_re=".*Sinictek-训练.*")
            main_window = app.window(title_re="Sinictek-训练")
            if main_window.exists(timeout=10):
                logger.info("成功连接到窗口")
                main_window.wait('ready', timeout=10)
                main_window.set_focus()
                # main_window.bring_to_front()
                # main_window.maximize()
                main_window.wait('ready', timeout=10)
            else:
                logger.error("未找到窗口")
                raise Exception("未找到窗口")
        else:
            logger.error("未找到程序")
            raise Exception("未找到程序")

def sync_csv_to_xlsx(csv_path, xlsx_path):
    try:
        logger.info(f"将csv更新的数据剪切至xlsx...csv: {csv_path}, xlsx: {xlsx_path}")

        # 检测CSV文件编码
        with open(csv_path, 'rb') as file:
            content = file.read()
            result = chardet.detect(content)
            encoding = result['encoding'] if result['encoding'] else 'utf-8'

        # 读取CSV和XLSX文件
        df_csv = pd.read_csv(csv_path, encoding=encoding)
        df_xlsx = pd.read_excel(xlsx_path, engine='openpyxl')

        # 确保id列为字符串类型并去除空格
        df_csv['id'] = df_csv['id'].astype(str).str.strip()
        df_xlsx['id'] = df_xlsx['id'].astype(str).str.strip()

        # 遍历CSV中的每一行，只剪切非标题行外的非空的新数据到XLSX
        rows_to_remove = []
        for index, row in df_csv.iterrows():
            try:
                if row.isnull().all():
                    logger.info(f"CSV中的行为空，跳过: {row.to_dict()}")
                    continue

                csv_id = row['id']
                # 查找XLSX中对应的行
                xlsx_row = df_xlsx[df_xlsx['id'] == csv_id]
                if not xlsx_row.empty:
                    # 检查CSV行是否已存在于XLSX中
                    if (xlsx_row.iloc[0, 6:] == row).all():
                        logger.info(f"CSV中的行已存在于XLSX中，跳过同步: {row.to_dict()}")
                        continue
                    # 更新XLSX中对应的列，假设df_csv的列数为N
                    for col_index, col_name in enumerate(df_csv.columns):
                        # 在XLSX中对应的列索引为6 + col_index
                        df_xlsx.loc[xlsx_row.index, df_xlsx.columns[col_index + 6]] = row[col_name]
                    logger.info(f"已将CSV中的行剪切到XLSX: {row.to_dict()}")
                else:
                    # 如果XLSX中没有对应的行，则添加新行
                    new_row = pd.Series([None] * 6 + list(row), index=df_xlsx.columns[:6 + len(row)])
                    df_xlsx = pd.concat([df_xlsx, new_row.to_frame().T], ignore_index=True)
                    logger.info(f"已将CSV中的新行添加到XLSX: {row.to_dict()}")

                # 记录需要删除的CSV行索引
                rows_to_remove.append(index)
            except Exception as row_error:
                logger.error(f"处理CSV行时发生错误: {row_error}, 行内容: {row.to_dict()}")

        # 删除CSV中已处理的行，保留标题行
        if rows_to_remove:
            logger.info(f"将从CSV中删除以下行索引: {rows_to_remove}")
            df_csv.drop(rows_to_remove, inplace=True)
            df_csv.to_csv(csv_path, index=False, encoding=encoding)

        # 保存更新后的XLSX文件
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df_xlsx.to_excel(writer, index=False)

        logger.info("csv更新的数据剪切至xlsx同步完成")
    except Exception as e:
        logger.error(f"处理Excel文件时发生错误: {e}")
        raise Exception(f"处理Excel文件时发生错误: {e}")

# ai复判 总路径(包含train,test文件夹),结果路径（含定位图片及输出数据文档）
def rv_ai_test(train_eval_paths, result_path, mode):
    # 从提供的路径向上遍历，直到找到包含pyd的目录,用pyd_path保存
    logger.info(f"传入的train_eval_paths: {train_eval_paths}")
    logger.info(f"传入的result_path: {result_path}")
    logger.info(f"传入的mode: {mode}")
    logger.info(f"查询pyd路径")
    pyd_path = result_path
    trained_success_list = []
    substring = 'py'
    root_path = os.path.abspath(os.sep)
    found = False  # 用于标记是否找到包含子字符串的目录

    while pyd_path != root_path:
        if substring in os.path.basename(pyd_path):
            found = True  # 找到了包含子字符串的目录
            logger.info(f"找到pyd路径: {pyd_path}")
            break
        pyd_path = os.path.dirname(pyd_path)

    if not found:
        logger.error(f"在路径 {result_path} 中未找到包含 '{substring}' 的目录")
        raise Exception(f"在路径 {result_path} 中未找到包含 '{substring}' 的目录")

    result_xlsx_dir = os.path.join(pyd_path, "eval_logs")
    os.makedirs(result_xlsx_dir, exist_ok=True)
    # 仅仅将csv处理为xlsx
    if not train_eval_paths and result_path or mode == "file":

        logger.info("训练推理为空，结果路径非空，开始处理结果路径下的.csv文件")
        csv_files = glob.glob(os.path.join(result_path, '*.csv'))
        if not csv_files:
            logger.error("未找到任何.csv文件")
            raise Exception("未找到任何.csv文件")
        # 只处理文件模式
        for csv_file in csv_files:
            logger.info(f"处理文件: {csv_file}")
            csv_file_name = os.path.splitext(os.path.basename(csv_file))[0]
            eval_result_xlsx = os.path.join(result_path, f"{csv_file_name}.xlsx")
            completed_txt_path = os.path.join(pyd_path, "脚本已运行完成.txt")
            if os.path.exists(completed_txt_path):
                os.remove(completed_txt_path)
            try:
                with open(csv_file, 'rb') as file:
                    content = file.read()
                    result = chardet.detect(content)
                    encoding = result['encoding'] if result['encoding'] else 'utf-8'
                    try:
                        content = content.decode(encoding)
                    except UnicodeDecodeError:
                        content = content.decode('gbk', errors='ignore')

                df_csv = pd.read_csv(csv_file, encoding=encoding)
                # 添加新列
                df_csv = df_csv.assign(eval结果=None, good_ng=None, eval图片=None, 定位图片=None, train图片=None, eval路径=None)
                # 重新排序列，将新列放在前面
                columns_order = ['eval结果', 'good_ng', 'eval图片', '定位图片', 'train图片', 'eval路径'] + [col for col in df_csv.columns if col not in ['eval结果', 'good_ng', 'eval图片', '定位图片', 'train图片', 'eval路径']]
                df_csv = df_csv[columns_order]
                df_csv.to_excel(eval_result_xlsx, index=False, engine='openpyxl')
                logger.info("对xlsx作预处理")
                xlsx = pd.read_excel(eval_result_xlsx, engine='openpyxl')
                xlsx.columns = [col.strip() for col in xlsx.columns]
                wb = openpyxl.load_workbook(eval_result_xlsx)
                ws = wb.active
                for cell in ws[1]:
                    cell.font = Font(bold=False)
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 18
                wb.save(eval_result_xlsx)
                xlsx = pd.read_excel(eval_result_xlsx, engine='openpyxl')
                xlsx.columns = [re.sub(r'\s+', '', col) for col in xlsx.columns]
                logger.info("对xlsx作预处理完成")
                
                wb = openpyxl.load_workbook(eval_result_xlsx)
                ws = wb.active
                for index, row in xlsx.iterrows():
                    excel_row = index + 2
                    if excel_row > 1:
                        ws.row_dimensions[excel_row].height = 75
                    id = ws.cell(row=excel_row, column=8).value

                    # 检测train_eval_paths下（包括根目录）文件名为id的（不包括后缀） 后缀为jpg或者bmp的路径
                    eval_path = None
                    id_without_ext = os.path.splitext(id)[0]
                    if isinstance(train_eval_paths, list):
                        logger.info("train_eval_paths是一个列表")
                        for path in train_eval_paths:
                            for root, dirs, files in os.walk(path):
                                for file in files:
                                    if os.path.splitext(file)[0] == id_without_ext and file.lower().endswith(('.jpg', '.bmp')):
                                        eval_path = os.path.join(root, file)
                                        break
                                if eval_path:
                                    break
                            if eval_path:
                                break
                    else:
                        logger.info("train_eval_paths不是一个列表，假定其为单个路径")
                        for root, dirs, files in os.walk(train_eval_paths):
                            for file in files:
                                if os.path.splitext(file)[0] == id_without_ext and file.lower().endswith(('.jpg', '.bmp')):
                                    eval_path = os.path.join(root, file)
                                    break
                            if eval_path:
                                break
                    ws.cell(row=excel_row, column=6).value = eval_path if eval_path else "未找到路径"

                    eval_image_cell = ws.cell(row=excel_row, column=3)
                    eval_image_path = ws.cell(row=excel_row, column=9).value        
                    try:
                        if eval_image_path:
                            eval_image_path = os.path.normpath(eval_image_path)
                            if os.path.exists(eval_image_path):
                                utils.insert_image_limited(ws, eval_image_cell, eval_image_path)
                            else:
                                eval_image_cell.value = "空"
                                logger.info(f"id: {id}, eval图片路径不存在: {eval_image_path}")
                        else:
                            eval_image_cell.value = "空"
                    except Exception as e:
                        logger.info(f"id: {id}, 处理eval图片 {eval_image_path} 时出错: {e}")

                    cad_cell = ws.cell(row=excel_row, column=4)
                    cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.bmp")        
                    try:
                        if not os.path.exists(cad_loc_path):
                            cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.jpg")
                        if not os.path.exists(cad_loc_path):
                            cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.png")
                        cad_loc_path = os.path.normpath(cad_loc_path)
                        if cad_loc_path and os.path.exists(cad_loc_path):
                            utils.insert_image_limited(ws, cad_cell, cad_loc_path)
                        else:
                            cad_cell.value = "空"
                            logger.error(f"id: {id}, cad_loc图片路径不存在: {cad_loc_path}")
                    except Exception as e:
                        logger.error(f"id: {id}, cad_img处理图片 {cad_loc_path} 时出错: {e}")

                    if ws.cell(row=excel_row, column=10).value is None:
                        ws.cell(row=excel_row, column=5).value = "空"
                    else:
                        train_image_path = ws.cell(row=excel_row, column=10).value
                        train_cell = ws.cell(row=excel_row, column=5)
                        try:
                            logger.info(f"id: {id}, train_image_path: {train_image_path}")
                            if train_image_path and os.path.exists(train_image_path):
                                train_image_path = os.path.normpath(train_image_path)
                                utils.insert_image_limited(ws, train_cell, train_image_path)
                            else:
                                train_cell.value = "空"
                        except Exception as e:
                            logger.info(f"id: {id}, train_img处理图片 {train_image_path} 时出错: {e}")

                logger.info("所有图片处理完毕，正在保存工作簿...")
                wb.save(eval_result_xlsx)
                logger.info(f"处理完成: {eval_result_xlsx}")
            except Exception as e:
                logger.error(f"处理文件 {csv_file} 时发生错误: {e}")
                raise Exception(f"处理文件 {csv_file} 时发生错误: {e}")
        logger.info(f"处理完成")
        return 0

    front_rv_window()
    train_paths = [] # 用于存储train_path
    eval_results = [] # 用于存储各个eval_path及其结果
    train_statuses = [] # 用于存储各个train_path及其状态 用于后续显示
    # 搜索train_eval_path下层及更下层文件夹中同时拥有名字内含train的文件夹和含test文件夹的文件夹
    for train_eval_path in train_eval_paths:
        first_completed_date = datetime.datetime.now().strftime("%Y-%m-%d")
        # 结果xlsx文档，无论有没有过夜，生成的结果都在该文档 每次跑完个train路径就存一次 
        eval_result_xlsx = os.path.join(result_xlsx_dir, f"{first_completed_date}_{os.path.basename(train_eval_path)}.xlsx")

        # 清空列表以避免数据污染
        train_paths.clear()
        eval_results.clear()
        train_statuses.clear()
        dir_suffix = os.path.basename(os.path.normpath(train_eval_path))
        
        # 确保路径是 Unicode 格式
        train_eval_path = os.path.abspath(train_eval_path)
        
        for root, dirs, files in os.walk(train_eval_path):
            if root != train_eval_path:  # 忽略train_eval_path该级文件夹
                train_folder = None
                test_folder = None
                for d in dirs:
                    # 使用 Unicode 字符串处理
                    if 'train' in d:
                        train_folder = os.path.join(root, d)
                    if 'test' in d:
                        test_folder = os.path.join(root, d)
                if train_folder and test_folder:
                    if os.path.isdir(train_folder) and os.path.isdir(test_folder):
                        train_paths.append(train_folder)
        if not train_paths:
            raise Exception(f"{train_eval_path}下未读取到训练路径")
        logger.success(f"读取train_eval_path下所有的训练路径train_paths: {train_paths}")


        # 使用 检测pyd_path下是否有train_logs的文件夹,没有的话创建该文件夹,在文件夹内创建 年-月-日.csv(若不存在的话)
        train_logs_dir = os.path.join(pyd_path, 'train_logs')
        os.makedirs(train_logs_dir, exist_ok=True)
        train_control_csv = os.path.join(train_logs_dir, f"train_control_{dir_suffix}.csv")
        trained_success_list.clear()

        # 读取里边为训练完成的训练路径装入trained_success_list。
        if os.path.exists(train_control_csv):
            with open(train_control_csv, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                next(reader)  # 跳过表头
                for row in reader:
                    if row[1] == '训练完成':
                        trained_success_list.append(row[0])
        else:
            # 文件不存在的话把检索到的train_paths装入csv
            with open(train_control_csv, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['train路径', 'train状态'])
                for train_path in train_paths:
                    writer.writerow([train_path, '待训练'])
        # 是否跳过大循环
        should_continue_outer = True
        for train_path in train_paths:
            try:
                # 如果有训练完成的，则跳过
                if train_path in trained_success_list:
                    logger.info(f"训练路径{train_path}训练状态为完成，跳过")
                    continue
                should_continue_outer = False
                pyperclip.copy(train_path)
                logger.info(f"开始训练{train_path}")
                utils.click_by_png(config.RV_AI_SIMULATE_TO_TRAIN)
                if utils.search_symbol_erroring(config.RV_AI_TRAIN_SUCCESS, 180):
                    time.sleep(0.5)
                    pyautogui.press('enter')
                time.sleep(0.5)
                refresh_count = 0
                while refresh_count < 300:
                    if refresh_count == 1:
                        # 重新训练
                        time.sleep(1)
                        utils.click_by_png(config.RV_AI_TRAIN_STATUS, timeout=60, if_click_right=1)
                        time.sleep(0.5)
                        pyautogui.press('down', 8)
                        time.sleep(0.5)
                        pyautogui.press('enter')
                        time.sleep(2)
                    # 刷新
                    utils.click_by_png(config.RV_AI_JOB_NAME, timeout=60, if_click_right=1)
                    time.sleep(0.5)
                    pyautogui.press('down', 2)
                    pyautogui.press('enter')
                    logger.info(f"训练刷新第{refresh_count}次")
                    time.sleep(0.5)
                    # 查看训练状态
                    utils.click_by_png(config.RV_AI_TRAIN_STATUS)
                    time.sleep(0.5)
                    pyautogui.hotkey('ctrl', 'a')
                    time.sleep(0.5)
                    pyautogui.hotkey('ctrl', 'c')
                    train_result = pyperclip.paste()
                    logger.info(train_result)
                    if '训练完成' in train_result:
                        train_result = '训练完成'
                        train_statuses.append((train_path, train_result))
                        logger.info("训练状态: 训练完成")
                        break
                    elif '训练失败' in train_result:
                        train_result = '训练失败'
                        train_statuses.append((train_path, train_result))
                        logger.error("训练状态: 训练失败")
                        break
                    elif '待训练' in train_result:
                        refresh_count += 1
                        logger.info(f"训练状态: 待训练, 刷新次数: {refresh_count}")
                        if refresh_count == 1:
                            time.sleep(0.5)
                        else:
                            time.sleep(7)
                    else:
                        refresh_count = 1
                        logger.error("复制到的训练状态未知，重试中")
                        time.sleep(5)

                logger.info(f"{train_path}训练完毕，开始推理")
                
                # 查询train_path同级文件夹下名字包含test的文件夹 装入eval_path
                eval_path = glob.glob(os.path.join(os.path.dirname(train_path), '*test*'))
                if not eval_path:  # 检查列表是否为空
                    logger.error("未找到推理的路径")
                    raise Exception("未找到推理的路径")
                elif len(eval_path) > 1:
                    logger.error("推理文件夹有多个")
                    raise Exception("推理文件夹有多个")
                else:
                    eval_path = eval_path[0]
                    logger.info(f"训练路径为:{train_path}, 推理路径为:{eval_path}")
                    
                # 训练失败的话就不用推理了
                if train_result == '训练失败':
                    eval_result = '训练失败，因此未推理'
                    eval_results.append((eval_path, eval_result, good_ng))
                    break

                if mode == "normal":
                    good_ng = "空"
                    # 复制路径处理
                    if os.path.exists("temp_eval_path.txt"):
                        os.remove("temp_eval_path.txt")
                    with open("temp_eval_path.txt", "w", encoding='utf-8') as temp_file:
                        temp_file.write(eval_path)
                    # 从记事本读取eval_path到剪切板
                    with open("temp_eval_path.txt", "r", encoding='utf-8') as temp_file:
                        eval_path_from_txt = temp_file.read()
                        pyperclip.copy('')
                        time.sleep(2)
                        pyperclip.copy(eval_path_from_txt)
                    
                    # 确保剪切板内容正确
                    clipboard_content = pyperclip.paste()
                    for i in range(5):
                        pyperclip.copy(eval_path_from_txt)
                        time.sleep(2)
                        clipboard_content = pyperclip.paste()
                        logger.info(f"第{i+1}次复制检测，剪切板内容: {clipboard_content}，期望内容: {eval_path_from_txt}")
                        if clipboard_content == eval_path_from_txt:
                            break
                    
                    if clipboard_content != eval_path_from_txt:
                        logger.error("无法确保剪切板内容正确，操作中止")
                        raise Exception("剪切板内容不匹配")
                    
                    time.sleep(1)
                    logger.warning(f"剪切板内容: {clipboard_content}，准备点击推理")
                    time.sleep(1)
                    # 开始推理
                    utils.click_by_png(config.RV_AI_SIMULATE_TO_EVAL, tolerance=0.95)
                    logger.warning(f"点击完模拟至推理了，目前的剪切板内容：{pyperclip.paste()},先前的剪切板内容为：{clipboard_content}")
                    time.sleep(0.5)
                    os.remove("temp_eval_path.txt")
                    utils.search_symbol_erroring(config.RV_AI_EVAL_SUCCESS, 30)
                    time.sleep(0.5)
                    pyautogui.press('enter')
                    # 刷新任务状态并提取结果
                    utils.click_by_png(config.RV_AI_JOB_NAME, timeout=6, if_click_right=1)
                    time.sleep(0.5)
                    pyautogui.press('down', 2)
                    pyautogui.press('enter')
                    time.sleep(2)
                    utils.click_by_png(config.RV_AI_TRAIN_STATUS, if_click_right=1)
                    time.sleep(0.5)
                    if not utils.search_symbol(config.RV_AI_CLICK_RESTART_EVAL, 5):
                        time.sleep(5)
                    utils.click_by_png(config.RV_AI_CLICK_RESTART_EVAL)
                    click_time = datetime.datetime.now()
                    logger.info(f"点击重新推理时间: {click_time.strftime('%Y-%m-%d %H:%M:%S')}")
                    retry_count = 0
                    while retry_count < 300:
                        logger.info(f"推理刷新第{retry_count}次")
                        time.sleep(5)
                        utils.click_by_png(config.RV_AI_MISSION_MANAGE, tolerance=0.7)
                        time.sleep(0.5)
                        pyautogui.hotkey('ctrl', 'a')
                        time.sleep(1.5)
                        pyautogui.hotkey('ctrl', 'c')
                        logger.info("开始解析推理状态")
                        clipboard_content = pyperclip.paste()
                        if clipboard_content is None:
                            logger.error("复制到的推理状态为空")
                            time.sleep(10)
                            pyautogui.hotkey('ctrl', 'a')
                            time.sleep(1.5)
                            pyautogui.hotkey('ctrl', 'c')
                            clipboard_content = pyperclip.paste()
                        task_lines = clipboard_content.split('\n')
                        latest_time = None
                        latest_record = None
                        for line in task_lines[1:]:
                            parts = line.split('\t')
                            if len(parts) >= 6:
                                task_id, task_type, task_status, result, add_time, next_time = parts
                                # 将添加时间字符串转换为datetime对象
                                try:
                                    add_time_dt = datetime.datetime.strptime(add_time.strip(), '%Y/%m/%d %H:%M:%S')
                                except ValueError as e:
                                    logger.warning(f"日期解析失败：{add_time}，错误：{e}")
                                # 检查添加时间是否在click_time的前后2分钟内
                                if click_time - datetime.timedelta(minutes=2) <= add_time_dt <= click_time + datetime.timedelta(minutes=2):
                                    # 检查是否是最近的时间
                                    if latest_time is None or add_time_dt > latest_time:
                                        latest_time = add_time_dt
                                        latest_record = (task_type, task_status, result)
                                        logger.info(f"task_type = {task_type},task_status = {task_status},task_result = {result}")
                        if latest_record:
                            logger.info(f"第{retry_count}次推理,记录为：{latest_record}")
                            logger.info(f"找到最新记录：任务类型={task_type}，任务状态={task_status}，结果={result}，添加时间={add_time}")
                        else:
                            logger.error("没有找到符合条件的记录")
                        # 取两分钟内的记录，如果在队列中的话，则记录其task_id，3分钟后如果还是显示队列中则
                        if latest_record and '推理' in latest_record[0] and '成功' in latest_record[2]:
                            eval_result = '推理成功'
                            if train_result == '待训练':
                                eval_result = f"{train_result}, {eval_result}"
                            utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                            break
                        elif latest_record and '推理' in latest_record[0] and '失败' in latest_record[2]:
                            eval_result = '推理失败'
                            if train_result == '待训练':
                                eval_result = f"{train_result}, {eval_result}"
                            utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                            break
                        elif "队列" in latest_record[1]:
                            retry_count += 0.5
                            logger.info(f"推理处于队列中")
                            utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                        else:
                            utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                            time.sleep(10)  # 等待10秒后再次尝试
                            retry_count += 1

                    time.sleep(1)
                    eval_results.clear()
                    eval_results.append((eval_path, eval_result, good_ng))
                elif mode == "good_ng":
                    # 获取eval_path下的good和ng文件夹路径
                    good_path = os.path.join(eval_path, 'good')
                    ng_path = os.path.join(eval_path, 'ng')
                    # 分为good和ng两种情况
                    if os.path.exists(good_path):
                        logger.info(f"good存在: {good_path}")
                        good_ng = "good"
                        # 复制路径处理
                        if os.path.exists("temp_good_path.txt"):
                            os.remove("temp_good_path.txt")
                        with open("temp_good_path.txt", "w", encoding='utf-8') as temp_file:
                            temp_file.write(good_path)
                        # 从记事本读取eval_path到剪切板
                        with open("temp_good_path.txt", "r", encoding='utf-8') as temp_file:
                            good_path_from_txt = temp_file.read()
                            pyperclip.copy('')
                            time.sleep(2)
                            pyperclip.copy(good_path_from_txt)
                            
                        # 确保剪切板内容正确
                        clipboard_content = pyperclip.paste()
                        for i in range(5):
                            pyperclip.copy(good_path_from_txt)
                            time.sleep(2)
                            clipboard_content = pyperclip.paste()
                            logger.info(f"第{i+1}次复制检测，剪切板内容: {clipboard_content}，期望内容: {good_path_from_txt}")
                            if clipboard_content == good_path_from_txt:
                                break
                        
                        if clipboard_content != good_path_from_txt:
                            logger.error("无法确保剪切板内容正确，操作中止")
                            raise Exception("剪切板内容不匹配")
                        
                        time.sleep(1)
                        logger.warning(f"剪切板内容: {clipboard_content}，准备点击推理")
                        time.sleep(1)
                        # 开始推理
                        utils.click_by_png(config.RV_AI_SIMULATE_TO_EVAL, tolerance=0.95)
                        time.sleep(0.5)
                        os.remove("temp_good_path.txt")
                        utils.search_symbol_erroring(config.RV_AI_EVAL_SUCCESS, 30)
                        time.sleep(0.5)
                        pyautogui.press('enter')
                        # 刷新任务状态
                        utils.click_by_png(config.RV_AI_JOB_NAME, timeout=6, if_click_right=1)
                        time.sleep(0.5)
                        pyautogui.press('down',2)
                        pyautogui.press('enter')
                        # 点击重新推理 推理完成后再去点手动筛选
                        time.sleep(2)
                        utils.click_by_png(config.RV_AI_TRAIN_STATUS, if_click_right=1)
                        if not utils.search_symbol(config.RV_AI_CLICK_RESTART_EVAL, 5):
                            time.sleep(5)
                        utils.click_by_png(config.RV_AI_CLICK_RESTART_EVAL)
                        click_time = datetime.datetime.now()
                        logger.info(f"点击重新推理时间: {click_time.strftime('%Y-%m-%d %H:%M:%S')}")
                        # 重新推理，点击任务状态，获取eval_result
                        retry_count = 0
                        while retry_count < 300:
                            logger.info(f"推理刷新第{retry_count}次")
                            time.sleep(5)
                            utils.click_by_png(config.RV_AI_MISSION_MANAGE, tolerance=0.7)
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl', 'a')
                            time.sleep(1)
                            pyautogui.hotkey('ctrl', 'c')
                            logger.info("开始解析推理状态")
                            clipboard_content = pyperclip.paste()
                            task_lines = clipboard_content.split('\n')
                            latest_time = None
                            latest_record = None
                            # 解析任务状态
                            for line in task_lines[1:]:  # 忽略第一行标题
                                parts = line.split('\t')
                                if len(parts) >= 6:
                                    task_id, task_type, task_status, result, add_time, next_time = parts
                                    # 将添加时间字符串转换为datetime对象
                                    try:
                                        add_time_dt = datetime.datetime.strptime(add_time.strip(), '%Y/%m/%d %H:%M:%S')
                                    except ValueError as e:
                                        logger.warning(f"日期解析失败：{add_time}，错误：{e}")
                                    # 检查添加时间是否在click_time的前后2分钟内
                                    if click_time - datetime.timedelta(minutes=2) <= add_time_dt <= click_time + datetime.timedelta(minutes=2):
                                        # 检查是否是最近的时间
                                        if latest_time is None or add_time_dt > latest_time:
                                            latest_time = add_time_dt
                                            latest_record = (task_type, task_status, result)
                                            logger.info(f"task_type = {task_type},task_status = {task_status},task_result = {result}")
                            if latest_record:
                                logger.info(f"第{retry_count}次推理,记录为：{latest_record}")
                                logger.info(f"找到最新记录：任务类型={task_type}，任务状态={task_status}，结果={result}，添加时间={add_time}")
                            else:
                                logger.error("没有找到符合条件的记录")
                            # 取两分钟内的记录，如果在队列中的话，则记录其task_id，3分钟后如果还是显示队列中则
                            if latest_record and '推理' in latest_record[0] and '成功' in latest_record[2]:
                                eval_result = '推理成功'
                                if train_result == '待训练':
                                    eval_result = f"{train_result}, {eval_result}"
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                break
                            elif latest_record and '推理' in latest_record[0] and '失败' in latest_record[2]:
                                eval_result = '推理失败'
                                if train_result == '待训练':
                                    eval_result = f"{train_result}, {eval_result}"
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                break
                            elif "队列" in latest_record[1]:
                                retry_count += 0.5
                                logger.info(f"推理处于队列中")
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                            else:
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                time.sleep(10)  # 等待10秒后再次尝试
                                retry_count += 1
                        
                        time.sleep(1)
                        eval_results.clear()
                        eval_results.append((good_path, eval_result, good_ng))
                        # 推理完成后，待确认列表，选中图片，同板批量删除
                        if utils.search_symbol(config.RV_AI_CONFIRM_NO):
                            utils.click_by_png(config.RV_AI_CONFIRM_NO)
                        utils.click_by_png(config.RV_AI_NO_CONFIRMED)
                        pyautogui.rightClick()
                        pyautogui.press('up', presses=3)
                        pyautogui.press('enter')
                        time.sleep(1)


                    if os.path.exists(ng_path):
                        logger.info(f"ng存在: {ng_path}")
                        good_ng = "ng"
                        # 复制路径处理
                        if os.path.exists("temp_ng_path.txt"):
                            os.remove("temp_ng_path.txt")
                        with open("temp_ng_path.txt", "w", encoding='utf-8') as temp_file:
                            temp_file.write(ng_path)
                        # 从记事本读取eval_path到剪切板
                        with open("temp_ng_path.txt", "r", encoding='utf-8') as temp_file:
                            ng_path_from_txt = temp_file.read()
                            pyperclip.copy('')
                            time.sleep(2)
                            pyperclip.copy(ng_path_from_txt)

                        # 确保剪切板内容正确
                        clipboard_content = pyperclip.paste()
                        for i in range(5):
                            pyperclip.copy(ng_path_from_txt)
                            time.sleep(2)
                            clipboard_content = pyperclip.paste()
                            logger.info(f"第{i+1}次复制检测，剪切板内容: {clipboard_content}，期望内容: {ng_path_from_txt}")
                            if clipboard_content == ng_path_from_txt:
                                break
                        
                        if clipboard_content != ng_path_from_txt:
                            logger.error("无法确保剪切板内容正确，操作中止")
                            raise Exception("剪切板内容不匹配")
                        
                        time.sleep(1)
                        logger.warning(f"剪切板内容: {clipboard_content}，准备点击推理")
                        time.sleep(1)  
                        # 开始推理
                        utils.click_by_png(config.RV_AI_SIMULATE_TO_EVAL, tolerance=0.95)
                        time.sleep(0.5)
                        os.remove("temp_ng_path.txt")
                        utils.search_symbol_erroring(config.RV_AI_EVAL_SUCCESS, 30)
                        time.sleep(1)
                        pyautogui.press('enter')
                        # 刷新任务状态
                        utils.click_by_png(config.RV_AI_JOB_NAME, timeout=6, if_click_right=1)
                        time.sleep(0.5)
                        pyautogui.press('down',2)
                        pyautogui.press('enter')
                        # 点击重新推理
                        time.sleep(2)
                        utils.click_by_png(config.RV_AI_TRAIN_STATUS, if_click_right=1)
                        if not utils.search_symbol(config.RV_AI_CLICK_RESTART_EVAL, 5):
                            time.sleep(5)
                        utils.click_by_png(config.RV_AI_CLICK_RESTART_EVAL)
                        click_time = datetime.datetime.now()
                        logger.info(f"点击重新推理时间: {click_time.strftime('%Y-%m-%d %H:%M:%S')}")
                        # 获取eval_result
                        retry_count = 0
                        while retry_count < 300:
                            logger.info(f"推理刷新第{retry_count}次")
                            time.sleep(5)
                            utils.click_by_png(config.RV_AI_MISSION_MANAGE, tolerance=0.7)
                            time.sleep(0.5)
                            pyautogui.hotkey('ctrl', 'a')
                            time.sleep(1)
                            pyautogui.hotkey('ctrl', 'c')
                            logger.info("开始解析推理状态")
                            clipboard_content = pyperclip.paste()
                            task_lines = clipboard_content.split('\n')
                            latest_time = None
                            latest_record = None
                            # 解析任务状态
                            for line in task_lines[1:]:  # 忽略第一行标题
                                parts = line.split('\t')
                                if len(parts) >= 6:
                                    task_id, task_type, task_status, result, add_time, next_time = parts
                                    # 将添加时间字符串转换为datetime对象
                                    try:
                                        add_time_dt = datetime.datetime.strptime(add_time.strip(), '%Y/%m/%d %H:%M:%S')
                                    except ValueError as e:
                                        logger.warning(f"日期解析失败：{add_time}，错误：{e}")
                                    # 检查添加时间是否在click_time的前后2分钟内
                                    if click_time - datetime.timedelta(minutes=2) <= add_time_dt <= click_time + datetime.timedelta(minutes=2):
                                        # 检查是否是最近的时间
                                        if latest_time is None or add_time_dt > latest_time:
                                            latest_time = add_time_dt
                                            latest_record = (task_type, task_status, result)
                                            logger.info(f"task_type = {task_type},task_status = {task_status},task_result = {result}")
                            if latest_record:
                                logger.info(f"第{retry_count}次推理,记录为：{latest_record}")
                                logger.info(f"找到最新记录：任务类型={task_type}，任务状态={task_status}，结果={result}，添加时间={add_time}")
                            else:
                                logger.error("没有找到符合条件的记录")
                            # 取两分钟内的记录，如果在队列中的话，则记录其task_id，3分钟后如果还是显示队列中则
                            if latest_record and '推理' in latest_record[0] and '成功' in latest_record[2]:
                                eval_result = '推理成功'
                                if train_result == '待训练':
                                    eval_result = f"{train_result}, {eval_result}"
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                break
                            elif latest_record and '推理' in latest_record[0] and '失败' in latest_record[2]:
                                eval_result = '推理失败'
                                if train_result == '待训练':
                                    eval_result = f"{train_result}, {eval_result}"
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                break
                            elif "队列" in latest_record[1]:
                                retry_count += 0.5
                                logger.info(f"推理处于队列中")
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                            else:
                                utils.click_by_png(config.RV_AI_CLOSE_MISSION_MANAGE, timeout=5)
                                time.sleep(10)  # 等待10秒后再次尝试
                                retry_count += 1

                        time.sleep(1)
                        eval_results.clear()
                        eval_results.append((ng_path, eval_result, good_ng))



            except Exception as e:
                logger.error(f"在处理{train_path}时发生错误: {e}")
                raise Exception(f"在处理{train_path}时发生错误: {e}")                    

            # 删除job
            utils.click_by_png(config.RV_AI_JOB_NAME, if_click_right=1)
            pyautogui.press('down')
            pyautogui.press('enter')


            have_trained_before = False
            final_completed_txt_path = None
            # tep下小路径训练-推理结束，获得了eval_results。将训练推理生成的csv转为xlsx，并将eval_results装至xlsx。
            try:
                res_static_dir = os.path.join(pyd_path, "logs", "res_static")
                csv_files = glob.glob(os.path.join(res_static_dir, "*.csv"))
                if not csv_files:
                    logger.error("res_static目录中未找到任何csv文件")
                    raise Exception("res_static目录中未找到任何csv文件")
                # 找到最新的csv
                eval_result_newest_csv = max(csv_files, key=os.path.getctime)
                if not os.path.exists(eval_result_newest_csv):
                    logger.error("未找到推理后生成的csv文件")
                    raise Exception("未找到推理后生成的csv文件")
                if len(train_eval_paths) == 1 or train_eval_path == train_eval_paths[-1]:
                    completed_txt_path = os.path.join(result_xlsx_dir, f"{os.path.basename(train_eval_path)}_脚本已运行完成.txt")
                    final_completed_txt_path = os.path.join(result_xlsx_dir, "脚本已运行完成.txt")
                else:
                    completed_txt_path = os.path.join(result_xlsx_dir, f"{os.path.basename(train_eval_path)}_脚本已运行完成.txt")
                    final_completed_txt_path = os.path.join(result_xlsx_dir, "脚本已运行完成.txt")
                if os.path.exists(completed_txt_path):
                    os.remove(completed_txt_path)
                if final_completed_txt_path and os.path.exists(final_completed_txt_path):
                    os.remove(final_completed_txt_path)
                logger.info(f"推理结果生成路径: {eval_result_xlsx}")

                # 如果xlsx不存在 则复制csv为xlsx 
                if not os.path.exists(eval_result_xlsx):
                    with open(eval_result_newest_csv, 'rb') as file:
                        content = file.read()
                        result = chardet.detect(content)
                        encoding = result['encoding'] if result['encoding'] else 'utf-8'
                        # 尝试使用探测到的编码打开文件，如果失败则使用GBK编码尝试
                        try:
                            content = content.decode(encoding)
                        except UnicodeDecodeError as e:
                            logger.error(f"编码解码错误: {e}")
                            content = content.decode('gbk', errors='ignore')

                    # 使用探测到的编码读取文件
                    try:
                        df_csv = pd.read_csv(eval_result_newest_csv, encoding=encoding)
                    except UnicodeEncodeError as e:
                        logger.error(f"读取CSV文件时发生编码错误: {e}")
                        raise Exception(f"读取CSV文件时发生编码错误: {e}")

                    # 直接保存为XLSX，不对CSV进行列调整
                    try:
                        # 在XLSX中添加新列
                        df_xlsx = df_csv.assign(eval结果=None, good_ng=None, eval图片=None, 定位图片=None, train图片=None, eval路径=None)
                        # 重新排序列，将新列放在前面
                        columns_order = ['eval结果', 'good_ng', 'eval图片', '定位图片', 'train图片', 'eval路径'] + [col for col in df_xlsx.columns if col not in ['eval结果', 'good_ng', 'eval图片', '定位图片', 'train图片', 'eval路径']]
                        df_xlsx = df_xlsx[columns_order]
                        df_xlsx.to_excel(eval_result_xlsx, index=False, engine='openpyxl')
                    except Exception as e:
                        logger.error(f"保存为XLSX时发生错误: {e}")
                        raise Exception(f"保存为XLSX时发生错误: {e}")

                    logger.info("对xlsx作预处理")
                    # 打开xlsx
                    xlsx = pd.read_excel(eval_result_xlsx, engine='openpyxl')
                    # 去除列名中的空白符
                    xlsx.columns = [col.strip() for col in xlsx.columns]
                    # 使用openpyxl直接加载工作簿以修改列名样式
                    wb = openpyxl.load_workbook(eval_result_xlsx)
                    ws = wb.active
                    # 取消列名加粗
                    for cell in ws[1]:  # 第一行是列名
                        cell.font = Font(bold=False)
                    # 设置B列和D列的列宽为18字符
                    ws.column_dimensions['C'].width = 18
                    ws.column_dimensions['D'].width = 18
                    ws.column_dimensions['E'].width = 18
                    # 保存修改后的工作簿
                    wb.save(eval_result_xlsx)
                    # 重新加载DataFrame以确保列名更改生效
                    xlsx = pd.read_excel(eval_result_xlsx, engine='openpyxl')
                    # 去除列名中的空白符和其他特殊字符
                    xlsx.columns = [re.sub(r'\s+', '', col) for col in xlsx.columns]
                    logger.info("推理结束后对xlsx作预处理完成，此后每条train路径完成后都会进行一次同步")
                    
                    # 新增完xlsx后删除csv的原有数据，保留标题行
                    df_csv.iloc[0:0].to_csv(eval_result_newest_csv, index=False, encoding='utf-8-sig')
                    logger.info("已删除csv中的原有数据，仅保留标题行")
                else:
                    have_trained_before = True
                    if os.path.exists(eval_result_newest_csv):
                        try:
                            sync_csv_to_xlsx(eval_result_newest_csv, eval_result_xlsx)
                        except Exception as e:
                            logger.error(f"同步CSV到XLSX时发生错误: {e}")
                            raise

            except Exception as e:
                logger.error(f"转csv为xlsx并同步数据至xlsx时发生错误: {e}")
                raise Exception(f"转csv为xlsx并同步数据至xlsx时发生错误: {e}")
            def clean_id(id_str):
                return str(id_str).strip().replace('\n', '').replace('\r', '').replace('\t', '')

            logger.info("开始将推理结果同步至xlsx")
            if not os.path.exists(eval_result_xlsx):
                logger.warning(f"文件 {eval_result_xlsx} 不存在，跳过")
                continue

            wb = openpyxl.load_workbook(eval_result_xlsx)
            ws = wb.active
            logger.info(f"工作簿和工作表已初始化: {eval_result_xlsx}")

            # 更新推理结果至eval_xlsx
            xlsx = pd.read_excel(eval_result_xlsx, engine='openpyxl')
            xlsx.columns = [col.strip() for col in xlsx.columns]
            xlsx.columns = [re.sub(r'\s+', '', col) for col in xlsx.columns]

            logger.info(f"开始将单条tep推理结果同步至eval_xlsx: {eval_result_xlsx}")
            logger.warning(f"eval_results: {eval_results}")
            # 将eval_results内的数据同步至eval_xlsx 因为是根据id对应，且考虑到隔夜，对开始时间以及开始时间的后一天对应的文件都去匹配
            for eval_path, eval_result, good_ng in eval_results:
                eval_image_paths = []
                eval_image_paths.extend(glob.glob(os.path.join(eval_path, '**/*.bmp'), recursive=True))
                eval_image_paths.extend(glob.glob(os.path.join(eval_path, '**/*.jpg'), recursive=True))
                eval_image_paths.extend(glob.glob(os.path.join(eval_path, '**/*.png'), recursive=True))
                logger.info(f"推理路径下找到的图片路径为: {eval_image_paths}")
                # 图片名称即id。通过id查找匹配的行 给对应行赋值
                for eval_image_path in eval_image_paths:
                    eval_image_name = clean_id(os.path.splitext(os.path.basename(eval_image_path))[0])
                    logger.info(f"处理的图片id为: {eval_image_name}")

                    try:
                        # 打印调试信息
                        logger.warning(f"待匹配的eval_image_name: '{eval_image_name}'")
                        logger.warning(f"所有行中的id值: {xlsx['id'].tolist()}")
                        logger.info("除标题行外所有行内容如下：")
                        for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始打印
                            logger.info(row)
                        # 除了标题行外 每一个id值不为空的行都要log
                        for index, row in xlsx.iterrows():
                            if pd.notna(row['id']):
                                logger.info(f"行 {index + 2} 的id值: {row['id']}")
                        logger.warning(f"所有行中的id值: {xlsx['id'].tolist()}")

                        # 确保id列为字符串类型并去除空格、换行符、回车符和制表符
                        xlsx['id'] = xlsx['id'].apply(clean_id).astype(str)
                        
                        # 搜索所有id列有值的行（除了标题行外）
                        for index, row in xlsx.iterrows():
                            if pd.notna(row['id']):
                                logger.info(f"行 {index + 2} 的id列数据类型: {type(row['id'])}")

                        # 使用字符串包含关系进行匹配
                        matching_rows = xlsx[xlsx['id'].astype(str).str.contains(eval_image_name, na=False)]
                        if not matching_rows.empty:
                            for index in matching_rows.index:
                                excel_row = index + 2
                                logger.info(f"匹配到的行号：{index + 1}")
                                logger.info(f"更新eval_results至xslx内查找到的对应行：eval_path: {eval_path}, eval_image_path: {eval_image_path}, eval_image_name: {eval_image_name}, eval_result: {eval_result}")
                                ws.cell(row=excel_row, column=1, value=eval_result)
                                ws.cell(row=excel_row, column=2, value=good_ng)
                                ws.cell(row=excel_row, column=3, value=eval_image_path)
                                ws.cell(row=excel_row, column=6, value=eval_image_path)
                                logger.info(f"已匹配到第 {index + 1} 行")
                            wb.save(eval_result_xlsx)
                        else:
                            logger.warning(f"未找到匹配的行，id为: {eval_image_name}")
                    except Exception as e:
                        logger.error(f"更新eval_results至xslx时发生错误: {e}")

                logger.info("推理结束，开始记录训练结果（该路径已 训练-推理 完成）")
                try:
                    # 训练-推理完后处理train.csv，填入train路径及train状态，以达到控制指定路径训练
                    logger.info(f"读取训练日志文件: {train_control_csv}")
                    df_train_logs = pd.read_csv(train_control_csv, encoding='utf-8')
                    logger.info(f"训练日志文件内容: {df_train_logs}")
                    
                    if train_path in df_train_logs['train路径'].values:
                        logger.info(f"更新训练路径 {train_path} 的训练状态为 {train_result}")
                        df_train_logs.loc[df_train_logs['train路径'] == train_path, 'train状态'] = train_result
                        # 将train状态的单元格标注为黄色
                        for index, row in df_train_logs.iterrows():
                            if row['train路径'] == train_path:
                                df_train_logs.at[index, 'train状态'] = train_result
                        df_train_logs.to_csv(train_control_csv, index=False, encoding='utf-8')
                        logger.info(f"训练路径 {train_path} 的训练状态已更新并保存")
                    else:
                        logger.info(f"训练路径 {train_path} 不存在于训练日志中，添加新记录")
                        new_row = pd.DataFrame([[train_path, train_result]], columns=['train路径', 'train状态'])
                        df_train_logs = pd.concat([df_train_logs, new_row], ignore_index=True)
                        df_train_logs.to_csv(train_control_csv, index=False, encoding='utf-8')
                        logger.info(f"新记录已添加并保存到训练日志文件")
                    
                    logger.info("训练日志文件已关闭")
                except Exception as e:
                    logger.error(f"记录训练结果时发生错误: {e}")
                    raise Exception(f"记录训练结果时发生错误: {e}")
                finally:
                    logger.info("记录训练结果操作完成")
        # 所有路径都被跳过，则进入下一个tep
        if should_continue_outer:
            logger.warning("所有路径都被跳过，进入下一个tep")
            continue
        # 如果有多条train_eval_path（tep）的话 每条tep每次把底下te小路径训练推理完后装入xlsx内。每个tep只有1个xlsx。
        # 单条tep装入完后再排序和插入图 
        if os.path.exists(eval_result_xlsx):
            logger.info(f"开始给结果文档排序和插入图片: {eval_result_xlsx}")
            wb = openpyxl.load_workbook(eval_result_xlsx)
            ws = wb.active
            
            if have_trained_before:
                # logger.info(f"检测到之前已训练，开始删除xlsx内所有图片...删除前图片数量：{len(ws._images)}")
                ws._images = [image for image in ws._images if image.anchor._from.col in [2, 3, 4]]
                logger.critical(f"图片数量: {len(ws._images)}")
                # logger.info("已删除所有在C, D, E列的图片")
                # logger.info(f"删除后图片数量: {len(ws._images)}")

            data = list(ws.values)
            header = data[0]  # 保存标题行
            non_empty_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
            logger.info(f"xlsx内有数据的行数: {len(non_empty_rows)}")
            # 根据 id 列名排序，忽略标题行，并去重
            seen = set()
            unique_data = []
            for row in sorted(data[1:], key=lambda x: x[header.index('id')]):
                if row[header.index('id')] not in seen:
                    unique_data.append(row)
                    seen.add(row[header.index('id')])
            for row_idx, row in enumerate([header] + unique_data, 1):
                for col_idx, value in enumerate(row, 1):
                    if value is None:
                        ws.cell(row=row_idx, column=col_idx).value = None
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
            logger.info("行数据已排序，正在保存工作簿...")
            wb.save(eval_result_xlsx)

            # 重新加载工作簿以验证数据
            logger.info("排序完毕，开始处理图片")
            # 单独处理图片
            for index, row in pd.DataFrame(ws.values).iterrows():
                excel_row = index + 2
                if excel_row > 1:
                    ws.row_dimensions[excel_row].height = 75
                id = ws.cell(row=excel_row, column=8).value

                if id is None:
                    logger.info(f"行 {excel_row} 的id为None，跳过图片处理")
                    continue

                # 处理eval图片
                eval_image_cell = ws.cell(row=excel_row, column=3)
                eval_image_path = eval_image_cell.value
                try:
                    if eval_image_path:
                        eval_image_path = os.path.normpath(eval_image_path)
                        if os.path.exists(eval_image_path):
                            utils.insert_image_limited(ws, eval_image_cell, eval_image_path)
                        else:
                            eval_image_cell.value = "空"
                            logger.info(f"id: {id}, eval图片路径不存在: {eval_image_path}")
                    else:
                        eval_image_cell.value = "空"
                except Exception as e:
                    logger.info(f"id: {id}, 处理eval图片 {eval_image_path} 时出错: {e}")

                # 处理定位图片
                cad_cell = ws.cell(row=excel_row, column=4)
                cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.bmp")
                try:
                    if not os.path.exists(cad_loc_path):
                        cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.jpg")
                    if not os.path.exists(cad_loc_path):
                        cad_loc_path = os.path.join(pyd_path, 'cad_com_loc', f"{id}.png")
                    cad_loc_path = os.path.normpath(cad_loc_path)
                    if cad_loc_path and os.path.exists(cad_loc_path):
                        utils.insert_image_limited(ws, cad_cell, cad_loc_path)
                    else:
                        cad_cell.value = "空"
                        logger.error(f"id: {id}, cad_loc图片路径不存在: {cad_loc_path}")
                except Exception as e:
                    logger.error(f"id: {id}, cad_img处理图片 {cad_loc_path} 时出错: {e}")

                # 处理train_img_path
                if ws.cell(row=excel_row, column=10).value is None:
                    ws.cell(row=excel_row, column=5).value = "空"
                else:
                    train_image_path = ws.cell(row=excel_row, column=10).value
                    train_cell = ws.cell(row=excel_row, column=5)
                    try:
                        logger.info(f"id: {id}, train_image_path: {train_image_path}")
                        if train_image_path and os.path.exists(train_image_path):
                            train_image_path = os.path.normpath(train_image_path)
                            utils.insert_image_limited(ws, train_cell, train_image_path)
                        else:
                            train_cell.value = "空"
                    except Exception as e:
                        logger.info(f"id: {id}, train_img处理图片 {train_image_path} 时出错: {e}")

            # 处理完所有图片后再次保存工作簿
            logger.info("所有图片处理完毕，正在保存工作簿...")
            wb.save(eval_result_xlsx)

        with open(completed_txt_path, "w") as f:
            pass
        if len(train_eval_paths) == 1 or train_eval_path == train_eval_paths[-1]:
            with open(final_completed_txt_path, "w") as f:
                pass

    logger.info(f"ai复判完成")
    logger.info(f"train_statuses: {train_statuses}")
    return 0, train_statuses

