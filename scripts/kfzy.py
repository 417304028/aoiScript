import datetime
import os
import queue
import re
import subprocess
import threading
import time
import win32api
import win32con
import pynput
from tkinter import messagebox
import ctypes
import pandas as pd
import psutil,win32gui
import pyautogui
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from ctypes import Structure, c_long, c_ulong, sizeof, byref
from ctypes import c_uint64  # 可用于64位时

import config
import utils
def a_zhaolu_loop_open():
    logger.info("开始检查并启动AOI程序...")
    pyautogui.hotkey('win', 'd')
    loop_count = 0
    while True:
        loop_count += 1
        logger.info(f"循环次数: {loop_count}")
        # AOI存在的话先杀掉
        utils.minimize_service_process_manager()
        for proc in psutil.process_iter(['pid', 'name']):
            if "AOI.exe" in proc.info['name']:
                try:
                    logger.info(f"检测到AOI进程: {proc.info['name']} (PID: {proc.info['pid']})，准备杀掉...")
                    parent_proc = psutil.Process(proc.info['pid'])
                    parent_proc.kill()
                    parent_proc.wait()  # 确认父进程被杀掉
                    
                    # 再次检查进程是否已被杀掉
                    if not psutil.pid_exists(proc.info['pid']):
                        logger.info(f"确认进程 {proc.info['name']} (PID: {proc.info['pid']}) 已被成功杀掉")
                    else:
                        raise Exception(f"进行杀掉AOI进程操作之后，进程 {proc.info['name']} (PID: {proc.info['pid']}) 仍在运行")
                except Exception as kill_error:
                    logger.error(f"无法杀掉进程 {proc.info['name']} (PID: {proc.info['pid']}): {kill_error}")
        aoi_running = any("AOI.exe" == p.name() for p in psutil.process_iter())
        # aoi不存在的话，启动aoi
        if not aoi_running:
            logger.info("AOI程序未运行，正在启动...")
            # 使用命令行启动AOI程序
            current_dir = os.getcwd()
            os.chdir(os.path.dirname(config.AOI_EXE_PATH))
            subprocess.Popen(f'cmd /c "{os.path.basename(config.AOI_EXE_PATH)}"', shell=True)
            # 切换回原来的工作目录
            os.chdir(current_dir)
            logger.info("等待AOI程序启动...")
            failure_count = 0
            while True:
                if utils.search_symbol(config.LOGINING, 5, tolerance=0.5):
                    logger.info("检测到登录窗口")
                    failure_count = 0
                else:
                    failure_count += 1
                    logger.warning(f"未检测到登录窗口，累计失败次数: {failure_count}")
                    utils.minimize_service_process_manager()
                    if failure_count >= 2:
                        break
                time.sleep(3)
            if not utils.search_symbol(config.LOGINING, 3, tolerance=0.5) and not utils.search_symbol(config.AOI_TOPIC, 3, tolerance=0.6):
                # 都没找到，记录
                try:
                    # 获取当前时间
                    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    # 截取屏幕截图
                    screenshot = pyautogui.screenshot()
                    # 将截图缩放为原图的0.5倍大小
                    original_width, original_height = screenshot.size
                    new_size = (int(original_width * 0.5), int(original_height * 0.5))
                    screenshot = screenshot.resize(new_size)
                    # 暂时保存截图为临时文件
                    temp_image_path = "temp_screenshot.png"
                    screenshot.save(temp_image_path)
                    
                    # 将当前时间和截图插入到 error.xlsx 中（如果文件已存在则插入下一行）
                    error_file = "error.xlsx"
                    if os.path.exists(error_file):
                        wb = load_workbook(error_file)
                        ws = wb.active
                    else:
                        wb = Workbook()
                        ws = wb.active
                    # 找到下一行
                    next_row = ws.max_row + 1
                    ws.cell(row=next_row, column=1, value=current_time)
                    from openpyxl.drawing.image import Image as XLImage
                    img = XLImage(temp_image_path)
                    ws.add_image(img, f"B{next_row}")
                    wb.save(error_file)
                    logger.error("已将错误信息记录到 error.xlsx")
                    # 删除临时图片文件
                    os.remove(temp_image_path)
                except Exception as e:
                    logger.error(f"记录错误信息失败: {e}")
        else:
            logger.critical("AOI进程仍未杀掉")

    
def play_and_stop():
    utils.click_by_png(config.PLAY, 2)
    for _ in range (3):
        time.sleep(6)
        pyautogui.press("enter")
        time.sleep(5)
    utils.search_symbol_erroring(config.TESTING_INTERFACE_INFORMATION,100,tolerance=0.75)
    start_time = time.time()
    while True:
        if utils.search_symbol(config.TESTING_INTERFACE_PERCENT_100, 5, tolerance=0.8):
            break
        if time.time() - start_time > 300:
            raise Exception("循环单次时疑似超过五分钟")
    if utils.search_symbol(config.QUESTION_MARK, 3, tolerance=0.75):
        utils.click_by_png(config.CLOSE,timeout=3)
    else:
        if not utils.click_by_ocr("停止"):
            raise Exception("未能识别到停止按钮")


def test_click_2():
    # 方法2：使用 win32gui 发送鼠标消息（基于窗口句柄及客户端坐标转换）
    coordinate = (1400, 860)
    click_delay = 0.2  # 点击延时（秒）
    logger.debug(f"目标屏幕坐标: {coordinate}，点击延时: {click_delay}秒")
    try:
        hwnd = win32gui.WindowFromPoint(coordinate)
        logger.debug(f"通过 win32gui.WindowFromPoint 获得的窗口句柄: {hwnd}")
        if hwnd:
            # 为确保消息能被正确处理，先将窗口置于前台
            try:
                win32gui.SetForegroundWindow(hwnd)
                logger.debug(f"将窗口句柄 {hwnd} 设置为前台窗口")
            except Exception as ex:
                logger.warning(f"设置前台窗口失败: {ex}")
            window_title = win32gui.GetWindowText(hwnd)
            logger.info(f"成功获取窗口句柄: {hwnd}，窗口标题: {window_title}")
            
            # 转换屏幕坐标为客户端坐标
            client_point = win32gui.ScreenToClient(hwnd, coordinate)
            logger.debug(f"转换后的客户端坐标: {client_point}")
            
            # 计算 lParam 参数：低16位存 x，高16位存 y（基于客户端坐标）
            lParam = (client_point[1] << 16) | (client_point[0] & 0xFFFF)
            logger.debug(f"计算得到 lParam = {lParam} (x: {client_point[0]}, y: {client_point[1]})")
            
            WM_LBUTTONDOWN = 0x0201
            WM_LBUTTONUP = 0x0202
            logger.info(f"发送 WM_LBUTTONDOWN 消息到窗口句柄 {hwnd}")
            win32gui.SendMessage(hwnd, WM_LBUTTONDOWN, 0x00000001, lParam)
            time.sleep(click_delay)
            logger.info(f"发送 WM_LBUTTONUP 消息到窗口句柄 {hwnd}")
            win32gui.SendMessage(hwnd, WM_LBUTTONUP, 0, lParam)
        else:
            logger.error("未能通过 win32gui 获取有效窗口句柄")
    except Exception as e:
        logger.error(f"win32gui 点击出错：{e}")

def test_click_3():
    # 方法3：使用 win32api.mouse_event 模拟鼠标点击
    coordinate = (1400, 860)
    click_delay = 0.2  # 点击延时（秒）
    logger.debug(f"使用 win32api.mouse_event 在坐标 {coordinate} 处点击，延时 {click_delay} 秒")
    try:
        # 将鼠标光标移动到目标位置
        win32api.SetCursorPos(coordinate)
        logger.debug(f"鼠标定位到 {coordinate}")
        # 模拟鼠标左键按下
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0)
        time.sleep(click_delay)
        # 模拟鼠标左键释放
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0)
        logger.info("win32api.mouse_event 点击操作成功")
    except Exception as e:
        logger.error(f"win32api.mouse_event 点击出错：{e}")

def test_click_4():
    # 方法4：使用 pynput 模拟鼠标点击操作
    from pynput.mouse import Controller, Button
    coordinate = (1400, 860)
    click_delay = 0.2  # 点击延时（秒）
    logger.debug(f"使用 pynput 在坐标 {coordinate} 处点击，延时 {click_delay} 秒")
    try:
        mouse = Controller()
        mouse.position = coordinate
        time.sleep(0.1)
        mouse.press(Button.left)
        time.sleep(click_delay)
        mouse.release(Button.left)
        logger.info("pynput 点击操作成功")
    except Exception as e:
        logger.error(f"pynput 点击出错：{e}")




def check_export():
    utils.check_export_ok(if_export_all_ok=True)
def play():
    utils.open_program(if_specific=True)
    utils.is_checked((66,255),(78,267),True)
    utils.is_checked((84,273),(96,285),True)
    utils.click_by_png(config.PLAY, 2)
    for _ in range(3):
        time.sleep(2)
        pyautogui.press("enter")
        time.sleep(2)
    if not utils.search_symbol(config.TESTING_INTERFACE_INFORMATION,timeout=30,tolerance=0.8):
        raise Exception("未检测到测试板界面的特定标识，疑似能未进入测试板界面")
    time.sleep(3)
# 新增：加载并保存Excel时的重试函数，解决文件被占用问题
def load_workbook_with_retry(excel_path, retries=5, delay=1):
    for attempt in range(retries):
        try:
            wb = load_workbook(excel_path)
            return wb
        except PermissionError as e:
            logger.error(f"加载Excel文件 {excel_path} 失败，可能被占用，尝试第 {attempt+1} 次等待 {delay} 秒。错误：{e}")
            time.sleep(delay)
    raise Exception(f"无法加载Excel文件 {excel_path}，文件可能被占用")

def save_workbook_with_retry(wb, excel_path, retries=5, delay=1):
    for attempt in range(retries):
        try:
            wb.save(excel_path)
            return
        except PermissionError as e:
            logger.error(f"保存Excel文件 {excel_path} 失败，可能被占用，尝试第 {attempt+1} 次等待 {delay} 秒。错误：{e}")
            time.sleep(delay)
    raise Exception(f"无法保存Excel文件 {excel_path}，文件可能被占用")

# 获取当前脚本或可执行文件所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))

# 获取base_dir的上级的上级文件夹
parent_dir = os.path.dirname(os.path.dirname(base_dir))

# 在程序开始时记录初始日期
initial_date = datetime.datetime.now().strftime('%Y-%m-%d')

# 全局变量
alg_cal_times_data = []

# 用于在线程间传递内存数据
memory_queue = queue.Queue()

PROCESS_PER_MONITOR_DPI_AWARE = 2
ctypes.windll.shcore.SetProcessDpiAwareness(PROCESS_PER_MONITOR_DPI_AWARE)

@utils.screenshot_error_to_excel()
def launch_aoi():
    utils.check_and_launch_aoi()
    
@utils.screenshot_error_to_excel()
def launch_rv():
    utils.check_and_launch_rv()

@utils.screenshot_error_to_excel()
def launch_spc():
    utils.check_and_launch_spc()

def click_board_by_png():
    utils.click_by_png(config.TEST_BOARD)

def click_test_board_by_win32gui(timeout=5, tolerance=0.8):
    """
    通过win32gui.PostMessage查找固定参数config.TEST_BOARD图像的坐标并发送左键点击消息。

    参数:
        timeout (int): 超时时间，单位秒。默认为20秒。
        tolerance (float): 查找图像时的置信度阈值，默认为0.8。

    异常:
        如果在timeout内未能查找到图像或点击失败，将抛出异常。
    """
    import time
    import pyautogui
    import win32gui
    import win32process
    import config  # 假定config模块中存在TEST_BOARD属性

    start_time = time.time()
    clicked = False
    image_path = config.TEST_BOARD  # 固定图像参数

    logger.info(f"开始查找图像 {image_path} 并尝试点击")
    click_delay = 0.1  # 点击延时（秒）
    
    while time.time() - start_time < timeout:
        try:
            # 固定坐标作为测试用例
            center_x = 1439
            center_y = 70
            coordinate = (center_x, center_y)
            # 通过win32gui获取坐标对应的窗口句柄
            hwnd = win32gui.WindowFromPoint(coordinate)
            logger.debug(f"通过win32gui获取的窗口句柄: {hwnd}")
            if hwnd and isinstance(hwnd, int) and hwnd != 0:
                window_title = win32gui.GetWindowText(hwnd)
                logger.info(f"找到有效窗口句柄: {hwnd}，窗口标题: {window_title}")

                # 计算lParam参数：低16位存x，高16位存y
                lParam = (center_y << 16) | (center_x & 0xFFFF)
                WM_LBUTTONDOWN = 0x0201
                WM_LBUTTONUP = 0x0202

                # 发送左键按下和抬起消息
                logger.info(f"发送WM_LBUTTONDOWN消息，lParam: {lParam}")
                win32gui.PostMessage(hwnd, WM_LBUTTONDOWN, 0x00000001, lParam)
                time.sleep(click_delay)
                logger.info(f"发送WM_LBUTTONUP消息，lParam: {lParam}")
                win32gui.PostMessage(hwnd, WM_LBUTTONUP, 0, lParam)
                time.sleep(click_delay)
                logger.info("通过win32gui成功发送左键点击消息")
                clicked = True
                break
            else:
                logger.warning(f"无法通过win32gui获取有效窗口句柄，当前坐标: {coordinate}")
        except Exception as e:
            logger.error(f"查找或点击过程中发生异常: {e}")
        time.sleep(1)

    if not clicked:
        error_msg = f"超时: 在{timeout}秒内未能点击图像 {image_path}"
        logger.error(error_msg)
        raise Exception(error_msg)

def click_test_board():
    coordinate = (1438, 72)
    click_delay = 0.1  # 点击时延时0.1秒
    # 确保程序DPI感知设置正确
    ctypes.windll.user32.SetProcessDPIAware()

    # 方法6（win32api）：使用pywin32通过win32api模拟鼠标点击（仅限Windows）
    try:
        win32api.SetCursorPos(coordinate)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        time.sleep(click_delay)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(click_delay)
        messagebox.showinfo("通知", "已执行方法6：win32api模拟鼠标点击")
    except ImportError:
        logger.error("pywin32库未安装，请执行：pip install pywin32")
    except Exception as e:
        logger.error(f"使用win32api点击失败: {e}")
    
    # 方法7（ctypes）：使用ctypes调用系统API模拟鼠标点击（适用于Windows）
    try:
        MOUSEEVENTF_LEFTDOWN = 0x0002
        MOUSEEVENTF_LEFTUP = 0x0004
        ctypes.windll.user32.SetCursorPos(coordinate[0], coordinate[1])
        ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        time.sleep(click_delay)
        ctypes.windll.user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        time.sleep(click_delay)
        messagebox.showinfo("通知", "已执行方法7：ctypes模拟鼠标点击")
    except Exception as e:
        logger.error(f"使用ctypes模拟点击失败: {e}")
    
    # 方法9（win32gui PostMessage）：直接向窗口发送鼠标点击消息（日志更详细）
    try:
        import win32gui
        logger.info("开始执行方法9：通过win32gui发送鼠标点击消息")
        logger.debug(f"目标坐标: {coordinate}，点击延时: {click_delay}秒")
        hwnd = win32gui.WindowFromPoint(coordinate)
        logger.debug(f"调用win32gui.WindowFromPoint后获得的窗口句柄: {hwnd}")
        if hwnd:
            window_title = win32gui.GetWindowText(hwnd)
            logger.info(f"成功获取窗口句柄: {hwnd}，对应窗口标题: {window_title}")
            # 计算lParam：高位为y，低位为x
            lParam = (coordinate[1] << 16) | (coordinate[0] & 0xFFFF)
            logger.debug(f"计算得到 lParam = {lParam} (x: {coordinate[0]}, y: {coordinate[1]})")
            WM_LBUTTONDOWN = 0x0201
            WM_LBUTTONUP = 0x0202
            logger.info(f"发送 WM_LBUTTONDOWN 消息到窗口句柄 {hwnd}")
            win32gui.PostMessage(hwnd, WM_LBUTTONDOWN, 0x00000001, lParam)
            time.sleep(click_delay)
            logger.info(f"发送 WM_LBUTTONUP 消息到窗口句柄 {hwnd}")
            win32gui.PostMessage(hwnd, WM_LBUTTONUP, 0, lParam)
            time.sleep(click_delay)
            logger.info("通过win32gui成功发送鼠标点击消息")
            messagebox.showinfo("通知", "已执行方法9：win32gui发送消息点击")
        else:
            logger.error(f"无法获取指定坐标 {coordinate} 下的有效窗口句柄")
    except Exception as e:
        logger.error(f"使用win32gui发送消息点击失败，异常信息: {e}")

@utils.screenshot_error_to_excel()
def check_close_all_algs():
    utils.check_close_all_algs()

@utils.screenshot_error_to_excel()
def write_text_textbox():
    utils.write_text_textbox(config.RV_PASSWORD, write_content=config.RV_PASSWORD_TEXT)

@utils.screenshot_error_to_excel()
def delete_bad_mark():
    utils.delete_bad_mark()

timeout_seconds = 600

def parse_log_file(log_file_path):
    latest_time = None
    latest_alg_cal_time = None

    try:
        with open(log_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            # 从下往上遍历日志文件
            for i in range(len(lines) - 1, -1, -1):
                line = lines[i]
                match = re.search(r'AlgCalTime\(S\):([\d.]+)', line)
                if match:
                    alg_cal_time = float(match.group(1))
                    
                    # 向上查找最近的时间戳行
                    for j in range(i, -1, -1):
                        time_line = lines[j]
                        time_match = re.match(r'(\d{2}:\d{2}:\d{2}\.\d{3})', time_line)
                        if time_match:
                            time_str = time_match.group(1)
                            try:
                                # 获取当前日期
                                current_date = datetime.datetime.now().date()
                                # 解析时间并附加当前日期
                                current_time = datetime.datetime.strptime(time_str, '%H:%M:%S.%f').replace(year=current_date.year, month=current_date.month, day=current_date.day)
                                # 只取第一条匹配的记录
                                if latest_time is None:
                                    latest_time = current_time
                                    latest_alg_cal_time = alg_cal_time
                                    logger.debug(f"更新最新时间: {latest_time}, AlgCalTime: {latest_alg_cal_time}")
                                    return latest_time, latest_alg_cal_time
                            except ValueError as e:
                                logger.error(f"解析时间时发生错误: {e}")
    except FileNotFoundError:
        logger.error(f"日志文件未找到: {log_file_path}")
        return None, None

    if latest_time is None:
        logger.info("日志文件解析失败或无数据")
    else:
        logger.info(f"最新记录: 时间={latest_time.strftime('%Y-%m-%d %H:%M:%S')}, AlgCalTime={latest_alg_cal_time}")
    return latest_time, latest_alg_cal_time

def save_memory_data(memory_data, title, date_str):
    # 构建按日期区分的Excel文件路径
    excel_path = os.path.join(parent_dir, f'offline_test_results_{date_str}.xlsx')
    logger.info(f"开始保存内存数据，文档位置: {excel_path}, 标题: {title}")
    
    df = pd.DataFrame(memory_data, columns=['Time', 'Memory'])

    # 如果文件不存在则创建
    if not os.path.exists(excel_path):
        logger.info(f"文件 {excel_path} 不存在，正在创建新文件")
        wb = Workbook()
        save_workbook_with_retry(wb, excel_path)
    
    # 使用重试机制写入数据，解决文件被占用问题
    attempt = 0
    while attempt < 5:
        try:
            with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=title, index=False)
            logger.info(f"内存数据已保存到Excel，sheet名称: {title}")
            break
        except PermissionError as e:
            logger.error(f"保存内存数据时Excel文件被占用，尝试第 {attempt+1} 次等待1秒。错误：{e}")
            time.sleep(1)
            attempt += 1
    else:
        raise Exception(f"超时：无法保存内存数据到 {excel_path}")

def generate_memory_chart(title, date_str):
    # 构建按日期区分的Excel文件路径
    excel_path = os.path.join(parent_dir, f'offline_test_results_{date_str}.xlsx')
    logger.info(f"开始生成内存数据折线图，文档位置: {excel_path}, 标题: {title}")
    create_excel_chart(excel_path, title, 'Memory')
    logger.info(f"内存数据折线图已生成，文档位置: {excel_path}, 标题: {title}")

def save_alg_cal_time_plot(alg_cal_times_data, title, date_str):
    excel_path = os.path.join(parent_dir, f'offline_test_results_{date_str}.xlsx')
    logger.info(f"开始保存算法计算时间数据，文档位置: {excel_path}, 标题: {title}")
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        save_workbook_with_retry(wb, excel_path)

    wb = load_workbook_with_retry(excel_path)
    if title not in wb.sheetnames:
        ws = wb.create_sheet(title)
        logger.info(f"创建新的 sheet: {title}")
    else:
        ws = wb[title]

    existing_data = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_data.add(row)

    new_data = [entry for entry in alg_cal_times_data if entry not in existing_data]

    for time_entry, alg_cal_time in new_data:
        ws.append([time_entry, alg_cal_time])

    save_workbook_with_retry(wb, excel_path)
    logger.info(f"算法计算时间数据已保存到Excel，sheet名称: {title}")

    if len(new_data) < 2:
        logger.error("数据不足，无法生成折线图")
        return

    create_excel_chart(excel_path, title, 'AlgCalTime')
    logger.info(f"算法计算时间折线图已生成，文档位置: {excel_path}, 标题: {title}")

def create_excel_chart(excel_path, sheet_name, y_axis_label):
    logger.info(f"开始创建Excel折线图，文档位置: {excel_path}, sheet名称: {sheet_name}, Y轴标签: {y_axis_label}")
    
    try:
        wb = load_workbook_with_retry(excel_path)
        logger.debug(f"成功加载Excel文件: {excel_path}")
    except Exception as e:
        logger.error(f"加载Excel文件失败: {e}")
        return

    if sheet_name not in wb.sheetnames:
        logger.warning(f"工作表 {sheet_name} 不存在，已创建新的空sheet用于生成图表")
        ws = wb.create_sheet(sheet_name)
        ws.append(["Time", y_axis_label])
    else:
        ws = wb[sheet_name]
        # 检查是否缺少表头，如果第一行不是"Time"，则插入表头
        if ws.max_row == 0 or ws.cell(row=1, column=1).value != "Time":
            ws.insert_rows(1)
            ws["A1"].value = "Time"
            ws["B1"].value = y_axis_label
            logger.info(f"在已有 sheet {sheet_name} 中添加了表头")
    logger.debug(f"成功获取工作表: {sheet_name}")

    for drawing in list(ws._charts):
        if drawing.title == f"{sheet_name} 折线图":
            ws._charts.remove(drawing)
            logger.info(f"已删除现有的折线图: {drawing.title}")

    chart = LineChart()
    chart.title = f"{sheet_name} 折线图"
    chart.style = 13
    chart.y_axis.title = y_axis_label
    chart.x_axis.title = 'Time'
    chart.width = 20
    chart.height = 10

    max_row = ws.max_row
    logger.debug(f"工作表最大行数: {max_row}")
    if max_row < 2:
        logger.error("数据不足，无法生成折线图")
        return

    data_column_empty = all(ws.cell(row=i, column=2).value is None for i in range(2, max_row + 1))
    logger.debug(f"数据列是否为空: {data_column_empty}")
    if data_column_empty:
        logger.error("数据列为空，无法生成折线图")
        return

    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    logger.debug("成功设置数据和类别引用")

    for series in chart.series:
        series.graphicalProperties.line.width = 20000
    logger.debug("成功设置折线图的线条宽度")

    chart.x_axis.majorUnit = 60
    logger.debug("成功设置X轴的主要单位为每分钟")

    try:
        ws.add_chart(chart, "C2")
        logger.info(f"折线图已添加到工作表，位置: C2")
    except Exception as e:
        logger.error(f"添加折线图到工作表失败: {e}")
        return

    try:
        save_workbook_with_retry(wb, excel_path)
        logger.info(f"Excel文档已保存，文档位置: {excel_path}")
    except Exception as e:
        logger.error(f"保存Excel文档失败: {e}")

def get_process_memory_usage(process_name):
    """获取指定进程的内存使用率"""
    try:
        for proc in psutil.process_iter(['name', 'memory_info']):
            if proc.info['name'] == process_name:
                return proc.info['memory_info'].rss / (1024 * 1024)
    except Exception as e:
        logger.error(f"获取进程内存使用率时发生错误: {e}")
    return 0

def test_current_window():
    if not utils.search_symbol(config.TEST_WINDOW, tolerance=0.7):
        messagebox.showwarning("警告", "请先选中窗口，并确保测试当前窗口的按钮存在")
        return

    mem_data_aoi = []
    mem_data_slave = []
    start_time = time.time()
    current_day = datetime.datetime.now().strftime('%Y-%m-%d')
    try:
        while True:
            if utils.search_symbol(config.TEST_WINDOW, timeout=1):
                utils.click_by_png(config.TEST_WINDOW, tolerance=0.7, timeout=1)
            else:
                logger.info("未找到测试窗口按钮")
                break
            time.sleep(1)
            now = datetime.datetime.now()
            timestamp = now.strftime('%Y-%m-%d %H:%M:%S')
            usage_aoi = get_process_memory_usage("AOI.exe")
            usage_slave = get_process_memory_usage("AOI_SLAVE.exe")
            mem_data_aoi.append((timestamp, usage_aoi))
            if usage_slave > 0:
                mem_data_slave.append((timestamp, usage_slave))
            current_time = time.time()
            if (current_time - start_time) >= 60:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Window Memory AOI", current_day)
                generate_memory_chart("Test Window Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Window Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Window Memory AOI_SLAVE", current_day)
                start_time = current_time
            # 检测日期变化，跨天则保存并清空数据
            new_day = datetime.datetime.now().strftime('%Y-%m-%d')
            if new_day != current_day:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Window Memory AOI", current_day)
                generate_memory_chart("Test Window Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Window Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Window Memory AOI_SLAVE", current_day)
                mem_data_aoi.clear()
                mem_data_slave.clear()
                current_day = new_day
    except KeyboardInterrupt:
        logger.info("用户中断，保存数据。")
    except Exception as e:
        logger.error(f"发生异常: {e}")
    finally:
        current_day = datetime.datetime.now().strftime('%Y-%m-%d')
        excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
        save_memory_data(mem_data_aoi, "Test Window Memory AOI", current_day)
        generate_memory_chart("Test Window Memory AOI", current_day)
        if mem_data_slave:
            save_memory_data(mem_data_slave, "Test Window Memory AOI_SLAVE", current_day)
            generate_memory_chart("Test Window Memory AOI_SLAVE", current_day)
        generate_charts_for_all_dates()

def test_current_group():
    if not utils.search_symbol(config.TEST_GROUP, tolerance=0.7):
        messagebox.showwarning("警告", "请先选中窗口，并确保测试当前分组的按钮存在")
        return

    mem_data_aoi = []
    mem_data_slave = []
    current_day = datetime.datetime.now().strftime('%Y-%m-%d')

    def record_memory_usage():
        while True:
            ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            usage_aoi = get_process_memory_usage("AOI.exe")
            usage_slave = get_process_memory_usage("AOI_SLAVE.exe")
            mem_data_aoi.append((ts, usage_aoi))
            if usage_slave > 0:
                mem_data_slave.append((ts, usage_slave))
            time.sleep(1)

    def save_memory_periodically():
        nonlocal current_day
        last_save_time = time.time()
        while True:
            time.sleep(1)
            now_time = time.time()
            if (now_time - last_save_time) >= 180:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Group Memory AOI", current_day)
                generate_memory_chart("Test Group Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Group Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Group Memory AOI_SLAVE", current_day)
                last_save_time = now_time
            new_day = datetime.datetime.now().strftime('%Y-%m-%d')
            if new_day != current_day:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Group Memory AOI", current_day)
                generate_memory_chart("Test Group Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Group Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Group Memory AOI_SLAVE", current_day)
                mem_data_aoi.clear()
                mem_data_slave.clear()
                current_day = new_day

    threading.Thread(target=record_memory_usage, daemon=True).start()
    threading.Thread(target=save_memory_periodically, daemon=True).start()

    try:
        while True:
            utils.click_by_png(config.TEST_GROUP, tolerance=0.7, timeout=1)
            while utils.search_symbol(config.QUESTION_MARK, tolerance=0.7, timeout=1):
                pyautogui.press("enter")
            while utils.search_symbol(config.TESTING_COMPONENT, tolerance=0.7, timeout=1):
                time.sleep(1)
    except KeyboardInterrupt:
        logger.info("用户中断，保存数据。")
    except Exception as e:
        logger.error(f"发生异常: {e}")
    finally:
        for day in [current_day, datetime.datetime.now().strftime('%Y-%m-%d')]:
            excel_path = os.path.join(parent_dir, f'offline_test_results_{day}.xlsx')
            save_memory_data(mem_data_aoi, "Test Group Memory AOI", day)
            generate_memory_chart("Test Group Memory AOI", day)
            if mem_data_slave:
                save_memory_data(mem_data_slave, "Test Group Memory AOI_SLAVE", day)
                generate_memory_chart("Test Group Memory AOI_SLAVE", day)
        generate_charts_for_all_dates()

def test_current_component():
    if not utils.search_symbol(config.TEST_COMPONENT, tolerance=0.7):
        messagebox.showwarning("警告", "请先选中窗口，并确保测试当前元件的按钮存在")
        return

    mem_data_aoi = []
    mem_data_slave = []
    current_day = datetime.datetime.now().strftime('%Y-%m-%d')

    def record_memory_usage():
        while True:
            ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            usage_aoi = get_process_memory_usage("AOI.exe")
            usage_slave = get_process_memory_usage("AOI_SLAVE.exe")
            mem_data_aoi.append((ts, usage_aoi))
            if usage_slave > 0:
                mem_data_slave.append((ts, usage_slave))
            time.sleep(1)

    def save_memory_periodically():
        nonlocal current_day
        last_save_time = time.time()
        while True:
            time.sleep(30)
            now_time = time.time()
            if (now_time - last_save_time) >= 30:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Component Memory AOI", current_day)
                generate_memory_chart("Test Component Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Component Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Component Memory AOI_SLAVE", current_day)
                last_save_time = now_time
            new_day = datetime.datetime.now().strftime('%Y-%m-%d')
            if new_day != current_day:
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                save_memory_data(mem_data_aoi, "Test Component Memory AOI", current_day)
                generate_memory_chart("Test Component Memory AOI", current_day)
                if mem_data_slave:
                    save_memory_data(mem_data_slave, "Test Component Memory AOI_SLAVE", current_day)
                    generate_memory_chart("Test Component Memory AOI_SLAVE", current_day)
                mem_data_aoi.clear()
                mem_data_slave.clear()
                current_day = new_day

    threading.Thread(target=record_memory_usage, daemon=True).start()
    threading.Thread(target=save_memory_periodically, daemon=True).start()

    try:
        while True:
            utils.click_by_png(config.TEST_COMPONENT, tolerance=0.7, timeout=1)
            while utils.search_symbol(config.TESTING_COMPONENT, tolerance=0.7, timeout=3):
                time.sleep(1)
    except KeyboardInterrupt:
        logger.info("用户中断，保存数据。")
    except Exception as e:
        logger.error(f"发生异常: {e}")
    finally:
        current_day = datetime.datetime.now().strftime('%Y-%m-%d')
        excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
        save_memory_data(mem_data_aoi, "Test Component Memory AOI", current_day)
        generate_memory_chart("Test Component Memory AOI", current_day)
        if mem_data_slave:
            save_memory_data(mem_data_slave, "Test Component Memory AOI_SLAVE", current_day)
            generate_memory_chart("Test Component Memory AOI_SLAVE", current_day)
        generate_charts_for_all_dates()

def test_current_board():
    # utils.bring_aoi_to_top()
    if not utils.search_symbol(config.TEST_BOARD, timeout=3, tolerance=0.7):
        messagebox.showwarning("警告", "请先选中窗口，并确保测试当前整板的按钮存在")
        return

    prev_alg_record_time = None  # 用于记录上次的算法计算时间数据
    current_day = datetime.datetime.now().strftime('%Y-%m-%d')
    excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
    if not os.path.exists(excel_path):
        wb = Workbook()
        save_workbook_with_retry(wb, excel_path)

    # 内存记录线程：同时记录 AOI.exe 和 AOI_SLAVE.exe 的内存数据
    def memory_record_thread():
        while True:
            mem_usage_aoi = get_process_memory_usage("AOI.exe")
            mem_usage_slave = get_process_memory_usage("AOI_SLAVE.exe")
            record_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            # 将两个进程的数据一起存入队列
            memory_queue.put((record_time, mem_usage_aoi, mem_usage_slave))
            time.sleep(1)

    # 数据保存和图表生成线程：按记录时间分组，不同日期的数据存入不同的Excel文件中
    def memory_save_and_chart_thread():
        while True:
            time.sleep(180)  # 每3分钟执行一次保存和图表生成
            records = []
            while not memory_queue.empty():
                records.append(memory_queue.get())
            if records:
                groups_aoi = {}
                groups_slave = {}
                for rec in records:
                    rec_day = rec[0][:10]
                    groups_aoi.setdefault(rec_day, []).append((rec[0], rec[1]))
                    groups_slave.setdefault(rec_day, []).append((rec[0], rec[2]))
                for rec_day, data in groups_aoi.items():
                    excel_path_day = os.path.join(parent_dir, f'offline_test_results_{rec_day}.xlsx')
                    save_memory_data(data, "Test Board Memory AOI", rec_day)
                    generate_memory_chart("Test Board Memory AOI", rec_day)
                for rec_day, data in groups_slave.items():
                    excel_path_day = os.path.join(parent_dir, f'offline_test_results_{rec_day}.xlsx')
                    save_memory_data(data, "Test Board Memory AOI_SLAVE", rec_day)
                    generate_memory_chart("Test Board Memory AOI_SLAVE", rec_day)

    # 算法计算时间监控线程保持不变
    def alg_time_monitor_thread():
        nonlocal prev_alg_record_time, current_day
        last_chart_update = time.time()
        sheet_name = "Test Board AlgCalTime"
        while True:
            log_file_path = r'D:\EYAOI\Logger\Lane_0\AlgSimpCalTime_' + datetime.datetime.now().strftime('%Y-%m-%d') + '.log'
            try:
                log_time, alg_cal_time = parse_log_file(log_file_path)
                if log_time is not None and alg_cal_time is not None:
                    current_day = datetime.datetime.now().strftime('%Y-%m-%d')
                    excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                    try:
                        wb = load_workbook_with_retry(excel_path)
                    except Exception as e:
                        logger.error(f"加载Excel失败: {e}")
                        wb = Workbook()
                        save_workbook_with_retry(wb, excel_path)
                    
                    # 如果sheet不存在，则创建并添加表头；如果存在但缺少表头，则插入表头
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(sheet_name)
                        ws.append(["Time", "AlgCalTime"])
                        logger.info(f"创建新的 sheet: {sheet_name} 并添加表头")
                    else:
                        ws = wb[sheet_name]
                        if ws.max_row == 0 or ws.cell(row=1, column=1).value != "Time":
                            ws.insert_rows(1)
                            ws["A1"].value = "Time"
                            ws["B1"].value = "AlgCalTime"
                            logger.info(f"在已有 sheet: {sheet_name} 中添加表头")
                    
                    # 判断数据是否重复以及插入新数据
                    last_row = ws.max_row
                    if last_row > 1:
                        last_time_in_sheet = ws.cell(row=last_row, column=1).value
                        last_alg_cal_in_sheet = ws.cell(row=last_row, column=2).value
                        if last_time_in_sheet == log_time and last_alg_cal_in_sheet == alg_cal_time:
                            logger.info("算法计算时间数据重复，跳过插入")
                        else:
                            ws.append([log_time, alg_cal_time])
                            logger.info(f"插入算法计算时间数据: 时间={log_time}, AlgCalTime={alg_cal_time}")
                    else:
                        ws.append([log_time, alg_cal_time])
                        logger.info(f"插入第一条算法计算时间数据: 时间={log_time}, AlgCalTime={alg_cal_time}")
                    
                    save_workbook_with_retry(wb, excel_path)
                    prev_alg_record_time = log_time
            except Exception as e:
                logger.error(f"解析日志文件时发生错误: {e}")
            
            if time.time() - last_chart_update >= 180:
                current_day = datetime.datetime.now().strftime('%Y-%m-%d')
                excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
                logger.info(f"AlgCalTime monitor: 重新生成折线图，文档位置: {excel_path}")
                create_excel_chart(excel_path, sheet_name, 'AlgCalTime')
                last_chart_update = time.time()
            
            time.sleep(1)

    threading.Thread(target=memory_record_thread, daemon=True).start()
    threading.Thread(target=memory_save_and_chart_thread, daemon=True).start()
    threading.Thread(target=alg_time_monitor_thread, daemon=True).start()

    try:
        while True:
            logger.info("开始点击")
            utils.click_by_png(config.TEST_BOARD, 2, timeout=3, tolerance=0.7)
            logger.info("点击完成")
            while utils.search_symbol(config.QUESTION_MARK, tolerance=0.7, timeout=1):
                utils.click_by_png(config.QUESTION_MARK, tolerance=0.7, timeout=1)
                pyautogui.press("enter")
            while utils.search_symbol(config.TESTING_COMPONENT, tolerance=0.7, timeout=3):
                time.sleep(1)
    except KeyboardInterrupt:
        logger.info("用户中断，保存数据。")
    except Exception as e:
        logger.error(f"发生异常: {e}")
    finally:
        current_day = datetime.datetime.now().strftime('%Y-%m-%d')
        excel_path = os.path.join(parent_dir, f'offline_test_results_{current_day}.xlsx')
        logger.info(f"最终保存数据到: {excel_path}")
        remaining_records = []
        while not memory_queue.empty():
            remaining_records.append(memory_queue.get())
        if remaining_records:
            groups_aoi = {}
            groups_slave = {}
            for rec in remaining_records:
                rec_day = rec[0][:10]
                groups_aoi.setdefault(rec_day, []).append((rec[0], rec[1]))
                groups_slave.setdefault(rec_day, []).append((rec[0], rec[2]))
            for rec_day, data in groups_aoi.items():
                excel_path_day = os.path.join(parent_dir, f'offline_test_results_{rec_day}.xlsx')
                save_memory_data(data, "Test Board Memory AOI", rec_day)
                generate_memory_chart("Test Board Memory AOI", rec_day)
            for rec_day, data in groups_slave.items():
                excel_path_day = os.path.join(parent_dir, f'offline_test_results_{rec_day}.xlsx')
                save_memory_data(data, "Test Board Memory AOI_SLAVE", rec_day)
                generate_memory_chart("Test Board Memory AOI_SLAVE", rec_day)
        create_excel_chart(excel_path, "Test Board AlgCalTime", 'AlgCalTime')
        generate_charts_for_all_dates()

def generate_charts_for_all_dates():
    logger.info("开始为所有日期生成折线图")
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    # 定义所有测试相关的关键字
    test_keywords = ["Test Board", "Test Group", "Test Component", "Test Window", "AOI", "AOI_SLAVE"]
    for file_name in os.listdir(parent_dir):
        if file_name.startswith('offline_test_results_') and file_name.endswith('.xlsx'):
            file_date = file_name[len('offline_test_results_'):-len('.xlsx')]
            if initial_date <= file_date <= current_date:
                excel_path = os.path.join(parent_dir, file_name)
                logger.info(f"处理文件: {excel_path}")
                try:
                    wb = load_workbook_with_retry(excel_path)
                    for sheet_name in wb.sheetnames:
                        # 检查sheet名称是否包含测试相关的关键字
                        if any(keyword in sheet_name for keyword in test_keywords):
                            if "AlgCalTime" in sheet_name:
                                create_excel_chart(excel_path, sheet_name, 'AlgCalTime')
                            elif "Memory" in sheet_name:
                                create_excel_chart(excel_path, sheet_name, 'Memory')
                            else:
                                logger.warning(f"未能确定图表类型的sheet: {sheet_name}")
                    save_workbook_with_retry(wb, excel_path)
                    logger.info(f"折线图已生成并保存，文档位置: {excel_path}")
                except Exception as e:
                    logger.error(f"处理文件 {excel_path} 时发生错误: {e}")

def bring_aoi_to_top():
    utils.bring_aoi_to_top()
